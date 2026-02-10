# Classe utilitária para análise geométrica das pilares
class AnalisadorGeometricoPilares:
    """Analisa e identifica tipos de barras para PILARES - EXPANDIDO"""

    @staticmethod
    def identificar_tipo_pilar(pos, bitola, comp):
        """
        Identificação baseada em padrões conhecidos de pilares
        Agora com TODOS os tipos de dobras possíveis
        """
        pos_upper = str(pos).upper()

        # REGRAS EXPANDIDAS PARA PILARES
        # N1 com comprimento < 1.5m = ESTRIBO
        if pos_upper == 'N1' and comp < 1.5:
            return "ESTRIBO"

        # N2, N3 com bitola pequena = ESTRIBO
        elif pos_upper in ['N2', 'N3'] and bitola <= 6.3:
            return "ESTRIBO"

        # N4, N5 = PORTA ESTRIBO (barras superiores)
        elif pos_upper in ['N4', 'N5']:
            return "PORTA_ESTRIBO"

        # N6, N7 = BARRAS NEGATIVAS (apoios)
        elif pos_upper in ['N6', 'N7']:
            return "NEGATIVO"

        # N8, N9 = BARRAS POSITIVAS (vão)
        elif pos_upper in ['N8', 'N9']:
            return "POSITIVO"

        # N10, N11 = DOBRA DUAS PONTAS
        elif pos_upper in ['N10', 'N11']:
            return "DOBRA_DUAS_PONTAS"

        # N12, N13 = DOBRA UMA PONTA
        elif pos_upper in ['N12', 'N13']:
            return "DOBRA_UMA_PONTA"

        # N14, N15 = BARRA U
        elif pos_upper in ['N14', 'N15']:
            return "BARRA_U"

        # N16, N17 = BARRA Z
        elif pos_upper in ['N16', 'N17']:
            return "BARRA_Z"

        # N18, N19 = CAVALETE
        elif pos_upper in ['N18', 'N19']:
            return "CAVALETE"

        # N20, N21 = GRAMPO
        elif pos_upper in ['N20', 'N21']:
            return "GRAMPO"

        # Bitola pequena (5.0 ou 6.3) = ESTRIBO
        elif bitola <= 6.3:
            return "ESTRIBO"

        # Comprimento grande = POSITIVO
        elif comp > 5.0:
            return "POSITIVO"

        # Comprimento médio com bitola média = NEGATIVO
        elif 2.0 < comp <= 5.0 and bitola >= 10.0:
            return "NEGATIVO"

        # Padrão = NEGATIVO
        else:
            return "NEGATIVO"

    @staticmethod
    def calcular_medidas_estribo_pilar(comprimento):
        """
        Calcula as medidas do estribo para pilares
        """
        perimetro_cm = int(comprimento * 100)

        # Estribos típicos de pilares
        if perimetro_cm <= 100:
            return (12, 40)  # Estribo 12x40
        elif perimetro_cm <= 120:
            return (12, 50)  # Estribo 12x50
        elif perimetro_cm <= 140:
            return (15, 55)  # Estribo 15x55
        elif perimetro_cm <= 160:
            return (15, 65)  # Estribo 15x65
        elif perimetro_cm <= 180:
            return (20, 70)  # Estribo 20x70
        else:
            return (20, 80)  # Estribo 20x80

    @staticmethod
    def calcular_medidas_negativo(comprimento):
        """
        Calcula as medidas da barra negativa (com dobras nos apoios)
        """
        comp_cm = int(comprimento * 100)

        # Típico: 20% dobra esquerda, 60% vão, 20% dobra direita
        dobra_esq = int(comp_cm * 0.20)
        vao = int(comp_cm * 0.60)
        dobra_dir = int(comp_cm * 0.20)

        return (dobra_esq, vao, dobra_dir)

    @staticmethod
    def calcular_medidas_positivo(comprimento):
        """
        Calcula as medidas da barra positiva (com ganchos)
        """
        comp_cm = int(comprimento * 100)

        # Descontar ganchos (15cm cada)
        gancho = 15
        vao_livre = comp_cm - (2 * gancho)

        return (gancho, vao_livre, gancho)

    @staticmethod
    def calcular_medidas_dobra_duas_pontas(comprimento):
        """
        Calcula as medidas da barra com dobra nas duas pontas
        """
        comp_cm = int(comprimento * 100)

        # Típico: 15% dobra esquerda, 70% central, 15% dobra direita
        dobra_esq = int(comp_cm * 0.15)
        trecho_central = int(comp_cm * 0.70)
        dobra_dir = int(comp_cm * 0.15)

        return (dobra_esq, trecho_central, dobra_dir)

    @staticmethod
    def calcular_medidas_dobra_uma_ponta(comprimento):
        """
        Calcula as medidas da barra com dobra em uma ponta
        """
        comp_cm = int(comprimento * 100)

        # Típico: 80% reto, 20% dobra
        trecho_reto = int(comp_cm * 0.80)
        dobra = int(comp_cm * 0.20)

        return (trecho_reto, dobra)

    @staticmethod
    def calcular_medidas_barra_u(comprimento):
        """
        Calcula as medidas da barra em U
        """
        comp_cm = int(comprimento * 100)

        # Proporção: 30% lateral esq, 40% base, 30% lateral dir
        lateral_esq = int(comp_cm * 0.30)
        base = int(comp_cm * 0.40)
        lateral_dir = int(comp_cm * 0.30)

        return (lateral_esq, base, lateral_dir)

    @staticmethod
    def calcular_medidas_barra_z(comprimento):
        """
        Calcula as medidas da barra em Z
        """
        comp_cm = int(comprimento * 100)

        # Proporção: 3 segmentos iguais
        segmento = int(comp_cm / 3)

        return (segmento, segmento, segmento)

    @staticmethod
    def calcular_medidas_tres_segmentos(comprimento):
        """
        Divide o comprimento em 3 segmentos iguais (em cm)
        """
        comp_cm = int(comprimento * 100)
        segmento = int(comp_cm / 3)
        return (segmento, segmento, segmento)

    @staticmethod
    def calcular_medidas_cavalete(comprimento):
        """
        Calcula as medidas do cavalete
        """
        comp_cm = int(comprimento * 100)

        # Base 50%, duas pernas 25% cada
        base = int(comp_cm * 0.50)
        perna = int(comp_cm * 0.25)

        return (perna, base, perna)

    @staticmethod
    def calcular_medidas_grampo(comprimento):
        """
        Calcula as medidas do grampo
        """
        comp_cm = int(comprimento * 100)

        # Abertura 60%, dobras 20% cada
        abertura = int(comp_cm * 0.60)
        dobra = int(comp_cm * 0.20)

        return (dobra, abertura, dobra)
# pilares_app.py - SISTEMA COMPLETO COM ROMANEIO MELHORADO E CHECK LIST
import os
import math
import shutil
import subprocess
import tempfile
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
from datetime import datetime
from PIL import Image, ImageTk

# Flags globais para verificar disponibilidade de módulos
ETIQUETAS_GERADOR_DISPONIVEL = False
ETIQUETAS_HELPER_DISPONIVEL = False
ETIQUETAS_LAYOUT_CFG = False
GeradorEtiquetasDinamico = None
processar_pilares = None

try:
    from core.pilares_motor_dual import processar_pilares
    print("[MOTOR] Usando pilares_motor_dual.py (Dual: Rapido + Completo)")
except Exception as e:
    print(f"[MOTOR] Motor dual indisponivel: {e}")
    try:
        from core.pilares_motor import processar_pilares
        print("[MOTOR] Usando pilares_motor.py (fallback)")
    except Exception:
        print("[MOTOR] Nenhum motor disponivel!")
        pass
try:
    from core.etiquetas_generator import GeradorEtiquetasDinamico
    ETIQUETAS_GERADOR_DISPONIVEL = True
except Exception:
    pass

try:
    from core.etiquetas_helper import (
        gerar_codigo_identificador,
        gerar_codigo_barras_imagem,
        localizar_desenho_barra,
        carregar_desenho_redimensionado,
        formatar_os_numero,
    )
    ETIQUETAS_HELPER_DISPONIVEL = True
except Exception:
    pass

try:
    from core.etiquetas_layout_config import (
        PX_MM as CFG_PX_MM,
        MARGEM_EXTERNA_MM,
        TOPO_ALTURA_MM,
        SECAO_MICRO_ALTURA_MM,
        ESPACO_PICOTE_MM,
        LARGURA_ETIQUETA_MM,
        OS_BLOCO_LARGURA_MM,
        OS_BLOCO_ALTURA_MM,
        FAIXA_VERTICAL_LARGURA_MM,
        TABELA_ALTURA_HEADER_MM,
        TABELA_ALTURA_LINHA_MM,
        COL_BITOLA_MM,
        COL_COMPR_UNIT_MM,
        COL_PESO_MM,
        COL_QTDE_MM,
        DESENHO_X_OFFSET_MM,
        DESENHO_Y_OFFSET_MM,
        DESENHO_LARGURA_MM,
        DESENHO_ALTURA_MM,
        FONT_PEQ,
        FONT_MED,
        FONT_GRD,
    )
    ETIQUETAS_LAYOUT_CFG = True
except Exception:
    pass

def converter_lote_oda(input_files):
    """Converte uma lista de arquivos (DWG/DXF) para DXF ACAD2018 usando ODAFileConverter.
    Retorna lista de caminhos dos arquivos convertidos em ODA/out.
    """
    try:
        base = Path(__file__).resolve().parent
        oda_dir = base / 'ODA'
        oda_exe = oda_dir / 'ODAFileConverter.exe'
        in_dir = oda_dir / 'in'
        out_dir = oda_dir / 'out'

        if not oda_exe.exists():
            raise FileNotFoundError(f"ODAFileConverter.exe não encontrado em: {oda_exe}")

        # Preparar pastas
        in_dir.mkdir(parents=True, exist_ok=True)
        if out_dir.exists():
            shutil.rmtree(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        converted = []
        for f in input_files:
            fin = Path(f)
            if not fin.exists():
                continue

            # Copiar para ODA/in para simplificar filtro
            dst = in_dir / fin.name
            try:
                if dst.exists():
                    dst.unlink()
                shutil.copy2(str(fin), str(dst))
            except Exception:
                # Se cópia falhar, tentar converter direto da pasta original
                dst = fin

            # Converter para DXF ACAD2018
            cmd = [
                str(oda_exe),
                str(dst.parent),  # inputDir
                str(out_dir),     # outputDir
                'ACAD2018',       # output_version
                'DXF',            # output_type
                '0',              # recurse
                '1',              # audit
                dst.name          # filter (somente este arquivo)
            ]
            try:
                subprocess.check_call(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                # Arquivo convertido terá extensão .dxf
                out_name = Path(dst.name).with_suffix('.dxf').name
                fout = out_dir / out_name
                if fout.exists():
                    converted.append(str(fout))
            except Exception:
                # Ignorar falhas individuais
                continue

        return converted
    except Exception:
        return []

class PilaresApp(tk.Tk):
    def _marcar_todas_etiquetas(self):
        """Marca todas as etiquetas para impressão"""
        # Marcar todas as etiquetas (função específica)
        if not hasattr(self, 'etiquetas_selecionadas') or len(self.etiquetas_selecionadas) != len(self.dados_processados):
            self.etiquetas_selecionadas = {i: True for i in range(len(self.dados_processados))}
        else:
            for i in range(len(self.dados_processados)):
                self.etiquetas_selecionadas[i] = True
        self.desenhar_etiquetas_com_selecao()
        if hasattr(self, 'label_selecionadas'):
            total = sum(1 for v in self.etiquetas_selecionadas.values() if v)
            self.label_selecionadas.config(text=f"Selecionadas: {total}/{len(self.dados_processados)}")

    def _desmarcar_todas_etiquetas(self):
        """Desmarca todas as etiquetas"""
        # Desmarcar todas as etiquetas (função específica)
        if not hasattr(self, 'etiquetas_selecionadas') or len(self.etiquetas_selecionadas) != len(self.dados_processados):
            self.etiquetas_selecionadas = {i: False for i in range(len(self.dados_processados))}
        else:
            for i in range(len(self.dados_processados)):
                self.etiquetas_selecionadas[i] = False
        self.desenhar_etiquetas_com_selecao()
        if hasattr(self, 'label_selecionadas'):
            total = sum(1 for v in self.etiquetas_selecionadas.values() if v)
            self.label_selecionadas.config(text=f"Selecionadas: {total}/{len(self.dados_processados)}")
    def __init__(self):
        super().__init__()

        self.title("EngenhariaPlanPro - PILARES")
        self.configure(bg="#0d2818")
        
        # Abrir em tela cheia (maximizado)
        self.state('zoomed')

        self.arquivos_selecionados = []
        self.dados_processados = []
        self.total_kg = 0.0
        self.total_barras = 0
        self.tipos_personalizados = {}  # Dicionário para armazenar tipos personalizados
        self.desenhos_customizados = {}  # Para compatibilidade com código antigo
        self.medidas_customizadas = {}  # {(pilar, pos): {bitola, comp, qtde}}
        self.formas_customizadas = {}  # {(pilar, pos): 'reta'|'gancho'|'estribo'}
        self.estribo_lados = {}  # {(pilar, pos): (lado_a, lado_b, lado_c, lado_d)}
        self.etiquetas_por_pagina = 6
        self.pagina_atual = 0

        self._criar_interface()

    def _criar_interface(self):
        # Container principal
        main_container = tk.Frame(self, bg="#0d2818")
        main_container.pack(fill="both", expand=True, padx=10, pady=10)

        # Título
        header = tk.Frame(main_container, bg="#0d2818")
        header.pack(fill="x", pady=(0, 10))

        tk.Label(
            header,
            text="🏗️ PROCESSAMENTO DE PILARES",
            font=("Arial", 20, "bold"),
            bg="#0d2818",
            fg="#ff9800"
        ).pack()

        # Frame de controles
        control_frame = tk.Frame(main_container, bg="#1a3d2e")
        control_frame.pack(fill="x", pady=5)

        # Linha 1 - Dados do projeto
        row1 = tk.Frame(control_frame, bg="#1a3d2e")
        row1.pack(fill="x", padx=10, pady=10)

        tk.Label(row1, text="Obra:", bg="#1a3d2e", fg="white", font=("Arial", 10)).pack(side="left", padx=5)
        self.var_obra = tk.StringVar(value="OBRA 001")
        tk.Entry(row1, textvariable=self.var_obra, width=25, font=("Arial", 10)).pack(side="left", padx=5)

        tk.Label(row1, text="Pavimento:", bg="#1a3d2e", fg="white", font=("Arial", 10)).pack(side="left", padx=20)
        self.var_pavimento = tk.StringVar(value="TÉRREO")
        tk.Entry(row1, textvariable=self.var_pavimento, width=20, font=("Arial", 10)).pack(side="left", padx=5)

        # Botão Limpar na mesma linha para fácil acesso
        tk.Button(
            row1,
            text="🔄 Limpar",
            command=self.limpar,
            bg="#e74c3c",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=3,
            cursor="hand2"
        ).pack(side="right", padx=10)

        # Linha 2 - Botões principais
        row2 = tk.Frame(control_frame, bg="#1a3d2e")
        row2.pack(fill="x", padx=10, pady=(0, 10))

        tk.Button(
            row2,
            text="📁 Selecionar Arquivos",
            command=self.selecionar_arquivos,
            bg="#ff6f00",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)

        tk.Button(
            row2,
            text="⚙️ PROCESSAR (Rápido)",
            command=self.processar,
            bg="#ff8f00",
            fg="white",
            font=("Arial", 11, "bold"),
            padx=20,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)

        tk.Button(
            row2,
            text="⚙️ PROCESSAR 2.0 (Completo)",
            command=self.processar_v2,
            bg="#ff3d00",
            fg="white",
            font=("Arial", 11, "bold"),
            padx=20,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)

        tk.Button(
            row2,
            text="📄 Gerar Romaneio",
            command=self.gerar_romaneio,
            bg="#1976d2",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)

        tk.Button(
            row2,
            text="📋 Check List",
            command=self.gerar_romaneio_conferencia,
            bg="#00acc1",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)

        tk.Button(
            row2,
            text="🏷️ Etiquetas",
            command=self.imprimir_etiquetas,
            bg="#9c27b0",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)

        tk.Button(
            row2,
            text="📊 Exportar Excel",
            command=self.exportar_excel,
            bg="#4caf50",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)

        tk.Button(
            row2,
            text="🖨️ Imprimir",
            command=self.imprimir_direto,
            bg="#6a1b9a",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)

        # Frame da tabela
        table_frame = tk.Frame(main_container, bg="#0d2818")
        table_frame.pack(fill="both", expand=True, pady=10)

        # Criar Treeview com scrollbars
        tree_scroll_y = ttk.Scrollbar(table_frame)
        tree_scroll_y.pack(side="right", fill="y")

        tree_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal")
        tree_scroll_x.pack(side="bottom", fill="x")

        # Configurar estilo
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="#1e1e1e", foreground="white", fieldbackground="#1e1e1e")
        style.configure("Treeview.Heading", background="#ff6f00", foreground="white", font=("Arial", 10, "bold"))

        # Criar Treeview
        self.tree = ttk.Treeview(
            table_frame,
            columns=("pilar", "pos", "bitola", "qtde", "comp", "peso"),
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set
        )

        # Configurar colunas
        self.tree.heading("pilar", text="PILAR")
        self.tree.heading("pos", text="POSIÇÃO")
        self.tree.heading("bitola", text="BITOLA (mm)")
        self.tree.heading("qtde", text="QUANTIDADE")
        self.tree.heading("comp", text="COMP. (m)")
        self.tree.heading("peso", text="PESO (kg)")

        self.tree.column("pilar", width=150, anchor="center")
        self.tree.column("pos", width=100, anchor="center")
        self.tree.column("bitola", width=100, anchor="center")
        self.tree.column("qtde", width=100, anchor="center")
        self.tree.column("comp", width=150, anchor="center")
        self.tree.column("peso", width=120, anchor="center")

        self.tree.pack(fill="both", expand=True)

        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)

        # Status bar
        self.status_frame = tk.Frame(main_container, bg="#1a3d2e", relief="sunken", bd=1)
        self.status_frame.pack(fill="x", pady=(5, 0))

        self.status_label = tk.Label(
            self.status_frame,
            text="✅ Pronto para processar",
            bg="#1a3d2e",
            fg="white",
            font=("Arial", 10),
            anchor="w"
        )
        self.status_label.pack(side="left", padx=10, pady=5)

        self.info_label = tk.Label(
            self.status_frame,
            text="",
            bg="#1a3d2e",
            fg="#ff9800",
            font=("Arial", 10, "bold"),
            anchor="e"
        )
        self.info_label.pack(side="right", padx=10, pady=5)

    def selecionar_arquivos(self):
        """Seleciona arquivos DXF/DWG"""
        arquivos = filedialog.askopenfilenames(
            title="Selecionar arquivos DXF/DWG de PILARES",
            filetypes=[
                ("Arquivos CAD", "*.dxf *.dwg"),
                ("DXF", "*.dxf"),
                ("DWG", "*.dwg"),
                ("Todos", "*.*")
            ]
        )

        if arquivos:
            self.arquivos_selecionados = list(arquivos)
            self.status_label.config(text=f"📁 {len(arquivos)} arquivo(s) selecionado(s)")

            nomes = [os.path.basename(f) for f in arquivos]
            if len(nomes) > 3:
                texto = f"{', '.join(nomes[:3])}, ..."
            else:
                texto = ', '.join(nomes)
            self.info_label.config(text=texto)

    def imprimir_etiquetas(self):
        """Abre o editor visual de etiquetas para seleção e edição antes de imprimir."""
        if not self.dados_processados:
            from tkinter import messagebox
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return
        try:
            # Inicializa variáveis de paginação e seleção se necessário
            if not hasattr(self, 'pagina_atual'):
                self.pagina_atual = 0
            if not hasattr(self, 'etiquetas_por_pagina'):
                self.etiquetas_por_pagina = 6
            if not hasattr(self, 'etiquetas_selecionadas') or not self.etiquetas_selecionadas:
                self.etiquetas_selecionadas = {i: True for i in range(len(self.dados_processados))}
            import math
            self.total_paginas = max(1, math.ceil(len(self.dados_processados) / self.etiquetas_por_pagina))
            # Chama o editor visual (janela_editor)
            self.janela_editor = None
            self._abrir_editor_etiquetas()
        except Exception as e:
            import traceback
            traceback.print_exc()
            from tkinter import messagebox
            messagebox.showerror("Erro", f"Erro ao abrir editor de etiquetas:\n{e}")

    def _abrir_editor_etiquetas(self):
        """Abre a janela do editor visual de etiquetas (profissional)."""
        import traceback
        from tkinter import messagebox
        try:
            if not self.dados_processados:
                messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
                return
            self.janela_editor = tk.Toplevel(self)
            self.janela_editor.title("✏️ EDITOR DE ETIQUETAS - PILARES")
            self.janela_editor.configure(bg="#0d2818")
            self.janela_editor.update_idletasks()
            screen_w = self.janela_editor.winfo_screenwidth()
            screen_h = self.janela_editor.winfo_screenheight()
            win_w = min(1200, screen_w - 40)
            win_h = min(900, screen_h - 80)
            x = max(0, (screen_w // 2) - (win_w // 2))
            y = max(0, (screen_h // 2) - (win_h // 2))
            self.janela_editor.geometry(f"{win_w}x{win_h}+{x}+{y}")
            if not hasattr(self, 'zoom_factor'):
                if screen_h < 900 or screen_w < 1200:
                    self.zoom_factor = 0.85
                else:
                    self.zoom_factor = 1.0
            if not hasattr(self, 'medidas_customizadas'):
                self.medidas_customizadas = {}
            if not hasattr(self, 'formas_customizadas'):
                self.formas_customizadas = {}
            if not hasattr(self, 'pagina_atual'):
                self.pagina_atual = 0
            if not hasattr(self, 'etiquetas_por_pagina'):
                self.etiquetas_por_pagina = 6
            if not hasattr(self, 'etiquetas_selecionadas'):
                self.etiquetas_selecionadas = {i: True for i in range(len(self.dados_processados))}
            # Garantir gerador para carregar desenhos antes da edição
            if ETIQUETAS_GERADOR_DISPONIVEL:
                try:
                    arquivos_dxf = []
                    if hasattr(self, 'arquivos_processados') and self.arquivos_processados:
                        arquivos_dxf = list(self.arquivos_processados)
                    elif hasattr(self, 'arquivos_selecionados') and self.arquivos_selecionados:
                        arquivos_dxf = list(self.arquivos_selecionados)

                    if not hasattr(self, 'gerador_etiquetas_dinamico') or self.gerador_etiquetas_dinamico is None:
                        self.gerador_etiquetas_dinamico = GeradorEtiquetasDinamico(
                            arquivos_dxf=arquivos_dxf,
                            pasta_etiquetas="output/etiquetas",
                            obra=self.var_obra.get() if hasattr(self, 'var_obra') else "OBRA 001",
                            pavimento=self.var_pavimento.get() if hasattr(self, 'var_pavimento') else "TÉRREO"
                        )
                    # Atualizar dados e customizações
                    self.gerador_etiquetas_dinamico.dados = list(self.dados_processados)
                    if hasattr(self, 'medidas_customizadas'):
                        self.gerador_etiquetas_dinamico.medidas_customizadas = self.medidas_customizadas
                    if hasattr(self, 'formas_customizadas'):
                        self.gerador_etiquetas_dinamico.formas_customizadas = self.formas_customizadas
                    
                    # CRÍTICO: Garantir que arquivo_dxf_base está setado para localizar_desenho_barra
                    if arquivos_dxf and not getattr(self.gerador_etiquetas_dinamico, 'arquivo_dxf_base', None):
                        self.gerador_etiquetas_dinamico.arquivo_dxf_base = arquivos_dxf[0]
                except Exception as e:
                    print(f"[WARN] Falha ao preparar gerador de desenhos: {e}")
            import math
            self.total_paginas = max(1, math.ceil(len(self.dados_processados) / self.etiquetas_por_pagina))
            # FRAME SUPERIOR - TÍTULO
            titulo_frame = tk.Frame(self.janela_editor, bg="#ff6f00")
            titulo_frame.pack(fill="x")
            tk.Label(
                titulo_frame,
                text="✏️ EDITOR DE ETIQUETAS - EDITE, SELECIONE E IMPRIMA",
                bg="#ff6f00",
                fg="white",
                font=("Arial", 12, "bold"),
                pady=8
            ).pack()
            # FRAME DO CANVAS
            canvas_frame = tk.Frame(self.janela_editor, bg="#0d2818")
            canvas_frame.pack(fill="both", expand=True, padx=10, pady=5)
            scrollbar = tk.Scrollbar(canvas_frame)
            scrollbar.pack(side="right", fill="y")
            self.canvas_etiq = tk.Canvas(
                canvas_frame,
                bg="white",
                yscrollcommand=scrollbar.set,
                highlightthickness=0
            )
            self.canvas_etiq.pack(side="left", fill="both", expand=True)
            scrollbar.config(command=self.canvas_etiq.yview)
            def _on_mousewheel(event):
                try:
                    self.canvas_etiq.yview_scroll(int(-1 * (event.delta / 120)), "units")
                except Exception:
                    pass
            self.canvas_etiq.bind("<MouseWheel>", _on_mousewheel)
            self.canvas_etiq.bind("<Button-1>", self._handle_canvas_click)
            try:
                self.desenhar_etiquetas_com_selecao()
            except Exception as e:
                self.canvas_etiq.create_text(
                    600, 450,
                    text=f"⚠️ Erro ao renderizar preview\n{str(e)}",
                    font=("Arial", 12),
                    fill="red"
                )
                print(f"[ERRO] Erro ao renderizar preview: {e}")
                messagebox.showerror("Erro ao renderizar preview", str(e))
            nav_frame = tk.Frame(self.janela_editor, bg="#34495e")
            nav_frame.pack(fill="x", padx=10, pady=5)
            self.label_pagina = tk.Label(
                nav_frame,
                text=f"Página {self.pagina_atual + 1} de {self.total_paginas}",
                bg="#34495e",
                fg="white",
                font=("Arial", 10, "bold")
            )
            self.label_pagina.pack(side="left", padx=10, pady=5)
            nav_btn_frame = tk.Frame(nav_frame, bg="#34495e")
            nav_btn_frame.pack(side="right", padx=5, pady=5)
            tk.Button(nav_btn_frame, text="⏮️ Primeira", command=self._ir_primeira_pagina_etiquetas,
                      bg="#16a085", fg="white", font=("Arial", 9), padx=8, pady=4, cursor="hand2").pack(side="left", padx=2)
            tk.Button(nav_btn_frame, text="◀ Anterior", command=self._ir_pagina_anterior_etiquetas,
                      bg="#27ae60", fg="white", font=("Arial", 9), padx=8, pady=4, cursor="hand2").pack(side="left", padx=2)
            tk.Button(nav_btn_frame, text="Próxima ▶", command=self._ir_proxima_pagina_etiquetas,
                      bg="#27ae60", fg="white", font=("Arial", 9), padx=8, pady=4, cursor="hand2").pack(side="left", padx=2)
            tk.Button(nav_btn_frame, text="Última ⏭️", command=self._ir_ultima_pagina_etiquetas,
                      bg="#16a085", fg="white", font=("Arial", 9), padx=8, pady=4, cursor="hand2").pack(side="left", padx=2)
            sel_frame = tk.Frame(self.janela_editor, bg="#1a3d2e")
            sel_frame.pack(fill="x", padx=10, pady=5)
            tk.Label(sel_frame, text="🔘 Seleção:", bg="#1a3d2e", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)
            tk.Button(sel_frame, text="☑️ MARCAR TODAS", command=self._marcar_todas_etiquetas,
                      bg="#27ae60", fg="white", font=("Arial", 9), padx=10, pady=4, cursor="hand2").pack(side="left", padx=3)
            tk.Button(sel_frame, text="☐ DESMARCAR TODAS", command=self._desmarcar_todas_etiquetas,
                      bg="#e74c3c", fg="white", font=("Arial", 9), padx=10, pady=4, cursor="hand2").pack(side="left", padx=3)
            total_selecionadas = sum(1 for v in self.etiquetas_selecionadas.values() if v)
            self.label_selecionadas = tk.Label(
                sel_frame,
                text=f"Selecionadas: {total_selecionadas}/{len(self.dados_processados)}",
                bg="#1a3d2e", fg="#ff9800", font=("Arial", 9, "bold")
            )
            self.label_selecionadas.pack(side="right", padx=10, pady=4)
            btn_frame = tk.Frame(self.janela_editor, bg="#1a3d2e")
            btn_frame.pack(fill="x", padx=10, pady=10)
            info_text = f"📋 Total: {len(self.dados_processados)} etiquetas | 💡 Clique nos valores para editar | ☑️ Selecione quais imprimir"
            tk.Label(btn_frame, text=info_text, bg="#1a3d2e", fg="white", font=("Arial", 9)).pack(side="left", padx=10, pady=5)
            btn_actions = tk.Frame(btn_frame, bg="#1a3d2e")
            btn_actions.pack(side="right", padx=5)
            tk.Button(btn_actions, text="ℹ️ COMO", command=self._mostrar_ajuda_edicao,
                      bg="#3498db", fg="white", font=("Arial", 9), padx=10, pady=5, cursor="hand2").pack(side="left", padx=3)
            tk.Button(btn_actions, text="✅ IMPRIMIR SELECIONADAS",
                      command=self._confirmar_e_imprimir_etiquetas,
                      bg="#27ae60", fg="white", font=("Arial", 10, "bold"), padx=15, pady=5, cursor="hand2").pack(side="left", padx=3)
            tk.Button(btn_actions, text="✕ FECHAR", command=self._fechar_editor_etiquetas,
                      bg="#e74c3c", fg="white", font=("Arial", 9), padx=10, pady=5, cursor="hand2").pack(side="left", padx=3)
        except Exception as e:
            print(f"[ERRO] Falha ao abrir editor: {e}")
            traceback.print_exc()
            messagebox.showerror("Erro ao abrir editor", str(e))

    def processar(self):
        """Processa os arquivos selecionados"""
        if not self.arquivos_selecionados:
            messagebox.showwarning("Atenção", "Por favor, selecione os arquivos DXF/DWG primeiro!")
            return

        if processar_pilares is None:
            messagebox.showerror(
                "Erro",
                "Motor de processamento indisponível.\n"
                "Verifique o módulo core.pilares_motor_v2 ou core.pilares_motor."
            )
            return

        try:
            self.status_label.config(text="⏳ Processando arquivos...")
            self.update()

            # Converter DWG para DXF no modo rápido
            arquivos = list(self.arquivos_selecionados)
            dxf_files = [f for f in arquivos if f.lower().endswith(".dxf")]
            dwg_files = [f for f in arquivos if f.lower().endswith(".dwg")]

            converted = []
            if dwg_files:
                converted = converter_lote_oda(dwg_files)
                if not converted and not dxf_files:
                    messagebox.showerror(
                        "Erro",
                        "Falha ao converter arquivos DWG para DXF.\n"
                        "Tente usar 'PROCESSAR 2.0 (Completo)' ou converta para DXF primeiro."
                    )
                    return

            arquivos_processar = dxf_files + converted
            if not arquivos_processar:
                messagebox.showwarning("Atenção", "Nenhum arquivo DXF válido para processar.")
                return

            self.arquivos_processados = list(arquivos_processar)
            self.dados_processados, self.total_kg, self.total_barras = processar_pilares(arquivos_processar)

            for item in self.tree.get_children():
                self.tree.delete(item)

            for dado in self.dados_processados:
                self.tree.insert("", "end", values=dado)

            self.status_label.config(text=f"✅ Processamento concluído")
            self.info_label.config(text=f"Total: {self.total_barras} barras | {self.total_kg:.2f} kg")

            if len(self.dados_processados) == 0:
                messagebox.showinfo("Informação", "Nenhum pilar foi encontrado nos arquivos.\n\nVerifique se o formato está correto:\n- Pilares: P1, P2, P3...\n- Barras: 4 N1 ø12.5 C=280")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar:\n{str(e)}")
            self.status_label.config(text="❌ Erro no processamento")
    


    def processar_v2(self):
        """Processa usando o motor DXF 2.0"""
        # LIMPAR CUSTOMIZAÇÕES ANTIGAS (PASSO 4 - sincronização)
        self.medidas_customizadas = {}  # {(pilar, pos): {bitola, comp, qtde}}
        self.formas_customizadas = {}   # {(pilar, pos): 'reta'|'gancho'|'estribo'}
        self.desenhos_customizados = {}  # {(pilar, pos): caminho_para_DXF}
        print("[INFO] Customizações antigas limpas para novo processamento")
        
        try:
            from core.pilares_motor_dual import processar_pilares as motor_v2
        except Exception as e:
            messagebox.showerror(
                "Erro",
                f"Motor DXF (Dual) não encontrado!\ncore/pilares_motor_dual.py ausente.\n{e}"
            )
            return  # << CORRETO

        if not self.arquivos_selecionados:
            messagebox.showwarning("Atenção", "Selecione arquivos primeiro!")
            return  # << CORRETO

        # Converter tudo para ACAD2018 (AC1032)
        conv = converter_lote_oda(self.arquivos_selecionados)
        if not conv:
            messagebox.showerror(
                "Erro",
                "Conversão ODA falhou ou nenhum arquivo pôde ser convertido para ACAD2018."
            )
            return

        try:
            dados, peso, barras = motor_v2(conv)

            self.dados_processados = dados
            self.total_kg = peso
            self.total_barras = barras
            # Guardar DXFs convertidos para alimentar o gerador
            self.arquivos_processados = list(conv)
            # Gerar etiquetas PNG usando o gerador dinâmico (layout 100% igual)
            try:
                if ETIQUETAS_GERADOR_DISPONIVEL:
                    print("[INFO] Criando gerador de etiquetas...")
                    os.makedirs(r"c:\EngenhariaPlanPro\etiquetas", exist_ok=True)
                    
                    # LIMPAR etiquetas antigas antes de gerar novas
                    from pathlib import Path
                    etiq_pasta = Path(r"c:\EngenhariaPlanPro\etiquetas")
                    for old_png in etiq_pasta.glob("*.png"):
                        old_png.unlink()
                    
                    print("[INFO] Gerando etiquetas PNG do editor...")
                    # GERAR PNGs DIRETO DO EDITOR (não usar GeradorEtiquetasDinamico)
                    self.caminhos_etiquetas_geradas = []
                    for idx, dado in enumerate(self.dados_processados):
                        pilar, pos, bitola, qtde, comp = dado[0], dado[1], dado[2], dado[3], dado[4]
                        png_path = self._gerar_png_etiqueta(idx, pilar, pos, bitola, qtde, comp, str(etiq_pasta))
                        self.caminhos_etiquetas_geradas.append(png_path)
                    
                    print(f"[INFO] {len(self.caminhos_etiquetas_geradas)} etiquetas PNG geradas do editor!")
                    
                    # VALIDAR SINCRONIZAÇÃO
                    assert len(self.caminhos_etiquetas_geradas) == len(self.dados_processados), \
                        f"Incompatibilidade de dados! PNGs={len(self.caminhos_etiquetas_geradas)} vs Dados={len(self.dados_processados)}"
                    print("[OK] Sincronização válida: dados_processados == caminhos_etiquetas_geradas")
            except Exception as e:
                print(f"[WARN] Erro ao gerar etiquetas PNG: {e}")

            self.preencher_tabela()

            self.status_label.config(text="Processamento DXF 2.0 concluído!")
            self.info_label.config(text=f"Total: {self.total_barras} barras | {self.total_kg:.2f} kg")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar DXF 2.0:\n{e}")

    def preencher_tabela(self):
        """Preenche a Treeview com os dados processados"""
        # Limpar a tabela
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Preencher
        for dado in self.dados_processados:
            if not dado or len(dado) < 5:
                print(f"[WARN] Dado inválido ou incompleto: {dado}")
                continue
            
            try:
                pilar, pos, bitola, qtd, comp = dado[0], dado[1], dado[2], dado[3], dado[4]
                peso = dado[5] if len(dado) > 5 else 0.0
            except (ValueError, TypeError, IndexError) as e:
                print(f"[WARN] Erro ao extrair dados da linha {dado}: {e}")
                continue
            
            chave = (pilar, pos)
            
            # Verificar se tem desenho customizado
            if chave in self.desenhos_customizados:
                if self.desenhos_customizados[chave] is None:
                    desenho_status = "(sem desenho)"
                else:
                    desenho_status = os.path.basename(self.desenhos_customizados[chave])
            else:
                desenho_status = "(duplo-clique para editar)"
            
            self.tree.insert("", "end", values=(
                pilar,
                pos,
                f"{bitola:.1f}",
                qtd,
                f"{comp:.2f}",
                f"{peso:.2f}",
                desenho_status
            ))

    def gerar_romaneio_conferencia(self):
        """Gera romaneio de conferência com checkboxes COMPLETO"""
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return

        # Criar janela
        janela_conf = tk.Toplevel(self)
        janela_conf.title("📋 CHECK LIST - Conferência de Pilares")
        janela_conf.geometry("1100x700")
        janela_conf.configure(bg="#0d2818")

        # Centralizar
        janela_conf.update_idletasks()
        x = (janela_conf.winfo_screenwidth() // 2) - 550
        y = (janela_conf.winfo_screenheight() // 2) - 350
        janela_conf.geometry(f"1100x700+{x}+{y}")

        # Frame superior com informações
        header_frame = tk.Frame(janela_conf, bg="#1a3d2e", relief="raised", bd=2)
        header_frame.pack(fill="x", padx=10, pady=10)

        tk.Label(
            header_frame,
            text="📋 CHECK LIST - CONFERÊNCIA DE PILARES",
            bg="#1a3d2e",
            fg="white",
            font=("Arial", 14, "bold")
        ).pack(pady=5)

        info_frame = tk.Frame(header_frame, bg="#1a3d2e")
        info_frame.pack()

        tk.Label(
            info_frame,
            text=f"Obra: {self.var_obra.get()}",
            bg="#1a3d2e",
            fg="white",
            font=("Arial", 10)
        ).pack(side="left", padx=20)

        tk.Label(
            info_frame,
            text=f"Pavimento: {self.var_pavimento.get()}",
            bg="#1a3d2e",
            fg="white",
            font=("Arial", 10)
        ).pack(side="left", padx=20)

        tk.Label(
            info_frame,
            text=f"Data: {datetime.now().strftime('%d/%m/%Y')}",
            bg="#1a3d2e",
            fg="white",
            font=("Arial", 10)
        ).pack(side="left", padx=20)

        tk.Label(
            info_frame,
            text=f"Total: {self.total_barras} barras | {self.total_kg:.2f} kg",
            bg="#1a3d2e",
            fg="#ff9800",
            font=("Arial", 10, "bold")
        ).pack(side="left", padx=20)

        # Frame de botões
        btn_frame = tk.Frame(janela_conf, bg="#2c3e50")
        btn_frame.pack(fill="x", padx=10, pady=5)

        tk.Button(
            btn_frame,
            text="✅ Marcar Todos",
            command=lambda: self.marcar_todos_checks(True),
            bg="#27ae60",
            fg="white",
            font=("Arial", 9, "bold")
        ).pack(side="left", padx=5, pady=5)

        tk.Button(
            btn_frame,
            text="❌ Desmarcar Todos",
            command=lambda: self.marcar_todos_checks(False),
            bg="#e74c3c",
            fg="white",
            font=("Arial", 9, "bold")
        ).pack(side="left", padx=5, pady=5)

        tk.Button(
            btn_frame,
            text="💾 Salvar Check List",
            command=lambda: self.salvar_checklist(),
            bg="#3498db",
            fg="white",
            font=("Arial", 9, "bold")
        ).pack(side="left", padx=5, pady=5)

        tk.Button(
            btn_frame,
            text="🖨️ Imprimir",
            command=lambda: self.imprimir_checklist(),
            bg="#9b59b6",
            fg="white",
            font=("Arial", 9, "bold")
        ).pack(side="left", padx=5, pady=5)

        # Frame principal com Canvas e Scrollbar
        main_frame = tk.Frame(janela_conf, bg="#f5f5f5")
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Editor com barras de rolagem e tamanho ajustável
        canvas = tk.Canvas(main_frame, bg="white")
        scrollbar_y = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollbar_x = ttk.Scrollbar(main_frame, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        canvas.pack(side="left", fill="both", expand=True)

        # Frame interno para o conteúdo, com largura mínima
        self.frame_checks = tk.Frame(canvas, bg="white")
        canvas.create_window((0, 0), window=self.frame_checks, anchor="nw")

        def on_configure(event):
            # Ajusta o scrollregion sempre que o frame interno muda de tamanho
            canvas.configure(scrollregion=canvas.bbox("all"))
        self.frame_checks.bind("<Configure>", on_configure)

        # Permitir redimensionamento mínimo para facilitar seleção
        self.frame_checks.update_idletasks()
        min_width = max(900, self.frame_checks.winfo_reqwidth())
        min_height = max(400, self.frame_checks.winfo_reqheight())
        canvas.config(width=min_width, height=min_height)

        # Dicionário para armazenar checkboxes
        self.checkboxes_conf = {}

        # Cabeçalho da tabela
        header_check = tk.Frame(self.frame_checks, bg="#34495e", relief="raised", bd=1)
        header_check.grid(row=0, column=0, columnspan=9, sticky="ew", padx=1, pady=1)

        headers = ["PILAR", "POS", "BITOLA", "QTD", "COMP", "CORTADO", "DOBRADO", "CONFERIDO", "CARREGADO"]
        for col, header in enumerate(headers):
            tk.Label(
                header_check,
                text=header,
                bg="#34495e",
                fg="white",
                font=("Arial", 8, "bold"),
                width=8 if col < 5 else 6
            ).grid(row=0, column=col, padx=1, pady=2)

        # Criar checkboxes para cada item
        pilar_atual = None
        row_num = 1

        for i, dado in enumerate(self.dados_processados):
            pilar = dado[0]
            pos = dado[1]
            bitola = f"ø{dado[2]:.1f}"
            qtde = dado[3]
            comp = f"{dado[4]:.2f}m"

            # Separador entre pilares diferentes
            if pilar != pilar_atual:
                pilar_atual = pilar
                separator = tk.Frame(self.frame_checks, bg="#ff9800", height=1)
                separator.grid(row=row_num, column=0, columnspan=9, sticky="ew", pady=2)
                row_num += 1

            # Frame para cada linha
            row_frame = tk.Frame(self.frame_checks, bg="#ecf0f1" if row_num % 2 == 0 else "white")
            row_frame.grid(row=row_num, column=0, columnspan=9, sticky="ew", padx=1, pady=0)

            # Dados compactos
            tk.Label(row_frame, text=pilar, font=("Arial", 8), width=8).grid(row=0, column=0)
            tk.Label(row_frame, text=pos, font=("Arial", 8), width=8).grid(row=0, column=1)
            tk.Label(row_frame, text=bitola, font=("Arial", 8), width=8).grid(row=0, column=2)
            tk.Label(row_frame, text=qtde, font=("Arial", 8), width=8).grid(row=0, column=3)
            tk.Label(row_frame, text=comp, font=("Arial", 8), width=8).grid(row=0, column=4)

            # Checkboxes compactos
            chave = f"{pilar}_{pos}"
            self.checkboxes_conf[chave] = {
                'cortado': tk.BooleanVar(value=False),
                'dobrado': tk.BooleanVar(value=False),
                'conferido': tk.BooleanVar(value=False),
                'carregado': tk.BooleanVar(value=False)
            }

            tk.Checkbutton(
                row_frame,
                variable=self.checkboxes_conf[chave]['cortado'],
                bg=row_frame['bg'],
                width=2,
                padx=0,
                pady=0
            ).grid(row=0, column=5)

            tk.Checkbutton(
                row_frame,
                variable=self.checkboxes_conf[chave]['dobrado'],
                bg=row_frame['bg'],
                width=2,
                padx=0,
                pady=0
            ).grid(row=0, column=6)

            tk.Checkbutton(
                row_frame,
                variable=self.checkboxes_conf[chave]['conferido'],
                bg=row_frame['bg'],
                width=2,
                padx=0,
                pady=0
            ).grid(row=0, column=7)

            tk.Checkbutton(
                row_frame,
                variable=self.checkboxes_conf[chave]['carregado'],
                bg=row_frame['bg'],
                width=2,
                padx=0,
                pady=0
            ).grid(row=0, column=8)

            row_num += 1

        # Atualizar scroll region
        self.frame_checks.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

        # Armazenar referências para uso posterior
        self.janela_checklist = janela_conf
        self.canvas_checklist = canvas

    def marcar_todos_checks(self, valor):
        """Marca ou desmarca todos os checkboxes"""
        if hasattr(self, 'checkboxes_conf'):
            for chave in self.checkboxes_conf:
                for tipo in self.checkboxes_conf[chave]:
                    self.checkboxes_conf[chave][tipo].set(valor)

    def salvar_checklist(self):
        """Salva o checklist em arquivo"""
        if not hasattr(self, 'checkboxes_conf'):
            return

        arquivo = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Texto", "*.txt"), ("CSV", "*.csv")],
            initialfile=f"checklist_pilares_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
        )

        if arquivo:
            with open(arquivo, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("CHECK LIST - CONFERÊNCIA DE PILARES\n")
                f.write("=" * 80 + "\n")
                f.write(f"Obra: {self.var_obra.get()}\n")
                f.write(f"Pavimento: {self.var_pavimento.get()}\n")
                f.write(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
                f.write("=" * 80 + "\n\n")

                f.write(f"{'PILAR':<10} {'POS':<8} {'CORTADO':<10} {'DOBRADO':<10} {'CONFERIDO':<12} {'CARREGADO':<10}\n")
                f.write("-" * 70 + "\n")

                for i, dado in enumerate(self.dados_processados):
                    pilar = dado[0]
                    pos = dado[1]
                    chave = f"{pilar}_{pos}"

                    if chave in self.checkboxes_conf:
                        cortado = "✓" if self.checkboxes_conf[chave]['cortado'].get() else "□"
                        dobrado = "✓" if self.checkboxes_conf[chave]['dobrado'].get() else "□"
                        conferido = "✓" if self.checkboxes_conf[chave]['conferido'].get() else "□"
                        carregado = "✓" if self.checkboxes_conf[chave]['carregado'].get() else "□"

                        f.write(f"{pilar:<10} {pos:<8} {cortado:<10} {dobrado:<10} {conferido:<12} {carregado:<10}\n")

                f.write("\n" + "=" * 80 + "\n")
                f.write("Assinatura: _______________________________\n")
                f.write("Data: ____/____/________\n")

            messagebox.showinfo("Sucesso", f"Check List salvo em:\n{arquivo}")

    def imprimir_checklist(self):
        """Imprime o checklist em HTML"""
        try:
            # Gerar conteúdo do checklist
            linhas_html = []
            
            pilar_atual = None
            for dado in self.dados_processados:
                if dado[0] != pilar_atual:
                    pilar_atual = dado[0]
                    linhas_html.append(f'<tr class="pilar-header"><td colspan="6">{pilar_atual}</td></tr>')

                linhas_html.append(
                    f'<tr>'
                    f'<td><input type="checkbox"></td>'
                    f'<td>{dado[1]}</td>'
                    f'<td>Ø{dado[2]:.1f}mm</td>'
                    f'<td>{dado[3]}</td>'
                    f'<td>{dado[4]:.2f}m</td>'
                    f'<td>{dado[5]:.2f}kg</td>'
                    f'</tr>'
                )

            html_content = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <title>Check List de Pilares</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: white;
        }}
        .header {{
            text-align: center;
            margin-bottom: 30px;
            border-bottom: 3px solid #2c3e50;
            padding-bottom: 15px;
        }}
        .header h1 {{
            color: #2c3e50;
            margin: 10px 0;
        }}
        .info {{
            background-color: #f8f9fa;
            padding: 15px;
            margin-bottom: 20px;
            border-left: 4px solid #e74c3c;
        }}
        .btn-container {{
            text-align: center;
            margin: 20px 0;
        }}
        .btn-print {{
            background-color: #27ae60;
            color: white;
            border: none;
            padding: 12px 30px;
            font-size: 16px;
            border-radius: 5px;
            cursor: pointer;
            margin: 10px;
        }}
        .btn-print:hover {{
            background-color: #229954;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
        }}
        th {{
            background-color: #34495e;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .pilar-header {{
            background-color: #e74c3c !important;
            color: white;
            font-weight: bold;
            font-size: 14px;
        }}
        .pilar-header td {{
            padding: 8px 12px;
        }}
        .assinatura {{
            margin-top: 40px;
            display: flex;
            justify-content: space-around;
        }}
        .assinatura-box {{
            text-align: center;
            border-top: 1px solid #000;
            width: 200px;
            margin-top: 40px;
        }}
        @media print {{
            body {{ margin: 0; }}
            .btn-container {{ display: none; }}
            input {{ print-color-adjust: exact; -webkit-print-color-adjust: exact; }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>✅ CHECK LIST DE PILARES</h1>
        <p>EngenhariaPlanPro</p>
    </div>
    
    <div class="info">
        <strong>Obra:</strong> {self.var_obra.get()}<br>
        <strong>Pavimento:</strong> {self.var_pavimento.get()}<br>
        <strong>Data:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}
    </div>
    
    <div class="btn-container">
        <button class="btn-print" onclick="window.print();">🖨️ IMPRIMIR AGORA</button>
    </div>
    
    <table>
        <thead>
            <tr>
                <th>✓</th>
                <th>Posição</th>
                <th>Bitola</th>
                <th>Quantidade</th>
                <th>Comprimento</th>
                <th>Peso</th>
            </tr>
        </thead>
        <tbody>
            {''.join(linhas_html)}
        </tbody>
    </table>
    
    <div class="assinatura">
        <div class="assinatura-box">
            <p>Responsável</p>
            <p>Data: ____/____/________</p>
        </div>
        <div class="assinatura-box">
            <p>Conferente</p>
            <p>Data: ____/____/________</p>
        </div>
    </div>
</body>
</html>
"""

            # Criar arquivo HTML temporário
            html_temp = tempfile.mktemp(suffix=".html")
            with open(html_temp, 'w', encoding='utf-8') as f:
                f.write(html_content)

            # Abrir no navegador
            import webbrowser
            webbrowser.open('file://' + html_temp)
            print(f"[CHECK LIST HTML] Aberto no navegador.")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar check list:\n{str(e)}")

    def gerar_romaneio(self):
        """Gera romaneio detalhado com múltiplas visualizações como em pilares"""
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return

        janela = tk.Toplevel(self)
        janela.title("Romaneio de PILARES - Múltiplas Visualizações")
        janela.geometry("1000x700")
        janela.configure(bg="#0d2818")

        janela.update_idletasks()
        x = (janela.winfo_screenwidth() // 2) - 500
        y = (janela.winfo_screenheight() // 2) - 350
        janela.geometry(f"1000x700+{x}+{y}")

        # Frame superior com botões
        btn_frame = tk.Frame(janela, bg="#1a3d2e")
        btn_frame.pack(fill="x", padx=10, pady=10)

        # Variável para controlar aba ativa
        self.aba_ativa = None
        self.texto_ativo = None

        def salvar_aba_ativa():
            if self.texto_ativo:
                self.salvar_txt(self.texto_ativo.get("1.0", "end-1c"))

        def imprimir_aba_ativa():
            # Pegar aba ativa
            tab_id = notebook.select()
            if not tab_id:
                messagebox.showwarning("Atenção", "Selecione uma aba!")
                return
            
            # Procurar pelo widget de texto na aba (busca recursiva)
            tab = notebook.nametowidget(tab_id)
            texto_widget = None
            
            def encontrar_texto(widget):
                nonlocal texto_widget
                if isinstance(widget, ScrolledText):
                    texto_widget = widget
                    return True
                for child in widget.winfo_children():
                    if encontrar_texto(child):
                        return True
                return False
            
            encontrar_texto(tab)
            
            if texto_widget:
                conteudo = texto_widget.get("1.0", "end-1c")
                if conteudo.strip():
                    self.imprimir_com_preview(conteudo)
                else:
                    messagebox.showwarning("Aviso", "A aba está vazia!")
            else:
                messagebox.showwarning("Erro", "Não foi encontrado conteúdo para imprimir!")

        tk.Button(
            btn_frame,
            text="💾 Salvar TXT",
            command=salvar_aba_ativa,
            bg="#2e7d32",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=10
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="🖨️ Imprimir",
            command=imprimir_aba_ativa,
            bg="#6a1b9a",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=10
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="📊 Exportar Tudo Excel",
            command=self.exportar_excel,
            bg="#ff6f00",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=10
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="❌ Fechar",
            command=janela.destroy,
            bg="#757575",
            fg="white",
            font=("Arial", 10),
            padx=10
        ).pack(side="right", padx=5)

        # Notebook para abas
        notebook = ttk.Notebook(janela)
        notebook.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Callback para detectar aba ativa
        def ao_trocar_aba(event=None):
            tab_id = notebook.select()
            if tab_id:
                tab = notebook.nametowidget(tab_id)
                for widget in tab.winfo_children():
                    if isinstance(widget, tk.Widget) and hasattr(widget, 'get'):
                        self.texto_ativo = widget
                        break

        notebook.bind("<<NotebookTabChanged>>", ao_trocar_aba)

        # Estilo para o notebook
        style = ttk.Style()
        style.configure('TNotebook.Tab', padding=[20, 10])

        # ========== ABA 1: ROMANEIO GERAL ==========
        frame_geral = tk.Frame(notebook, bg="#1e1e1e")
        notebook.add(frame_geral, text="📋 Romaneio Geral")

        texto_geral = ScrolledText(
            frame_geral,
            bg="#1e1e1e",
            fg="#ff9800",
            font=("Courier New", 10),
            wrap="none"
        )
        texto_geral.pack(fill="both", expand=True)

        # Gerar conteúdo do romaneio geral
        conteudo_geral = []
        conteudo_geral.append("=" * 80)
        conteudo_geral.append("                    ROMANEIO DE PILARES - GERAL")
        conteudo_geral.append("=" * 80)
        conteudo_geral.append(f"Obra:      {self.var_obra.get()}")
        conteudo_geral.append(f"Pavimento: {self.var_pavimento.get()}")
        conteudo_geral.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        conteudo_geral.append("=" * 80)
        conteudo_geral.append("")

        pilar_atual = None
        for dado in self.dados_processados:
            if dado[0] != pilar_atual:
                pilar_atual = dado[0]
                conteudo_geral.append("")
                conteudo_geral.append(f">>> {pilar_atual}")
                conteudo_geral.append("-" * 60)

            conteudo_geral.append(
                f"    {dado[1]:<6} Ø{dado[2]:<6.1f} Qtd: {dado[3]:<4} "
                f"Comp: {dado[4]:<6.2f}m  Peso: {dado[5]:<8.2f}kg"
            )

        conteudo_geral.append("")
        conteudo_geral.append("=" * 80)
        conteudo_geral.append(f"TOTAL GERAL: {self.total_kg:.2f} kg")
        conteudo_geral.append("=" * 80)

        texto_geral.insert("1.0", "\n".join(conteudo_geral))

        # ========== ABA 2: POR TIPO DE BARRA ==========
        frame_tipo = tk.Frame(notebook, bg="#1e1e1e")
        notebook.add(frame_tipo, text="🔧 Por Tipo")

        texto_tipo = ScrolledText(
            frame_tipo,
            bg="#1e1e1e",
            fg="#00bcd4",
            font=("Courier New", 10),
            wrap="none"
        )
        texto_tipo.pack(fill="both", expand=True)

        # Separar por tipo
        tipos_agrupados = {}
        for dado in self.dados_processados:
            pos = dado[1]
            bitola = dado[2]
            comp = dado[4]
            tipo = AnalisadorGeometricoPilares.identificar_tipo_pilar(pos, bitola, comp)

            if tipo not in tipos_agrupados:
                tipos_agrupados[tipo] = []
            tipos_agrupados[tipo].append(dado)

        conteudo_tipo = []
        conteudo_tipo.append("=" * 80)
        conteudo_tipo.append("                ROMANEIO DE PILARES - POR TIPO DE BARRA")
        conteudo_tipo.append("=" * 80)
        conteudo_tipo.append(f"Obra:      {self.var_obra.get()}")
        conteudo_tipo.append(f"Pavimento: {self.var_pavimento.get()}")
        conteudo_tipo.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        conteudo_tipo.append("=" * 80)

        for tipo in sorted(tipos_agrupados.keys()):
            conteudo_tipo.append("")
            conteudo_tipo.append(f"╔{'═' * 78}╗")
            conteudo_tipo.append(f"║ TIPO: {tipo:<70} ║")
            conteudo_tipo.append(f"╚{'═' * 78}╝")
            conteudo_tipo.append("")

            pilar_atual = None
            subtotal_tipo = 0.0
            subtotal_barras = 0

            for dado in tipos_agrupados[tipo]:
                if dado[0] != pilar_atual:
                    pilar_atual = dado[0]
                    conteudo_tipo.append(f"  ▶ {pilar_atual}")
                    conteudo_tipo.append("  " + "-" * 60)

                conteudo_tipo.append(
                    f"      {dado[1]:<6} Ø{dado[2]:<6.1f} Qtd: {dado[3]:<4} "
                    f"Comp: {dado[4]:<6.2f}m  Peso: {dado[5]:<8.2f}kg"
                )
                subtotal_tipo += dado[5]
                subtotal_barras += dado[3]

            conteudo_tipo.append("  " + "-" * 60)
            conteudo_tipo.append(f"  Subtotal {tipo}:")
            conteudo_tipo.append(f"    Barras: {subtotal_barras} | Peso: {subtotal_tipo:.2f} kg")

        conteudo_tipo.append("")
        conteudo_tipo.append("=" * 80)
        conteudo_tipo.append(f"TOTAL GERAL: {self.total_barras} barras | {self.total_kg:.2f} kg")
        conteudo_tipo.append("=" * 80)

        texto_tipo.insert("1.0", "\n".join(conteudo_tipo))

        # ========== ABA 3: POR BITOLA ==========
        frame_bitola = tk.Frame(notebook, bg="#1e1e1e")
        notebook.add(frame_bitola, text="📏 Por Bitola")

        texto_bitola = ScrolledText(
            frame_bitola,
            bg="#1e1e1e",
            fg="#4caf50",
            font=("Courier New", 10),
            wrap="none"
        )
        texto_bitola.pack(fill="both", expand=True)

        # Agrupar por bitola
        bitolas_agrupadas = {}
        for dado in self.dados_processados:
            bitola = dado[2]
            if bitola not in bitolas_agrupadas:
                bitolas_agrupadas[bitola] = []
            bitolas_agrupadas[bitola].append(dado)

        conteudo_bitola = []
        conteudo_bitola.append("=" * 80)
        conteudo_bitola.append("                  ROMANEIO DE PILARES - POR BITOLA")
        conteudo_bitola.append("=" * 80)
        conteudo_bitola.append(f"Obra:      {self.var_obra.get()}")
        conteudo_bitola.append(f"Pavimento: {self.var_pavimento.get()}")
        conteudo_bitola.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        conteudo_bitola.append("=" * 80)

        for bitola in sorted(bitolas_agrupadas.keys()):
            conteudo_bitola.append("")
            conteudo_bitola.append(f"╔{'═' * 78}╗")
            conteudo_bitola.append(f"║ BITOLA: Ø {bitola:.1f} mm{' ' * (66 - len(f'{bitola:.1f}'))} ║")
            conteudo_bitola.append(f"╚{'═' * 78}╝")
            conteudo_bitola.append("")

            pilar_atual = None
            subtotal_peso = 0.0
            subtotal_barras = 0

            for dado in bitolas_agrupadas[bitola]:
                if dado[0] != pilar_atual:
                    pilar_atual = dado[0]
                    conteudo_bitola.append(f"  ▶ {pilar_atual}")
                    conteudo_bitola.append("  " + "-" * 60)

                conteudo_bitola.append(
                    f"      {dado[1]:<6} Qtd: {dado[3]:<4} "
                    f"Comp: {dado[4]:<6.2f}m  Peso: {dado[5]:<8.2f}kg"
                )
                subtotal_peso += dado[5]
                subtotal_barras += dado[3]

            conteudo_bitola.append("  " + "-" * 60)
            conteudo_bitola.append(f"  Subtotal Ø {bitola:.1f}mm:")
            conteudo_bitola.append(f"    Barras: {subtotal_barras} | Peso: {subtotal_peso:.2f} kg")
            percentual = (subtotal_peso / self.total_kg) * 100
            conteudo_bitola.append(f"    Percentual do total: {percentual:.1f}%")

        conteudo_bitola.append("")
        conteudo_bitola.append("=" * 80)
        conteudo_bitola.append(f"TOTAL GERAL: {self.total_barras} barras | {self.total_kg:.2f} kg")
        conteudo_bitola.append("=" * 80)

        texto_bitola.insert("1.0", "\n".join(conteudo_bitola))

        # ========== ABA 4: RESUMO EXECUTIVO ==========
        frame_resumo = tk.Frame(notebook, bg="#1e1e1e")
        notebook.add(frame_resumo, text="📊 Resumo Executivo")

        texto_resumo = ScrolledText(
            frame_resumo,
            bg="#1e1e1e",
            fg="#ffc107",
            font=("Courier New", 10),
            wrap="none"
        )
        texto_resumo.pack(fill="both", expand=True)

        conteudo_resumo = []
        conteudo_resumo.append("=" * 80)
        conteudo_resumo.append("                    RESUMO EXECUTIVO - PILARES")
        conteudo_resumo.append("=" * 80)
        conteudo_resumo.append(f"Obra:      {self.var_obra.get()}")
        conteudo_resumo.append(f"Pavimento: {self.var_pavimento.get()}")
        conteudo_resumo.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        conteudo_resumo.append("=" * 80)
        conteudo_resumo.append("")

        # Estatísticas gerais
        conteudo_resumo.append("ESTATÍSTICAS GERAIS:")
        conteudo_resumo.append("-" * 40)

        # Contar pilares únicas
        pilares_unicas = set(dado[0] for dado in self.dados_processados)
        conteudo_resumo.append(f"  Total de Pilares:        {len(pilares_unicas)}")
        conteudo_resumo.append(f"  Total de Posições:     {len(self.dados_processados)}")
        conteudo_resumo.append(f"  Total de Barras:       {self.total_barras}")
        
        conteudo_resumo.append("")
        conteudo_resumo.append("=" * 80)
        conteudo_resumo.append(f"TOTAL GERAL: {self.total_barras} barras | {self.total_kg:.2f} kg")
        conteudo_resumo.append("=" * 80)
        
        texto_resumo.insert("1.0", "\n".join(conteudo_resumo))

    def _atualizar_preview_zoom(self):
        if not hasattr(self, '_preview_img') or self._preview_img is None:
            return
        zoom = self.var_zoom_preview.get() / 100.0
        w, h = self._preview_img.width, self._preview_img.height
        new_w, new_h = int(w * zoom), int(h * zoom)
        if new_w < 10 or new_h < 10:
            return
        img_resized = self._preview_img.resize((new_w, new_h), Image.LANCZOS)
        self._preview_imgtk = ImageTk.PhotoImage(img_resized)
        self.canvas_etiq.delete("all")
        self.canvas_etiq.create_image(0, 0, anchor="nw", image=self._preview_imgtk)
        self.canvas_etiq.config(scrollregion=(0, 0, new_w, new_h))

    def desenhar_pagina_etiquetas_pilares(self):
        """Desenha a página atual de etiquetas para pilares usando o layout e dimensões reais do gerador dinâmico."""
        # O preview agora usa o mesmo layout da FASE 4 (10x15cm, 3 picotes, margens e escala reais)
        self.desenhar_pagina_etiquetas_pilares_fase4()

    # 🆕 FASE 4: Novo método para layout 10x15cm com 3 picotes
    def desenhar_pagina_etiquetas_pilares_fase4(self):
        # Desenha múltiplas etiquetas 10x15 cm por página (modo rolo)
        try:
            yview_prev = getattr(self, '_pending_yview', None)
        except Exception:
            yview_prev = None
        if yview_prev is None:
            try:
                yview_prev = self.canvas_etiq.yview()
            except Exception:
                yview_prev = None
        self.canvas_etiq.delete("all")
        self._barcode_images = []
        self._desenho_images = []

        # Escala e dimensões reais
        PX_MM = CFG_PX_MM if 'CFG_PX_MM' in globals() and ETIQUETAS_LAYOUT_CFG else 4
        MARGEM = (MARGEM_EXTERNA_MM if ETIQUETAS_LAYOUT_CFG else 10) * PX_MM
        LARGURA_ETIQ = (LARGURA_ETIQUETA_MM if ETIQUETAS_LAYOUT_CFG else 100) * PX_MM
        ALTURA_TOPO = (TOPO_ALTURA_MM if ETIQUETAS_LAYOUT_CFG else 93) * PX_MM
        ALTURA_MICRO = (SECAO_MICRO_ALTURA_MM if ETIQUETAS_LAYOUT_CFG else 19) * PX_MM
        ESPACO_PICOTE = (ESPACO_PICOTE_MM if ETIQUETAS_LAYOUT_CFG else 2) * PX_MM

        # Cálculo de altura total da etiqueta (sem bloco verde de medidas)
        # Adicionar margem extra para evitar corte do último picote
        altura_etiqueta = ALTURA_TOPO + (ESPACO_PICOTE // 2) + 3 * (ALTURA_MICRO + ESPACO_PICOTE) + ALTURA_MICRO

        # Centro horizontal na tela
        canvas_w = int(self.canvas_etiq.cget('width'))
        x_base = (canvas_w - (LARGURA_ETIQ + 2 * MARGEM)) // 2 + MARGEM

        # Índice inicial baseado na página
        inicio = self.pagina_atual * self.etiquetas_por_pagina
        fim = min(len(self.dados_processados), inicio + self.etiquetas_por_pagina)

        y_cursor = MARGEM
        for i in range(inicio, fim):
            # Validar índice e dados
            if i < 0 or i >= len(self.dados_processados):
                print(f"[WARN] Índice {i} inválido em fase4 (0-{len(self.dados_processados)-1})")
                continue
            
            dado = self.dados_processados[i]
            if not dado or len(dado) < 5:
                print(f"[WARN] Dado inválido em fase4 no índice {i}: {dado}")
                continue
            
            try:
                pilar, pos, bitola, qtde, comp = (
                    dado[0],
                    str(dado[1]),
                    float(dado[2]),
                    int(dado[3]),
                    float(dado[4]),
                )
            except (ValueError, TypeError, IndexError) as e:
                print(f"[WARN] Erro ao extrair dados em fase4: {e}")
                continue

            # Moldura topo e conteúdo técnico (com desenho dentro)
            self._desenhar_moldura_etiqueta_fase4(x_base, y_cursor, LARGURA_ETIQ, ALTURA_TOPO)
            self._desenhar_topo_identico_fase4(x_base, y_cursor, LARGURA_ETIQ, ALTURA_TOPO, pilar, pos, bitola, qtde, comp)

            # 3 seções inferiores (direto após o topo, sem bloco de medidas)
            base_y = y_cursor + ALTURA_TOPO + (ESPACO_PICOTE // 2)
            for idx in range(3):
                y_sec = base_y + idx * (ALTURA_MICRO + ESPACO_PICOTE)
                self.canvas_etiq.create_rectangle(x_base, y_sec, x_base + LARGURA_ETIQ, y_sec + ALTURA_MICRO, outline="#cccccc", width=1)
                self._desenhar_secao_micro_fase4(x_base + 4, y_sec + 2, LARGURA_ETIQ - 8, ALTURA_MICRO - 4, pilar, pos, bitola, qtde, comp)
                if idx < 2:
                    self._desenhar_picote_fase4(x_base, y_sec + ALTURA_MICRO + (ESPACO_PICOTE // 2), LARGURA_ETIQ)

            y_cursor += altura_etiqueta + (ESPACO_PICOTE * 3)

        # Calcular total de páginas
        self.total_paginas = max(1, math.ceil(len(self.dados_processados) / self.etiquetas_por_pagina))

        # Ajustar área de scroll para caber todas as etiquetas desta página + margem extra
        self.canvas_etiq.configure(scrollregion=(0, 0, canvas_w, max(1188, y_cursor + MARGEM + (ALTURA_MICRO * 2))))

        # Reaplicar zoom atual sem perder posição
        self._zoom_aplicado = 1.0
        self._reaplicar_zoom()
        if 'yview_prev' in locals() and yview_prev:
            try:
                self.canvas_etiq.yview_moveto(yview_prev[0])
            except Exception:
                pass
        if hasattr(self, '_pending_yview'):
            try:
                del self._pending_yview
            except Exception:
                pass

        # Atualizar label de página
        if hasattr(self, 'label_pagina'):
            self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {self.total_paginas}")

    def _desenhar_topo_identico_fase4(self, x, y, w, h, pilar, pos, bitola, qtde, comp):
        """Desenha o topo (9,3 cm) com OS, faixa vertical e tabela técnica."""
        pxmm = (CFG_PX_MM if 'CFG_PX_MM' in globals() and ETIQUETAS_LAYOUT_CFG else 4)
        def mm(v):
            return int(round(v * pxmm))

        faixa_larg = mm(FAIXA_VERTICAL_LARGURA_MM) if ETIQUETAS_LAYOUT_CFG else mm(10)
        os_w = mm(OS_BLOCO_LARGURA_MM) if ETIQUETAS_LAYOUT_CFG else mm(18)
        os_h = mm(OS_BLOCO_ALTURA_MM) if ETIQUETAS_LAYOUT_CFG else mm(30)

        area_os_x1 = x + w - os_w - faixa_larg
        area_os_y1 = y
        area_os_x2 = x + w - faixa_larg
        area_os_y2 = y + os_h
        self.canvas_etiq.create_rectangle(area_os_x1, area_os_y1, area_os_x2, area_os_y2, outline="#000", width=1)
        self.canvas_etiq.create_text(area_os_x1 + 6, area_os_y1 + 6, text="OS", font=("Arial", 10, "bold"), anchor="nw")
        
        # PASSO 3: Evitar truncamento/quebra de linhas na caixa OS
        os_txt = f"{self.pagina_atual + 1}-{self.total_paginas}" if hasattr(self, 'total_paginas') else "-"
        # Dividir em linhas com espaçamento fixo (15px entre linhas)
        linhas_os = os_txt.split("-")
        start_y = area_os_y1 + os_h // 2 - mm(6)
        espaco_linha = mm(8)
        for i, linha in enumerate(linhas_os):
            self.canvas_etiq.create_text((area_os_x1 + area_os_x2)//2, start_y + i*espaco_linha, 
                                         text=linha, font=("Arial", 11, "bold"), anchor="center")

        faixa_x1 = x + w - faixa_larg
        faixa_x2 = x + w
        self.canvas_etiq.create_rectangle(faixa_x1, y, faixa_x2, y + h, outline="#000", width=1)

        obra_nome = self.var_obra.get() if hasattr(self, 'var_obra') else "OBRA.001"
        if Image is None or ImageTk is None:
            self.canvas_etiq.create_text(faixa_x1 + faixa_larg//2, y + h//2, text=obra_nome, angle=90)
        else:
            try:
                img_tmp = Image.new('RGBA', (mm(60), faixa_larg), (255, 255, 255, 0))
                from PIL import ImageDraw, ImageFont
                draw = ImageDraw.Draw(img_tmp)
                try:
                    fnt = ImageFont.truetype("arial.ttf", 12)
                except Exception:
                    fnt = ImageFont.load_default()
                # PASSO 2: Centrar texto verticalmente (antes da rotação)
                tw, th = draw.textsize(obra_nome, font=fnt)
                text_x = (img_tmp.width - tw) // 2
                text_y = (img_tmp.height - th) // 2
                draw.text((text_x, text_y), obra_nome, fill=(0,0,0,255), font=fnt)
                # Rotacionar 90 graus para exibição vertical
                img_tmp = img_tmp.rotate(90, expand=True)
                photo = ImageTk.PhotoImage(img_tmp)
                if not hasattr(self, '_desenho_images'):
                    self._desenho_images = []
                self._desenho_images.append(photo)
                # Centralizar imagem rotacionada no espaço da faixa
                self.canvas_etiq.create_image(faixa_x1 + (faixa_larg//2), y + h//2, image=photo)
            except Exception:
                self.canvas_etiq.create_text(faixa_x1 + faixa_larg//2, y + h//2, text=obra_nome, angle=90)

        area_util_x1 = x + 6
        area_util_x2 = area_os_x1 - 6
        area_util_w = max(0, area_util_x2 - area_util_x1)

        # PASSO 2: Espaçamento fixo para textos do cabeçalho (8mm entre linhas) - AJUSTADO
        step_y = mm(8)  # 8mm entre linhas
        y_current = y + mm(12)  # Margem do topo aumentada para descer o cabeçalho

        # Linha 1: Sigla/Obra - FONTE MAIOR
        self.canvas_etiq.create_text(area_util_x1, y_current, text="Sigla/Obra", font=("Arial", 8), anchor="nw")
        self.canvas_etiq.create_text(area_util_x1 + mm(24), y_current, text=obra_nome.replace(" ", " - "), font=("Arial", 11, "bold"), anchor="nw")
        y_current += step_y

        # Linha 2: Desenho - FONTE MAIOR
        self.canvas_etiq.create_text(area_util_x1, y_current, text="Desenho", font=("Arial", 8), anchor="nw")
        self.canvas_etiq.create_text(area_util_x1 + mm(24), y_current, text="PEDIDO - 00000 - Rev.0", font=("Arial", 10, "bold"), anchor="nw")
        y_current += step_y

        # Linha 3: Pavimento - FONTE MAIOR
        self.canvas_etiq.create_text(area_util_x1, y_current, text="Pavimento", font=("Arial", 8), anchor="nw")
        self.canvas_etiq.create_text(area_util_x1 + mm(24), y_current, text="ARMADO", font=("Arial", 10, "bold"), anchor="nw")
        y_current += step_y

        # Linha 4: Elemento + POS - FONTE MAIOR
        self.canvas_etiq.create_text(area_util_x1, y_current, text="Elemento", font=("Arial", 8), anchor="nw")
        self.canvas_etiq.create_text(area_util_x1 + mm(24), y_current, text=f"{pilar}", font=("Arial", 10, "bold"), anchor="nw")
        self.canvas_etiq.create_text(area_util_x2 - mm(25), y_current, text="POS", font=("Arial", 8))
        self.canvas_etiq.create_text(area_util_x2 - mm(10), y_current, text=f"{pos}", font=("Arial", 12, "bold"))

        th = mm(TABELA_ALTURA_HEADER_MM) if ETIQUETAS_LAYOUT_CFG else mm(8)
        tr = mm(TABELA_ALTURA_LINHA_MM) if ETIQUETAS_LAYOUT_CFG else mm(10)
        tab_y1 = y + mm(45)  # DESCEU DE 40 PARA 45
        tab_y2 = tab_y1 + th + tr
        tab_x = area_util_x1
        cw1 = mm(COL_BITOLA_MM if ETIQUETAS_LAYOUT_CFG else 16)
        cw2 = mm(COL_COMPR_UNIT_MM if ETIQUETAS_LAYOUT_CFG else 34)
        cw3 = mm(COL_PESO_MM if ETIQUETAS_LAYOUT_CFG else 22)
        cw4 = mm(COL_QTDE_MM if ETIQUETAS_LAYOUT_CFG else 18)
        tot = cw1 + cw2 + cw3 + cw4
        if tot > area_util_w:
            escala = area_util_w / tot
            cw1 = int(cw1 * escala); cw2 = int(cw2 * escala); cw3 = int(cw3 * escala); cw4 = int(cw4 * escala)
        self.canvas_etiq.create_rectangle(tab_x, tab_y1, tab_x + cw1 + cw2 + cw3 + cw4, tab_y2, outline="#000", width=1)
        self.canvas_etiq.create_line(tab_x + cw1, tab_y1, tab_x + cw1, tab_y2)
        self.canvas_etiq.create_line(tab_x + cw1 + cw2, tab_y1, tab_x + cw1 + cw2, tab_y2)
        self.canvas_etiq.create_line(tab_x + cw1 + cw2 + cw3, tab_y1, tab_x + cw1 + cw2 + cw3, tab_y2)
        self.canvas_etiq.create_line(tab_x, tab_y1 + th, tab_x + cw1 + cw2 + cw3 + cw4, tab_y1 + th)
        self.canvas_etiq.create_text(tab_x + cw1//2, tab_y1 + th//2, text="Bitola", font=("Arial", 7, "bold"))
        self.canvas_etiq.create_text(tab_x + cw1 + cw2//2, tab_y1 + th//2, text="Compr. Unit.", font=("Arial", 7, "bold"))
        self.canvas_etiq.create_text(tab_x + cw1 + cw2 + cw3//2, tab_y1 + th//2, text="Peso", font=("Arial", 7, "bold"))
        self.canvas_etiq.create_text(tab_x + cw1 + cw2 + cw3 + cw4//2, tab_y1 + th//2, text="Qtde", font=("Arial", 7, "bold"))
        
        # Valores editáveis (com tags clicáveis)
        tag_bitola = f"edit_bitola_{pilar}_{pos}"
        tag_comp = f"edit_comp_{pilar}_{pos}"
        tag_qtde = f"edit_qtde_{pilar}_{pos}"
        
        # Usar valores customizados se existirem
        chave = (pilar, pos)
        if chave in self.medidas_customizadas:
            bitola = self.medidas_customizadas[chave].get('bitola', bitola)
            comp = self.medidas_customizadas[chave].get('comp', comp)
            qtde = self.medidas_customizadas[chave].get('qtde', qtde)
        
        txt_bitola = self.canvas_etiq.create_text(tab_x + cw1//2, tab_y1 + th + tr//2, text=f"{bitola:.2f}", font=("Arial", 8, "underline"), fill="#0066cc", tags=(tag_bitola,))
        txt_comp = self.canvas_etiq.create_text(tab_x + cw1 + cw2//2, tab_y1 + th + tr//2, text=f"{comp:.3f}", font=("Arial", 8, "underline"), fill="#0066cc", tags=(tag_comp,))
        
        peso_val = 0.0
        try:
            from core.peso import peso_linear_kg_m
            peso_val = peso_linear_kg_m(float(bitola)) * float(comp) * float(qtde)
        except Exception as e:
            print(f"[peso] Falha ao calcular: {e}")
        self.canvas_etiq.create_text(tab_x + cw1 + cw2 + cw3//2, tab_y1 + th + tr//2, text=f"{peso_val:.2f}", font=("Arial", 8))
        
        txt_qtde = self.canvas_etiq.create_text(tab_x + cw1 + cw2 + cw3 + cw4//2, tab_y1 + th + tr//2, text=f"{qtde}", font=("Arial", 8, "underline"), fill="#0066cc", tags=(tag_qtde,))
        
        # Bindings para editar ao clicar
        self.canvas_etiq.tag_bind(txt_bitola, "<Button-1>", lambda e, v=pilar, p=pos, tipo="bitola": self._editar_medida_etiqueta(v, p, tipo))
        self.canvas_etiq.tag_bind(txt_comp, "<Button-1>", lambda e, v=pilar, p=pos, tipo="comp": self._editar_medida_etiqueta(v, p, tipo))
        self.canvas_etiq.tag_bind(txt_qtde, "<Button-1>", lambda e, v=pilar, p=pos, tipo="qtde": self._editar_medida_etiqueta(v, p, tipo))


        # Desenho sempre renderiza (não depende de ETIQUETAS_HELPER_DISPONIVEL)
        try:
            # Área disponível para o desenho dentro do topo (sem sobrepor textos/tabela)
            margem_topo = mm(3)  # AUMENTADA
            margem_lateral = mm(3)  # REDUZIDA
            draw_area_x1 = area_util_x1 + margem_lateral
            draw_area_x2 = area_util_x2 - margem_lateral
            draw_area_y1 = tab_y2 + margem_topo
            draw_area_y2 = y + h - margem_topo
            
            print(f"DEBUG: area_util_x1={area_util_x1}, area_util_x2={area_util_x2}, tab_y2={tab_y2}, y+h={y+h}")
            print(f"DEBUG: draw_area_x1={draw_area_x1}, draw_area_x2={draw_area_x2}, draw_area_y1={draw_area_y1}, draw_area_y2={draw_area_y2}")
            
            if draw_area_x2 <= draw_area_x1 or draw_area_y2 <= draw_area_y1:
                print(f"DEBUG: Area insuficiente, retornando")
                return

            avail_w = draw_area_x2 - draw_area_x1
            avail_h = draw_area_y2 - draw_area_y1

            # Dimensões preferidas e ajuste para caber no box - DESENHO COLOSSAL
            pref_w = mm(DESENHO_LARGURA_MM) if ETIQUETAS_LAYOUT_CFG else mm(60)  # REDUZIDO PARA CABER AMPLIADO
            pref_h = mm(DESENHO_ALTURA_MM) if ETIQUETAS_LAYOUT_CFG else mm(50)  # REDUZIDO PARA CABER AMPLIADO
            escala = min(avail_w / max(1, pref_w), avail_h / max(1, pref_h), 6.0)  # AUMENTADO PARA 6.0
            dw = max(10, int(pref_w * escala))
            dh = max(10, int(pref_h * escala))
            dx = draw_area_x1 + (avail_w - dw) // 2
            dy = draw_area_y1 + (avail_h - dh) // 2
            
            # Verificar se há desenho customizado
            chave = (pilar, pos)
            tag = f"desenho_{pilar}_{pos}"
            
            # Área do desenho (clicável para editar medidas)
            rect_id = self.canvas_etiq.create_rectangle(
                dx, dy, dx + dw, dy + dh,
                outline="#e11d48", width=2, fill="white", tags=(tag,)
            )
            # Bind na tag para capturar cliques
            self.canvas_etiq.tag_bind(tag, "<Button-1>", lambda e, v=pilar, p=pos: self._editar_desenho_canvas(v, p))
            
            print(f"DEBUG: Desenhando {pilar}/{pos}, dx={dx}, dy={dy}, dw={dw}, dh={dh}")
            
            # Mostrar forma customizada, PNG técnico, ou detectar automaticamente
            try:
                forma_raw = self.formas_customizadas.get(chave)
                forma = None
                
                # Se houver forma customizada, normalizar para minúsculas sem espaço
                if forma_raw:
                    forma_lower = forma_raw.lower()
                    # Priorizar 'redondo' antes de verificar 'estribo' para não cair no quadrado
                    if 'redondo' in forma_lower or 'red.' in forma_lower:
                        forma = 'estribo_redondo'
                    elif 'estribo' in forma_lower or 'quadrado' in forma_lower:
                        forma = 'estribo'
                    elif 'dobra' in forma_lower and 'dupla' in forma_lower:
                        forma = 'dobra_dupla'
                    elif 'dobra' in forma_lower:
                        # Se tiver 'dobra' mas não 'dupla', é dobra simples (única)
                        forma = 'dobra'
                    elif 'gancho' in forma_lower:
                        forma = 'gancho'
                    else:
                        forma = 'reta'
                    print(f"✓ Forma customizada encontrada: '{forma_raw}' → normalizada: '{forma}'")
                
                # Se não houver forma customizada, detectar automaticamente
                if not forma:
                    tipo_barra = AnalisadorGeometricoPilares.identificar_tipo_pilar(pos, bitola, comp)
                    # Mapear tipo para forma simplificada
                    if 'ESTRIBO' in tipo_barra.upper():
                        forma = 'estribo'
                    elif 'GANCHO' in tipo_barra.upper() or 'POSITIVO' in tipo_barra.upper():
                        forma = 'gancho'
                    else:
                        forma = 'reta'
                    print(f"✓ Detectado {pilar}/{pos}: tipo={tipo_barra} → forma={forma}")
                
                # Desenhar forma detectada ou customizada (sempre desenhar)
                medida_dobra = None
                medidas_gancho = None
                
                # Tentar obter medidas do dicionário de customizações
                if chave in self.medidas_customizadas:
                    medidas_dict = self.medidas_customizadas[chave]
                    print(f"✓ MEDIDAS ENCONTRADAS para {chave}: {medidas_dict}")
                    
                    if forma == 'dobra':
                        # Para dobra simples, usar medida_dobra (dobra) e medida_dobra_2 (corpo)
                        m_dobra = medidas_dict.get('medida_dobra', 0.0)
                        m_corpo = medidas_dict.get('medida_dobra_2', 0.0)
                        medida_dobra = (m_dobra, m_corpo)
                        if m_dobra > 0 or m_corpo > 0:
                            print(f"✓ Dobra CARREGADA {pilar}/{pos}: dobra={m_dobra}cm, corpo={m_corpo}cm")
                        else:
                            print(f"⚠ Dobra {pilar}/{pos}: medida=0.0cm (VAZIA)")
                    elif forma == 'dobra_dupla':
                        # Para dobra dupla, usar medida_dobra (dobra1), medida_dobra_2 (dobra2) e lado1 (corpo)
                        m1 = medidas_dict.get('medida_dobra', 0.0)
                        m2 = medidas_dict.get('medida_dobra_2', 0.0)
                        m_corpo = medidas_dict.get('lado1', 0.0)
                        medida_dobra = (m1, m2, m_corpo)
                        if m1 > 0 or m2 > 0 or m_corpo > 0:
                            print(f"✓ Dobra Dupla {pilar}/{pos}: dobra1={m1}cm, dobra2={m2}cm, corpo={m_corpo}cm")
                    elif forma == 'gancho':
                        # Para gancho, usar lado1 e lado2
                        lado1 = medidas_dict.get('lado1', 0.0)
                        lado2 = medidas_dict.get('lado2', 0.0)
                        if lado1 > 0 or lado2 > 0:
                            medidas_gancho = (lado1, lado2)
                            print(f"✓ Gancho {pilar}/{pos}: medidas={lado1}cm, {lado2}cm")
                    elif forma == 'estribo':
                        # Para estribo, armazenar todos os lados
                        va = medidas_dict.get('lado1', 0.0)
                        vb = medidas_dict.get('lado2', 0.0)
                        vc = medidas_dict.get('lado3', 0.0)
                        vd = medidas_dict.get('lado4', 0.0)
                        if va > 0 or vb > 0 or vc > 0 or vd > 0:
                            # Armazenar em atributo para usar depois
                            if not hasattr(self, 'estribo_lados'):
                                self.estribo_lados = {}
                            self.estribo_lados[(pilar, pos)] = (va, vb, vc, vd)
                            print(f"✓ Estribo {pilar}/{pos}: lados={va}cm, {vb}cm, {vc}cm, {vd}cm")
                    elif forma == 'estribo_redondo':
                        # Para estribo redondo, usar raio
                        raio = medidas_dict.get('raio', 0.0)
                        if raio > 0:
                            medida_dobra = raio
                            print(f"✓ Estribo Redondo {pilar}/{pos}: raio={raio}cm (Ø={2*raio:.0f}cm)")
                else:
                    print(f"⚠ MEDIDAS NÃO ENCONTRADAS para {chave}")
                    print(f"  Chaves disponíveis: {list(self.medidas_customizadas.keys())}")
                
                # Fallback para dicionários legados (dobra_medidas, dobra_dupla_medidas, gancho_medidas, estribo_lados)
                if not medida_dobra and forma == 'dobra' and hasattr(self, 'dobra_medidas') and chave in self.dobra_medidas:
                    medida_dobra = self.dobra_medidas.get(chave)
                    print(f"✓ Dobra carregada do fallback (dobra_medidas): {medida_dobra}cm")
                if not medida_dobra and forma == 'dobra_dupla' and hasattr(self, 'dobra_dupla_medidas') and chave in self.dobra_dupla_medidas:
                    # Carregar dobra_dupla do dicionário legado (já é tupla)
                    medida_dobra = self.dobra_dupla_medidas.get(chave)
                    print(f"✓ Dobra dupla carregada do fallback (dobra_dupla_medidas): {medida_dobra}cm")
                if not medidas_gancho and forma == 'gancho' and hasattr(self, 'gancho_medidas') and chave in self.gancho_medidas:
                    medidas_gancho = self.gancho_medidas.get(chave)
                
                # Obter lados do estribo se houver
                estribo_lados = None
                if forma == 'estribo' and hasattr(self, 'estribo_lados') and chave in self.estribo_lados:
                    estribo_lados = self.estribo_lados[chave]
                    print(f"✓ Passando estribo_lados para desenho: {estribo_lados}")
                
                self._desenhar_forma_simplificada(self.canvas_etiq, dx, dy, dw, dh, forma, medida_dobra, medidas_gancho, estribo_lados, comp)
                
                # Se for estribo e houver lados definidos, anotar levemente
                if forma == 'estribo' and hasattr(self, 'estribo_lados') and (pilar, pos) in self.estribo_lados:
                    va, vb, vc, vd = self.estribo_lados[(pilar, pos)]
                    self.canvas_etiq.create_text(dx + dw/2, dy - 8, text=f"A={va:.0f}", font=("Arial",7), fill="#475569")
                    self.canvas_etiq.create_text(dx + dw/2, dy + dh + 8, text=f"C={vc:.0f}", font=("Arial",7), fill="#475569")
                    self.canvas_etiq.create_text(dx - 14, dy + dh/2, text=f"B={vb:.0f}", font=("Arial",7), fill="#475569")
                    self.canvas_etiq.create_text(dx + dw + 14, dy + dh/2, text=f"D={vd:.0f}", font=("Arial",7), fill="#475569")
                else:
                    # Manter apenas a área demarcada
                    pass
            except Exception:
                # Em erro, não desenhar placeholder
                pass
        except Exception as e:
            print(f"Erro ao desenhar etiqueta: {e}")

        self.canvas_etiq.create_text(area_util_x2, y + h - 12, text=f"Página {self.pagina_atual + 1} de {self.total_paginas}", font=("Arial", 8), anchor="e")

    def _desenhar_medidas_resumo_fase4(self, x, y, w, h, pilar, pos, bitola, qtde, comp):
        """Bloco compacto com medidas principais (SEM desenho; desenho fica apenas no quadrado do topo)."""
        # Usar valores customizados se existirem
        chave = (pilar, pos)
        if chave in self.medidas_customizadas:
            bitola = float(self.medidas_customizadas[chave].get('bitola', bitola))
            comp = float(self.medidas_customizadas[chave].get('comp', comp))
            qtde = int(self.medidas_customizadas[chave].get('qtde', qtde))

        # Moldura e faixa de confiança estilo V3
        self.canvas_etiq.create_rectangle(x, y, x + w, y + h, outline="#0f172a", width=1, fill="#f8fafc")
        faixa_w = 8
        self.canvas_etiq.create_rectangle(x, y, x + faixa_w, y + h, outline="", fill="#22c55e")  # verde (confiança alta)

        # Layout: esquerda medidas, direita comprimento
        left_x = x + faixa_w + 10
        right_x = x + w - 10

        # Tags para edição
        tag_bitola = f"med_blk_bitola_{pilar}_{pos}"
        tag_comp = f"med_blk_comp_{pilar}_{pos}"
        tag_qtde = f"med_blk_qtde_{pilar}_{pos}"

        # Linha superior: quantidade x Øbitola
        txt_qtd_bitola = self.canvas_etiq.create_text(
            left_x, y + h // 2 - 8,
            text=f"{qtde}x Ø{bitola:.1f}",
            font=("Arial", 14, "bold"),
            fill="#0f172a",
            anchor="w",
            tags=(tag_bitola, tag_qtde)
        )

        # Linha direita: comprimento
        txt_comp = self.canvas_etiq.create_text(
            right_x, y + h // 2 - 8,
            text=f"C={comp:.2f}m",
            font=("Arial", 14, "bold"),
            fill="#dc2626",
            anchor="e",
            tags=(tag_comp,)
        )

        # Linha inferior: PILAR e POS
        self.canvas_etiq.create_text(
            left_x, y + h - 6,
            text=f"PILAR: {pilar}",
            font=("Arial", 9, "bold"),
            fill="#64748b",
            anchor="sw"
        )
        self.canvas_etiq.create_text(
            right_x, y + h - 6,
            text=f"POS: {pos}",
            font=("Arial", 9, "bold"),
            fill="#64748b",
            anchor="se"
        )

        # Marca inferior direita estilo "EngenhariaPlanPro V3"
        self.canvas_etiq.create_text(
            right_x, y + h - 2,
            text="EngenhariaPlanPro V3",
            font=("Arial", 8),
            fill="#94a3b8",
            anchor="se"
        )

    def _desenhar_moldura_etiqueta(self, x, y, w, h):
        """Desenha moldura com marcas de corte (padrao vigas)."""
        self.canvas_etiq.create_rectangle(x, y, x + w, y + h, outline="#ff6f00", width=2, fill="white")
        tamanho_marca = 5
        cantos = [(x, y), (x + w, y), (x, y + h), (x + w, y + h)]
        for px, py in cantos:
            self.canvas_etiq.create_line(px - tamanho_marca, py, px + tamanho_marca, py, width=1, fill="black")
            self.canvas_etiq.create_line(px, py - tamanho_marca, px, py + tamanho_marca, width=1, fill="black")

    def _desenhar_topo_etiqueta(self, x, y, w, h, pilar, pos, bitola, qtde, comp):
        """Desenha o topo (93mm) com informacoes tecnicas (padrao vigas)."""
        pxmm = 4

        def mm(v):
            return int(round(v * pxmm))

        faixa_larg = mm(10)
        os_w = mm(18)
        os_h = mm(30)

        area_os_x1 = x + w - os_w - faixa_larg
        area_os_y1 = y
        area_os_x2 = x + w - faixa_larg
        area_os_y2 = y + os_h
        self.canvas_etiq.create_rectangle(area_os_x1, area_os_y1, area_os_x2, area_os_y2, outline="#000", width=1)
        self.canvas_etiq.create_text(area_os_x1 + 6, area_os_y1 + 6, text="OS", font=("Arial", 10, "bold"), anchor="nw")

        os_txt = f"{self.pagina_atual + 1}-{self.total_paginas}" if hasattr(self, 'total_paginas') else "-"
        linhas_os = os_txt.split("-")
        start_y = area_os_y1 + os_h // 2 - mm(6)
        espaco_linha = mm(8)
        for i, linha in enumerate(linhas_os):
            self.canvas_etiq.create_text((area_os_x1 + area_os_x2) // 2, start_y + i * espaco_linha,
                                         text=linha, font=("Arial", 11, "bold"), anchor="center")

        faixa_x1 = x + w - faixa_larg
        faixa_x2 = x + w
        self.canvas_etiq.create_rectangle(faixa_x1, y, faixa_x2, y + h, outline="#000", width=1)

        obra_nome = self.var_obra.get() if hasattr(self, 'var_obra') else "OBRA.001"
        if Image is None or ImageTk is None:
            self.canvas_etiq.create_text(faixa_x1 + faixa_larg // 2, y + h // 2, text=obra_nome, angle=90)
        else:
            try:
                from PIL import ImageDraw, ImageFont
                img_tmp = Image.new('RGBA', (mm(60), faixa_larg), (255, 255, 255, 0))
                draw = ImageDraw.Draw(img_tmp)
                try:
                    fnt = ImageFont.truetype("arial.ttf", 12)
                except Exception:
                    fnt = ImageFont.load_default()
                tw, th = draw.textsize(obra_nome, font=fnt)
                text_x = (img_tmp.width - tw) // 2
                text_y = (img_tmp.height - th) // 2
                draw.text((text_x, text_y), obra_nome, fill=(0, 0, 0, 255), font=fnt)
                img_tmp = img_tmp.rotate(90, expand=True)
                photo = ImageTk.PhotoImage(img_tmp)
                if not hasattr(self, '_desenho_images'):
                    self._desenho_images = []
                self._desenho_images.append(photo)
                self.canvas_etiq.create_image(faixa_x1 + (faixa_larg // 2), y + h // 2, image=photo)
            except Exception:
                self.canvas_etiq.create_text(faixa_x1 + faixa_larg // 2, y + h // 2, text=obra_nome, angle=90)

        area_util_x1 = x + 6
        area_util_x2 = area_os_x1 - 6
        area_util_w = max(0, area_util_x2 - area_util_x1)

        def _truncate_text(text, max_width, font):
            try:
                import tkinter.font as tkfont
                f = tkfont.Font(font=font)
                if f.measure(text) <= max_width:
                    return text
                ell = "..."
                max_w = max_width - f.measure(ell)
                if max_w <= 0:
                    return ell
                while text and f.measure(text) > max_w:
                    text = text[:-1]
                return text + ell
            except Exception:
                return text

        step_y = mm(8)
        y_current = y + mm(12)

        self.canvas_etiq.create_text(area_util_x1, y_current, text="Sigla/Obra", font=("Arial", 8), anchor="nw")
        self.canvas_etiq.create_text(area_util_x1 + mm(24), y_current, text=obra_nome.replace(" ", " - "), font=("Arial", 11, "bold"), anchor="nw")
        y_current += step_y

        self.canvas_etiq.create_text(area_util_x1, y_current, text="Desenho", font=("Arial", 8), anchor="nw")
        self.canvas_etiq.create_text(area_util_x1 + mm(24), y_current, text="PILAR", font=("Arial", 10, "bold"), anchor="nw", fill="#e11d48")
        y_current += step_y

        self.canvas_etiq.create_text(area_util_x1, y_current, text="Pavimento", font=("Arial", 8), anchor="nw")
        pavimento = self.var_pavimento.get() if hasattr(self, 'var_pavimento') else "TÉRREO"
        self.canvas_etiq.create_text(area_util_x1 + mm(24), y_current, text=pavimento, font=("Arial", 10, "bold"), anchor="nw")
        y_current += step_y

        self.canvas_etiq.create_text(area_util_x1, y_current, text="Elemento", font=("Arial", 8), anchor="nw")
        pos_block_w = mm(24)
        elem_x = area_util_x1 + mm(24)
        elem_max_w = max(0, (area_util_x2 - pos_block_w - mm(4)) - elem_x)
        elemento_txt = _truncate_text(f"{pilar}", elem_max_w, ("Arial", 10, "bold"))
        self.canvas_etiq.create_text(elem_x, y_current, text=elemento_txt, font=("Arial", 10, "bold"), anchor="nw")
        self.canvas_etiq.create_text(area_util_x2 - mm(18), y_current, text="POS", font=("Arial", 8), anchor="e")
        self.canvas_etiq.create_text(area_util_x2 - mm(2), y_current, text=f"{pos}", font=("Arial", 12, "bold"), anchor="e")

        th = mm(8)
        tr = mm(10)
        tab_y1 = y + mm(45)
        tab_y2 = tab_y1 + th + tr
        tab_x = area_util_x1
        cw1 = mm(16)
        cw2 = mm(34)
        cw3 = mm(22)
        cw4 = mm(18)
        tot = cw1 + cw2 + cw3 + cw4
        if tot > area_util_w and area_util_w > 0:
            escala = area_util_w / tot
            cw1 = int(cw1 * escala); cw2 = int(cw2 * escala); cw3 = int(cw3 * escala); cw4 = int(cw4 * escala)

        self.canvas_etiq.create_rectangle(tab_x, tab_y1, tab_x + cw1 + cw2 + cw3 + cw4, tab_y2, outline="#000", width=1)
        self.canvas_etiq.create_line(tab_x + cw1, tab_y1, tab_x + cw1, tab_y2)
        self.canvas_etiq.create_line(tab_x + cw1 + cw2, tab_y1, tab_x + cw1 + cw2, tab_y2)
        self.canvas_etiq.create_line(tab_x + cw1 + cw2 + cw3, tab_y1, tab_x + cw1 + cw2 + cw3, tab_y2)
        self.canvas_etiq.create_line(tab_x, tab_y1 + th, tab_x + cw1 + cw2 + cw3 + cw4, tab_y1 + th)

        self.canvas_etiq.create_text(tab_x + cw1 // 2, tab_y1 + th // 2, text="Bitola", font=("Arial", 7, "bold"))
        self.canvas_etiq.create_text(tab_x + cw1 + cw2 // 2, tab_y1 + th // 2, text="Compr. Unit.", font=("Arial", 7, "bold"))
        self.canvas_etiq.create_text(tab_x + cw1 + cw2 + cw3 // 2, tab_y1 + th // 2, text="Peso", font=("Arial", 7, "bold"))
        self.canvas_etiq.create_text(tab_x + cw1 + cw2 + cw3 + cw4 // 2, tab_y1 + th // 2, text="Qtde", font=("Arial", 7, "bold"))

        chave = (pilar, pos)
        if hasattr(self, 'medidas_customizadas') and chave in self.medidas_customizadas:
            bitola = self.medidas_customizadas[chave].get('bitola', bitola)
            comp = self.medidas_customizadas[chave].get('comp', comp)
            qtde = self.medidas_customizadas[chave].get('qtde', qtde)

        peso_val = 0.0
        try:
            from core.peso import peso_linear_kg_m
            peso_val = peso_linear_kg_m(float(bitola)) * float(comp) * float(qtde)
        except Exception:
            pass

        tag_bitola = f"edit_bitola_{pilar}_{pos}"
        tag_comp = f"edit_comp_{pilar}_{pos}"
        tag_qtde = f"edit_qtde_{pilar}_{pos}"

        txt_bitola = self.canvas_etiq.create_text(tab_x + cw1 // 2, tab_y1 + th + tr // 2, text=f"{bitola:.2f}",
                                                  font=("Arial", 8, "underline"), fill="#0066cc", tags=(tag_bitola,))
        txt_comp = self.canvas_etiq.create_text(tab_x + cw1 + cw2 // 2, tab_y1 + th + tr // 2, text=f"{comp:.3f}",
                                                font=("Arial", 8, "underline"), fill="#0066cc", tags=(tag_comp,))
        self.canvas_etiq.create_text(tab_x + cw1 + cw2 + cw3 // 2, tab_y1 + th + tr // 2, text=f"{peso_val:.2f}", font=("Arial", 8))
        txt_qtde = self.canvas_etiq.create_text(tab_x + cw1 + cw2 + cw3 + cw4 // 2, tab_y1 + th + tr // 2, text=f"{qtde}",
                                                font=("Arial", 8, "underline"), fill="#0066cc", tags=(tag_qtde,))

        self.canvas_etiq.tag_bind(txt_bitola, "<Button-1>", lambda e, v=pilar, p=pos, tipo="bitola": self._editar_medida_etiqueta(v, p, tipo))
        self.canvas_etiq.tag_bind(txt_comp, "<Button-1>", lambda e, v=pilar, p=pos, tipo="comp": self._editar_medida_etiqueta(v, p, tipo))
        self.canvas_etiq.tag_bind(txt_qtde, "<Button-1>", lambda e, v=pilar, p=pos, tipo="qtde": self._editar_medida_etiqueta(v, p, tipo))

        draw_area_x1 = area_util_x1 + mm(3)
        draw_area_x2 = area_util_x2 - mm(3)
        draw_area_y1 = tab_y2 + mm(3)
        draw_area_y2 = y + h - mm(3)
        if draw_area_x2 > draw_area_x1 and draw_area_y2 > draw_area_y1:
            tag = f"desenho_{pilar}_{pos}"
            self.canvas_etiq.create_rectangle(
                draw_area_x1,
                draw_area_y1,
                draw_area_x2,
                draw_area_y2,
                outline="#e11d48",
                width=2,
                fill="white",
                tags=(tag,)
            )
            self.canvas_etiq.tag_bind(tag, "<Button-1>", lambda e, v=pilar, p=pos: self._editar_desenho_canvas(v, p))
            forma, medida_dobra, medidas_gancho, estribo_lados = self._resolver_forma_medidas_pilar(pilar, pos, bitola, comp)
            self._desenhar_forma_simplificada(
                self.canvas_etiq,
                draw_area_x1,
                draw_area_y1,
                draw_area_x2 - draw_area_x1,
                draw_area_y2 - draw_area_y1,
                forma,
                medida_dobra,
                medidas_gancho,
                estribo_lados,
                comp
            )

    def _desenhar_secao_micro(self, x, y, w, h, pilar, pos, bitola, qtde, comp):
        """Desenha secao micro (picote) no padrao vigas."""
        obra_nome = self.var_obra.get() if hasattr(self, 'var_obra') else "OBRA 001"
        texto = f"{obra_nome} | {pilar} | {pos} | ø{bitola:.1f} | {qtde}x{comp:.2f}m"
        self.canvas_etiq.create_text(x + w // 2, y + h // 2, text=texto, font=("Arial", 8, "bold"), fill="#000")

    def _desenhar_picote(self, x, y, w):
        """Desenha linha tracejada de picote."""
        self.canvas_etiq.create_line(x, y, x + w, y, dash=(4, 4), fill="#999")

    def _desenhar_forma_simplificada(self, canvas, x, y, w, h, forma, medida_dobra=None, medidas_gancho=None, estribo_lados=None, comprimento=None):
        """Desenha forma simples (reta/gancho/estribo/dobra/estribo redondo)"""
        if medida_dobra is not None or medidas_gancho is not None or estribo_lados is not None:
            print(f"[DESENHO FORMA] Forma: {forma}, medida_dobra={medida_dobra}, medidas_gancho={medidas_gancho}, estribo_lados={estribo_lados}")
        
        cx, cy = x + w // 2, y + h // 2
        ctx_w, ctx_h = w, h
        cor = "#1e293b"
        esp = 4
        if forma == 'estribo':
            rx = int(ctx_w * 0.3); ry = int(ctx_h * 0.2)
            rw = int(ctx_w * 0.4); rh = int(ctx_h * 0.6)
            canvas.create_rectangle(x + rx, y + ry, x + rx + rw, y + ry + rh, outline=cor, width=esp)
            
            # Adicionar medidas dos lados se disponíveis
            if estribo_lados is not None:
                try:
                    va, vb, vc, vd = estribo_lados
                    # Lado superior (A)
                    if va > 0:
                        canvas.create_text(x + rx + rw // 2, y + ry - 8, text=f"A={va:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="s")
                    # Lado direito (B)
                    if vb > 0:
                        canvas.create_text(x + rx + rw + 8, y + ry + rh // 2, text=f"B={vb:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="w")
                    # Lado inferior (C)
                    if vc > 0:
                        canvas.create_text(x + rx + rw // 2, y + ry + rh + 8, text=f"C={vc:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="n")
                    # Lado esquerdo (D)
                    if vd > 0:
                        canvas.create_text(x + rx - 8, y + ry + rh // 2, text=f"D={vd:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="e")
                except Exception as e:
                    print(f"Erro ao desenhar medidas do estribo: {e}")
        elif forma == 'estribo_redondo':
            r = int(min(ctx_w, ctx_h) * 0.3)
            canvas.create_oval(cx - r, cy - r, cx + r, cy + r, outline=cor, width=esp)
            # Exibir raio/diâmetro se disponível via medida_dobra (raio)
            if medida_dobra is not None and isinstance(medida_dobra, (int, float)) and medida_dobra > 0:
                try:
                    diam = float(medida_dobra) * 2.0
                    canvas.create_text(cx, cy + r + 10, text=f"Ø={diam:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="n")
                except Exception:
                    pass
        elif forma == 'gancho':
            x1 = x + int(ctx_w * 0.1); y1 = y + int(ctx_h * 0.6)
            x2 = x + int(ctx_w * 0.9); y2 = y + int(ctx_h * 0.6)
            yb = y + int(ctx_h * 0.35)
            canvas.create_line(x1, y1, x1, yb, width=esp, fill=cor)
            canvas.create_line(x1, yb, x2, yb, width=esp, fill=cor)
            canvas.create_line(x2, yb, x2, y2, width=esp, fill=cor)
            if medidas_gancho is not None:
                try:
                    ga, gb = medidas_gancho
                    canvas.create_text(x1 - 8, (y1 + yb)//2, text=f"A={ga:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="e")
                    canvas.create_text(x2 + 8, (y2 + yb)//2, text=f"B={gb:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="w")
                except Exception:
                    pass
        elif forma == 'dobra':
            x1 = x + int(ctx_w * 0.15); y1 = y + int(ctx_h * 0.55)
            x2 = x + int(ctx_w * 0.75); y2 = y + int(ctx_h * 0.55)
            yb = y + int(ctx_h * 0.25)
            canvas.create_line(x1, y1, x2, y1, width=esp, fill=cor)
            canvas.create_line(x2, y1, x2, yb, width=esp, fill=cor)
            if medida_dobra is not None:
                # Tratar caso de tupla (dobra, corpo) ou float
                if isinstance(medida_dobra, (list, tuple)) and len(medida_dobra) >= 2:
                    # Dobra + Corpo
                    m_dobra, m_corpo = medida_dobra[0], medida_dobra[1]
                    canvas.create_text(x2 + 8, (y1 + yb)//2, text=f"{m_dobra:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="w")
                    canvas.create_text((x1 + x2)//2, y1 + 8, text=f"{m_corpo:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="n")
                else:
                    # Compatibilidade: só dobra
                    txt = f"{medida_dobra:.0f}cm"
                    canvas.create_text(x2 + 8, (y1 + yb)//2, text=txt, font=("Arial", 28, "bold"), fill="#000000", anchor="w")
        elif forma == 'dobra_dupla':
            # Duas dobras: uma em cada extremidade (formato tipo "U" com perninhas)
            x_esq = x + int(ctx_w * 0.15)
            x_dir = x + int(ctx_w * 0.85)
            y_base = y + int(ctx_h * 0.60)
            y_top = y + int(ctx_h * 0.25)

            # Barra horizontal principal
            canvas.create_line(x_esq, y_base, x_dir, y_base, width=esp, fill=cor)
            # Pernas (dobras) nas extremidades
            canvas.create_line(x_esq, y_base, x_esq, y_top, width=esp, fill=cor)
            canvas.create_line(x_dir, y_base, x_dir, y_top, width=esp, fill=cor)

            # Medidas: cada perna + corpo horizontal
            if isinstance(medida_dobra, (list, tuple)):
                if len(medida_dobra) >= 3:
                    m_esq, m_dir, m_corpo = medida_dobra[0], medida_dobra[1], medida_dobra[2]
                    canvas.create_text(x_esq - 8, (y_base + y_top)//2, text=f"{m_esq:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="e")
                    canvas.create_text(x_dir + 8, (y_base + y_top)//2, text=f"{m_dir:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="w")
                    if m_corpo > 0:
                        canvas.create_text((x_esq + x_dir)//2, y_base + 8, text=f"{m_corpo:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="n")
                elif len(medida_dobra) >= 2:
                    m_esq, m_dir = medida_dobra[0], medida_dobra[1]
                    canvas.create_text(x_esq - 8, (y_base + y_top)//2, text=f"{m_esq:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="e")
                    canvas.create_text(x_dir + 8, (y_base + y_top)//2, text=f"{m_dir:.0f}cm", font=("Arial", 28, "bold"), fill="#000000", anchor="w")
        else:
            x1 = x + int(ctx_w * 0.1); x2 = x + int(ctx_w * 0.9)
            yb = y + int(ctx_h * 0.5)
            canvas.create_line(x1, yb, x2, yb, width=esp, fill=cor)
            # Adicionar medida do comprimento na reta
            if comprimento and comprimento > 0:
                canvas.create_text((x1 + x2)//2, yb - 12, text=f"{comprimento:.2f}m", font=("Arial", 28, "bold"), fill="#000000", anchor="s")

        # Interações específicas ficam em outros blocos; não criar binds aqui para evitar NameError

    def _desenhar_forma_pil(self, draw, x, y, w, h, forma, medida_dobra=None, medidas_gancho=None, estribo_lados=None, font=None, comprimento=None):
        """Versão PIL da função _desenhar_forma_simplificada - IDÊNTICA ao canvas"""
        cx, cy = x + w // 2, y + h // 2
        cor = (30, 41, 59)  # #1e293b
        esp = 4
        
        if forma == 'estribo':
            rx = int(w * 0.3); ry = int(h * 0.2)
            rw = int(w * 0.4); rh = int(h * 0.6)
            draw.rectangle([x + rx, y + ry, x + rx + rw, y + ry + rh], outline=cor, width=esp)
            
            if estribo_lados is not None:
                try:
                    va, vb, vc, vd = estribo_lados
                    if va > 0:
                        draw.text((x + rx + rw // 2, y + ry - 8), f"A={va:.0f}cm", fill=(0,0,0), font=font, anchor='ms')
                    if vb > 0:
                        draw.text((x + rx + rw + 8, y + ry + rh // 2), f"B={vb:.0f}cm", fill=(0,0,0), font=font, anchor='lm')
                    if vc > 0:
                        draw.text((x + rx + rw // 2, y + ry + rh + 8), f"C={vc:.0f}cm", fill=(0,0,0), font=font, anchor='mt')
                    if vd > 0:
                        draw.text((x + rx - 8, y + ry + rh // 2), f"D={vd:.0f}cm", fill=(0,0,0), font=font, anchor='rm')
                except Exception:
                    pass
        elif forma == 'estribo_redondo':
            r = int(min(w, h) * 0.3)
            draw.ellipse([cx - r, cy - r, cx + r, cy + r], outline=cor, width=esp)
            if medida_dobra is not None and isinstance(medida_dobra, (int, float)) and medida_dobra > 0:
                try:
                    diam = float(medida_dobra) * 2.0
                    draw.text((cx, cy + r + 10), f"Ø={diam:.0f}cm", fill=(0,0,0), font=font, anchor='mt')
                except Exception:
                    pass
        elif forma == 'gancho':
            x1 = x + int(w * 0.1); y1 = y + int(h * 0.6)
            x2 = x + int(w * 0.9); y2 = y + int(h * 0.6)
            yb = y + int(h * 0.35)
            draw.line([x1, y1, x1, yb], fill=cor, width=esp)
            draw.line([x1, yb, x2, yb], fill=cor, width=esp)
            draw.line([x2, yb, x2, y2], fill=cor, width=esp)
            if medidas_gancho is not None:
                try:
                    ga, gb = medidas_gancho
                    draw.text((x1 - 8, (y1 + yb)//2), f"A={ga:.0f}cm", fill=(0,0,0), font=font, anchor='rm')
                    draw.text((x2 + 8, (y2 + yb)//2), f"B={gb:.0f}cm", fill=(0,0,0), font=font, anchor='lm')
                except Exception:
                    pass
        elif forma == 'dobra':
            x1 = x + int(w * 0.15); y1 = y + int(h * 0.55)
            x2 = x + int(w * 0.75); y2 = y + int(h * 0.55)
            yb = y + int(h * 0.25)
            draw.line([x1, y1, x2, y1], fill=cor, width=esp)
            draw.line([x2, y1, x2, yb], fill=cor, width=esp)
            if medida_dobra is not None:
                if isinstance(medida_dobra, (list, tuple)) and len(medida_dobra) >= 2:
                    m_dobra, m_corpo = medida_dobra[0], medida_dobra[1]
                    draw.text((x2 + 8, (y1 + yb)//2), f"{m_dobra:.0f}cm", fill=(0,0,0), font=font, anchor='lm')
                    draw.text(((x1 + x2)//2, y1 + 8), f"{m_corpo:.0f}cm", fill=(0,0,0), font=font, anchor='mt')
                else:
                    txt = f"{medida_dobra:.0f}cm"
                    draw.text((x2 + 8, (y1 + yb)//2), txt, fill=(0,0,0), font=font, anchor='lm')
        elif forma == 'dobra_dupla':
            x_esq = x + int(w * 0.15)
            x_dir = x + int(w * 0.85)
            y_base = y + int(h * 0.60)
            y_top = y + int(h * 0.25)
            draw.line([x_esq, y_base, x_dir, y_base], fill=cor, width=esp)
            draw.line([x_esq, y_base, x_esq, y_top], fill=cor, width=esp)
            draw.line([x_dir, y_base, x_dir, y_top], fill=cor, width=esp)
            if isinstance(medida_dobra, (list, tuple)):
                if len(medida_dobra) >= 3:
                    m_esq, m_dir, m_corpo = medida_dobra[0], medida_dobra[1], medida_dobra[2]
                    draw.text((x_esq - 8, (y_base + y_top)//2), f"{m_esq:.0f}cm", fill=(0,0,0), font=font, anchor='rm')
                    draw.text((x_dir + 8, (y_base + y_top)//2), f"{m_dir:.0f}cm", fill=(0,0,0), font=font, anchor='lm')
                    if m_corpo > 0:
                        draw.text(((x_esq + x_dir)//2, y_base + 8), f"{m_corpo:.0f}cm", fill=(0,0,0), font=font, anchor='mt')
                elif len(medida_dobra) >= 2:
                    m_esq, m_dir = medida_dobra[0], medida_dobra[1]
                    draw.text((x_esq - 8, (y_base + y_top)//2), f"{m_esq:.0f}cm", fill=(0,0,0), font=font, anchor='rm')
                    draw.text((x_dir + 8, (y_base + y_top)//2), f"{m_dir:.0f}cm", fill=(0,0,0), font=font, anchor='lm')
        else:
            x1 = x + int(w * 0.1); x2 = x + int(w * 0.9)
            yb = y + int(h * 0.5)
            draw.line([x1, yb, x2, yb], fill=cor, width=esp)
            # Adicionar medida do comprimento na reta
            if comprimento and comprimento > 0 and font:
                draw.text(((x1 + x2)//2, yb - 12), f"{comprimento:.2f}m", fill=(0,0,0), font=font, anchor='ms')

    def _desenhar_moldura_etiqueta_fase4(self, x, y, w, h):
        """Desenha moldura com marcas de corte"""
        # Moldura principal
        self.canvas_etiq.create_rectangle(x, y, x + w, y + h, outline="#ff6f00", width=2, fill="white")

        # Marcas de corte nos cantos (crop marks)
        tamanho_marca = 5
        cantos = [
            (x, y), (x + w, y),  # Superior esquerdo e direito
            (x, y + h), (x + w, y + h)  # Inferior esquerdo e direito
        ]

        for px, py in cantos:
            # Linhas horizontais
            self.canvas_etiq.create_line(
                px - tamanho_marca, py, px + tamanho_marca, py,
                width=1, fill="black"
            )
            # Linhas verticais
            self.canvas_etiq.create_line(
                px, py - tamanho_marca, px, py + tamanho_marca,
                width=1, fill="black"
            )

    def _editar_desenho_canvas(self, pilar, pos):
        """Editar o desenho (forma e comprimento) e, para estribo, os 4 lados (A,B,C,D)."""
        print(f"\n[CLIQUE DETECTADO!] _editar_desenho_canvas chamado para {pilar}/{pos}")
        chave = (pilar, pos)

        # Buscar dados atuais
        item_atual = None
        for item in self.dados_processados:
            if item[0] == pilar and item[1] == pos:
                item_atual = item
                break
        if not item_atual:
            messagebox.showwarning("Aviso", f"Item {pilar} {pos} não encontrado")
            return

        bitola_atual = float(item_atual[2])
        qtde_atual = int(item_atual[3])
        comp_atual = float(item_atual[4])
        if chave in self.medidas_customizadas:
            bitola_atual = float(self.medidas_customizadas[chave].get('bitola', bitola_atual))
            comp_atual = float(self.medidas_customizadas[chave].get('comp', comp_atual))
            qtde_atual = int(self.medidas_customizadas[chave].get('qtde', qtde_atual))
        forma_atual = self.formas_customizadas.get(chave, 'reta')

        # Controles integrados ao editor (sem dialog)
        controls = tk.Frame(self.janela_etiq, bg="#ecf0f1")
        controls.pack(fill="x", padx=14, pady=8)
        var_forma = tk.StringVar(value=forma_atual)
        tk.Label(controls, text="Forma:", bg="#ecf0f1").grid(row=0, column=0, sticky="w")
        formas_opcoes = [
            ("Reta", "reta"),
            ("Gancho", "gancho"),
            ("Dobra 1 lado", "dobra"),
            ("Dobra 2 lados", "dobra_dupla"),
            ("Estribo ret.", "estribo"),
            ("Estribo red.", "estribo_redondo"),
        ]
        for j, (label, f) in enumerate(formas_opcoes):
            tk.Radiobutton(controls, text=label, variable=var_forma, value=f, bg="#ecf0f1").grid(row=0, column=j+1, padx=4)

        tk.Label(controls, text="Bitola (mm):", bg="#ecf0f1").grid(row=1, column=0, sticky="w", pady=4)
        ent_bitola = tk.Entry(controls, width=10)
        ent_bitola.insert(0, f"{bitola_atual:.1f}")
        ent_bitola.grid(row=1, column=1)

        tk.Label(controls, text="Compr. (m):", bg="#ecf0f1").grid(row=1, column=2, sticky="w")
        ent_comp = tk.Entry(controls, width=10)
        ent_comp.insert(0, f"{comp_atual:.2f}")
        ent_comp.grid(row=1, column=3)

        tk.Label(controls, text="Qtde:", bg="#ecf0f1").grid(row=1, column=4, sticky="w")
        ent_qtde = tk.Entry(controls, width=6)
        ent_qtde.insert(0, str(qtde_atual))
        ent_qtde.grid(row=1, column=5)

        # Entradas de lados do estribo (em cm)
        lados_frame = tk.Frame(self.janela_etiq, bg="#ecf0f1")
        lados_frame.pack(fill="x", padx=14, pady=4)
        tk.Label(lados_frame, text="Estribo (A,B,C,D) cm:", bg="#ecf0f1").grid(row=0, column=0, sticky="w")
        ent_a = tk.Entry(lados_frame, width=6); ent_b = tk.Entry(lados_frame, width=6)
        ent_c = tk.Entry(lados_frame, width=6); ent_d = tk.Entry(lados_frame, width=6)
        ent_a.grid(row=0, column=1, padx=2); ent_b.grid(row=0, column=2, padx=2)
        ent_c.grid(row=0, column=3, padx=2); ent_d.grid(row=0, column=4, padx=2)

        # Entrada para medida da dobra (em cm)
        dobra_frame = tk.Frame(self.janela_etiq, bg="#ecf0f1")
        dobra_frame.pack(fill="x", padx=14, pady=4)
        tk.Label(dobra_frame, text="Dobra 1 (cm):", bg="#ecf0f1").grid(row=0, column=0, sticky="w")
        ent_dobra = tk.Entry(dobra_frame, width=10)
        ent_dobra.grid(row=0, column=1, padx=4)
        
        # Entrada para segunda medida da dobra dupla (em cm)
        tk.Label(dobra_frame, text="Dobra 2 (cm):", bg="#ecf0f1").grid(row=0, column=2, sticky="w")
        ent_dobra2 = tk.Entry(dobra_frame, width=10)
        ent_dobra2.grid(row=0, column=3, padx=4)

        # Entrada para medidas do gancho (A,B em cm)
        gancho_frame = tk.Frame(self.janela_etiq, bg="#ecf0f1")
        gancho_frame.pack(fill="x", padx=14, pady=4)
        tk.Label(gancho_frame, text="Gancho (A,B) cm:", bg="#ecf0f1").grid(row=0, column=0, sticky="w")
        ent_ga = tk.Entry(gancho_frame, width=6); ent_gb = tk.Entry(gancho_frame, width=6)
        ent_ga.grid(row=0, column=1, padx=2); ent_gb.grid(row=0, column=2, padx=2)

        # Pré-preencher se já existir
        if not hasattr(self, 'estribo_lados'):
            self.estribo_lados = {}
        if (pilar, pos) in self.estribo_lados:
            a,b,c,d = self.estribo_lados[(pilar,pos)]
            ent_a.insert(0, str(a)); ent_b.insert(0, str(b)); ent_c.insert(0, str(c)); ent_d.insert(0, str(d))

        if not hasattr(self, 'dobra_medidas'):
            self.dobra_medidas = {}
        if (pilar, pos) in self.dobra_medidas:
            try:
                ent_dobra.insert(0, str(self.dobra_medidas[(pilar, pos)]))
            except Exception:
                pass
        
        # Pré-preencher dobras do medidas_customizadas também
        if chave in self.medidas_customizadas:
            m1 = self.medidas_customizadas[chave].get('medida_dobra', 0.0)
            if m1 > 0:
                ent_dobra.delete(0, 'end')
                ent_dobra.insert(0, str(m1))

        if not hasattr(self, 'gancho_medidas'):
            self.gancho_medidas = {}
        if (pilar, pos) in self.gancho_medidas:
            try:
                ga, gb = self.gancho_medidas[(pilar, pos)]
                ent_ga.insert(0, str(ga)); ent_gb.insert(0, str(gb))
            except Exception:
                pass


        # Removido preview_canvas e dialog: controles e preview agora são integrados ao editor
        # Se desejar preview visual, implemente aqui usando self.canvas_etiq
        btns = tk.Frame(self.janela_etiq, bg="#ecf0f1")
        btns.pack(pady=12)

        def salvar():
            print(f"\n[SALVAR CHAMADO!] Função salvar foi executada para {chave}")
            try:
                novo_bitola = float(ent_bitola.get())
                novo_comp = float(ent_comp.get())
                novo_qtde = int(ent_qtde.get())
            except ValueError:
                messagebox.showerror("Erro", "Digite valores válidos")
                return
            forma_sel = var_forma.get()
            print(f"[SALVAR] chave={chave}, forma_sel='{forma_sel}'")
            if chave not in self.medidas_customizadas:
                self.medidas_customizadas[chave] = {}
            self.medidas_customizadas[chave]['bitola'] = novo_bitola
            self.medidas_customizadas[chave]['comp'] = novo_comp
            self.medidas_customizadas[chave]['qtde'] = novo_qtde
            self.formas_customizadas[chave] = forma_sel
            
            # Se for estribo, salvar lados (valores em cm)
            if forma_sel == 'estribo':
                try:
                    va = float(ent_a.get()); vb = float(ent_b.get()); vc = float(ent_c.get()); vd = float(ent_d.get())
                    self.medidas_customizadas[chave]['lado1'] = va
                    self.medidas_customizadas[chave]['lado2'] = vb
                    self.medidas_customizadas[chave]['lado3'] = vc
                    self.medidas_customizadas[chave]['lado4'] = vd
                    # Também salvar no estribo_lados para compatibilidade
                    if not hasattr(self, 'estribo_lados'):
                        self.estribo_lados = {}
                    self.estribo_lados[(pilar, pos)] = (va, vb, vc, vd)
                except Exception:
                    pass
            # Se for dobra simples, salvar medida (cm) no medidas_customizadas - MESMA LOGICA DO ESTRIBO
            if forma_sel == 'dobra':
                try:
                    vdobra = float(ent_dobra.get()) if ent_dobra.get() else 0.0
                    print(f"[SALVAR DOBRA] vdobra={vdobra}")
                    self.medidas_customizadas[chave]['medida_dobra'] = vdobra
                    self.medidas_customizadas[chave]['medida_dobra_2'] = 0.0
                    # Também salvar no dobra_medidas para compatibilidade (EXATAMENTE COMO ESTRIBO_LADOS)
                    if not hasattr(self, 'dobra_medidas'):
                        self.dobra_medidas = {}
                    self.dobra_medidas[chave] = vdobra
                except Exception:
                    pass
            # Se for dobra dupla, salvar ambas medidas - MESMA LOGICA DO ESTRIBO
            if forma_sel == 'dobra_dupla':
                try:
                    vdobra1 = float(ent_dobra.get()) if ent_dobra.get() else 0.0
                    vdobra2 = float(ent_dobra2.get()) if ent_dobra2.get() else 0.0
                    print(f"[SALVAR DOBRA_DUPLA] vdobra1={vdobra1}, vdobra2={vdobra2}")
                    self.medidas_customizadas[chave]['medida_dobra'] = vdobra1
                    self.medidas_customizadas[chave]['medida_dobra_2'] = vdobra2
                    # Também salvar em dobra_dupla_medidas para compatibilidade (EXATAMENTE COMO ESTRIBO_LADOS)
                    if not hasattr(self, 'dobra_dupla_medidas'):
                        self.dobra_dupla_medidas = {}
                    self.dobra_dupla_medidas[chave] = (vdobra1, vdobra2)
                except Exception:
                    pass
            # Se for gancho, salvar medidas (A,B em cm)
            if forma_sel == 'gancho':
                try:
                    vga = float(ent_ga.get()); vgb = float(ent_gb.get())
                    self.medidas_customizadas[chave]['lado1'] = vga
                    self.medidas_customizadas[chave]['lado2'] = vgb
                except Exception:
                    pass
            try:
                self._pending_yview = self.canvas_etiq.yview()
            except Exception:
                self._pending_yview = None
            # Não há mais dialog.destroy(), apenas atualiza o editor
            self.desenhar_etiquetas_com_selecao()

        tk.Button(btns, text="✓ Salvar", command=salvar, bg="#27ae60", fg="white", font=("Arial", 11, "bold"), width=12).pack(side="left", padx=10)
        print(f"[BOTÃO CRIADO] Botão Salvar criado para {chave} com command=salvar")
        # Botão cancelar apenas limpa campos, não fecha dialog
        tk.Button(btns, text="✗ Cancelar", command=lambda: None, bg="#e74c3c", fg="white", font=("Arial", 11, "bold"), width=12).pack(side="left", padx=10)



    def _editar_medida_etiqueta(self, pilar, pos, tipo):
        """
        Abre um diálogo para editar bitola, comp ou qtde diretamente na etiqueta.
        Armazena as mudanças em self.medidas_customizadas.
        """
        from tkinter import simpledialog
        
        chave = (pilar, pos)
        
        # Buscar valores atuais
        item_atual = None
        for item in self.dados_processados:
            if item[0] == pilar and item[1] == pos:
                item_atual = item
                break
        
        if not item_atual:
            messagebox.showwarning("Aviso", f"Item {pilar} {pos} não encontrado")
            return
        
        # item = (pilar, pos, bitola, comp, qtde, peso_un, tipo, cod_barras, desenho)
        bitola_atual = item_atual[2]
        comp_atual = item_atual[3]
        qtde_atual = item_atual[4]
        
        # Verificar se já tem customização
        if chave in self.medidas_customizadas:
            bitola_atual = self.medidas_customizadas[chave].get('bitola', bitola_atual)
            comp_atual = self.medidas_customizadas[chave].get('comp', comp_atual)
            qtde_atual = self.medidas_customizadas[chave].get('qtde', qtde_atual)
        
        # Diálogo para novo valor
        if tipo == "bitola":
            novo_valor = simpledialog.askfloat(
                "Editar Bitola",
                f"Nova bitola para {pilar} {pos}:",
                initialvalue=bitola_atual,
                minvalue=0.1,
                maxvalue=100.0
            )
        elif tipo == "comp":
            novo_valor = simpledialog.askfloat(
                "Editar Comprimento",
                f"Novo comprimento (cm) para {pilar} {pos}:",
                initialvalue=comp_atual,
                minvalue=1.0,
                maxvalue=10000.0
            )
        elif tipo == "qtde":
            novo_valor = simpledialog.askinteger(
                "Editar Quantidade",
                f"Nova quantidade para {pilar} {pos}:",
                initialvalue=int(qtde_atual),
                minvalue=1,
                maxvalue=1000
            )
        else:
            return
        
        if novo_valor is None:
            return  # Usuário cancelou
        
        # Armazenar customização
        if chave not in self.medidas_customizadas:
            self.medidas_customizadas[chave] = {}
        
        self.medidas_customizadas[chave][tipo] = novo_valor
        
        # Redesenhar a página de etiquetas
        self.desenhar_etiquetas_com_selecao()

    def _desenhar_conteudo_etiqueta_fase4(self, x, y, pilar, pos, bitola, qtde, comp, altura_max):
        """Desenha conteúdo da seção: cabeçalho, PNG técnico e barcode."""
        # Cabeçalho compacto (alinha com padrão de projeto)
        self.canvas_etiq.create_text(
            x, y,
            text=f"{pilar} {pos} • Ø{bitola:.0f} • Q{qtde} • {comp:.2f}m",
            font=("Arial", 6, "bold"),
            fill="black",
            anchor="nw"
        )


        # Espaços base para posicionamento interno
        conteudo_top = y + 14
        png_x, png_y = x + 4, conteudo_top
        barcode_x, barcode_y = x + 130, conteudo_top  # à direita do PNG

        # PNG técnico (90x60 aprox.)
        if ETIQUETAS_HELPER_DISPONIVEL and hasattr(self, 'gerador_etiquetas_dinamico'):
            try:
                pasta = getattr(self.gerador_etiquetas_dinamico, 'pasta_etiquetas', None)
                arq_base = getattr(self.gerador_etiquetas_dinamico, 'arquivo_dxf_base', None)
                
                # Tentar obter caminho completo do DXF original
                caminho_dxf_completo = None
                if hasattr(self.gerador_etiquetas_dinamico, 'arquivos_dxf') and self.gerador_etiquetas_dinamico.arquivos_dxf:
                    caminho_dxf_completo = self.gerador_etiquetas_dinamico.arquivos_dxf[0]
                
                if pasta and arq_base:
                    caminho_png = localizar_desenho_barra(
                        pasta,
                        caminho_dxf_completo if caminho_dxf_completo else arq_base,
                        pilar,
                        pos,
                        bitola,
                        qtde,
                        comp * 100.0  # m -> cm
                    )
                    if caminho_png:
                        # Se retornou marcador DXF, adicionar pilar e pos
                        if caminho_png.startswith("DXF:"):
                            caminho_png = f"{caminho_png}|{pilar}|{pos}"
                        
                        png_photo = carregar_desenho_redimensionado(caminho_png, 90, 60)
                        if png_photo:
                            if not hasattr(self, '_desenho_images'):
                                self._desenho_images = []
                            self._desenho_images.append(png_photo)
                            self.canvas_etiq.create_image(png_x, png_y, image=png_photo, anchor="nw")
                        else:
                            # Placeholder quando não carrega
                            self.canvas_etiq.create_rectangle(png_x, png_y, png_x + 90, png_y + 60, outline="#888")
                            self.canvas_etiq.create_text(png_x + 45, png_y + 30, text="PNG", font=("Arial", 6), fill="#888")
                    else:
                        # Sem arquivo correspondente
                        self.canvas_etiq.create_rectangle(png_x, png_y, png_x + 90, png_y + 60, outline="#888")
                        self.canvas_etiq.create_text(png_x + 45, png_y + 30, text="SEM DESENHO", font=("Arial", 6), fill="#888")
            except Exception:
                # Em caso de erro, insere placeholder
                self.canvas_etiq.create_rectangle(png_x, png_y, png_x + 90, png_y + 60, outline="#888")
                self.canvas_etiq.create_text(png_x + 45, png_y + 30, text="ERRO PNG", font=("Arial", 6), fill="#888")

        # Code128 redimensionado (85x22) para padrão do projeto
        if ETIQUETAS_HELPER_DISPONIVEL and hasattr(self, 'gerador_etiquetas_dinamico'):
            try:
                obra_nome = self.var_obra.get() if hasattr(self, 'var_obra') else "OBRA"
                os_num = f"{self.pagina_atual + 1}-{self.total_paginas}"
                codigo_id = gerar_codigo_identificador(
                    obra=obra_nome,
                    os_num=os_num,
                    elemento=pilar,
                    pos=pos,
                    bitola=bitola,
                    comp=comp
                )

                from core.etiquetas_helper import BARCODE_DISPONIVEL
                if BARCODE_DISPONIVEL:
                    barcode_img = gerar_codigo_barras_imagem(codigo_id, largura_px=85, altura_px=22)
                    barcode_photo = ImageTk.PhotoImage(barcode_img)
                    if not hasattr(self, '_barcode_images'):
                        self._barcode_images = []
                    self._barcode_images.append(barcode_photo)
                    self.canvas_etiq.create_image(barcode_x, barcode_y, image=barcode_photo, anchor="nw")
                else:
                    for j in range(28):
                        self.canvas_etiq.create_line(barcode_x + j * 3, barcode_y, barcode_x + j * 3, barcode_y + 18, width=1, fill="black")
            except Exception:
                # Fallback: barras simples
                for j in range(28):
                    self.canvas_etiq.create_line(barcode_x + j * 3, barcode_y, barcode_x + j * 3, barcode_y + 18, width=1, fill="black")

    def _desenhar_picote_fase4(self, x, y, w):
        """Desenha linha de picote tracejada com label"""
        # Linha tracejada em vermelho
        self.canvas_etiq.create_line(
            x, y, x + w, y,
            fill="red", width=2, dash=(4, 4)
        )

        # Label "PICOTE"
        self.canvas_etiq.create_text(
            x + w // 2, y - 8,
            text="✄ DESTACAR AQUI",
            font=("Arial", 5),
            fill="red"
        )

    def _desenhar_secao_micro_fase4(self, x, y, w, h, pilar, pos, bitola, qtde, comp):
        """Secção inferior (1,9 cm): sem barcode, 2 linhas compactas."""
        # Calcular OS por pilar (índice/total)
        try:
            # Procurar índice desta etiqueta nos dados processados
            idx = -1
            for i, dado in enumerate(self.dados_processados):
                if dado[0] == pilar and str(dado[1]) == str(pos):
                    idx = i
                    break
            
            if idx >= 0:
                pilar_index = sum(1 for i in range(idx) if self.dados_processados[i][0] == pilar)
                pilar_total = sum(1 for d in self.dados_processados if d[0] == pilar)
                os_num = f"{pilar_index + 1}-{pilar_total}"
            else:
                os_num = "-"
        except Exception:
            os_num = "-"

        # Centro horizontal da seção
        texto_center_x = x + w // 2
        
        # Obra e Pavimento para as 2 linhas
        obra = self.var_obra.get() if hasattr(self, 'var_obra') else ""
        pavimento = self.var_pavimento.get() if hasattr(self, 'var_pavimento') else ""
        
        # Linha 1 (topo): Obra/Pav • Pilar/Pos • OS
        y_linha1 = y + h * 0.3
        texto_linha1 = f"{obra}|{pavimento} • {pilar}/{pos} • OS:{os_num}"
        self.canvas_etiq.create_text(
            texto_center_x, y_linha1,
            text=texto_linha1,
            font=("Arial", 7, "bold"),
            fill="black",
            anchor="center"
        )
        
        # Linha 2 (base): Bitola • Qtd • Comp
        y_linha2 = y + h * 0.6
        texto_linha2 = f"Ø{bitola:.1f}mm • Q:{qtde} • C:{comp:.2f}m"
        self.canvas_etiq.create_text(
            texto_center_x, y_linha2,
            text=texto_linha2,
            font=("Arial", 7, "bold"),
            fill="black",
            anchor="center"
        )

    def _desenhar_codigo_fake(self, x, y, largura_etiq, altura_etiq, pilar, pos, bitola):
        """Desenha código de barras fake (fallback)"""
        codigo_y = y + altura_etiq - 40
        codigo_x = x + (largura_etiq - 160) / 2

        # Barras verticais
        for j in range(20):
            bar_w = 3 if j % 2 == 0 else 2
            bar_x = codigo_x + j * 8
            self.canvas_etiq.create_line(bar_x, codigo_y, bar_x, codigo_y + 20,
                                         width=bar_w, fill="black")

        # Texto do código
        codigo = f"{pilar}{pos}{int(bitola * 10):03d}"
        self.canvas_etiq.create_text(x + largura_etiq / 2, codigo_y + 28,
                                     text=codigo, font=("Courier", 9), fill="black")

    def editar_tipo_barra_pilar(self, idx_barra):
        """Abre janela para editar o tipo da barra de pilar com TODOS os tipos"""
        dado = self.dados_processados[idx_barra]
        pilar = dado[0]
        pos = dado[1]
        bitola = dado[2]
        qtde = dado[3]
        comp = dado[4]

        # Criar janela de edição
        janela_edit = tk.Toplevel(self.janela_etiq)
        janela_edit.title(f"Editar Tipo - {pilar}/{pos}")
        janela_edit.geometry("650x800")
        janela_edit.configure(bg="#f5f5f5")

        # Centralizar
        janela_edit.update_idletasks()
        x = (janela_edit.winfo_screenwidth() // 2) - 325
        y = (janela_edit.winfo_screenheight() // 2) - 400
        janela_edit.geometry(f"650x800+{x}+{y}")

        # Informações da barra
        info_frame = tk.Frame(janela_edit, bg="#ff6f00", relief="raised", bd=2)
        info_frame.pack(fill="x", padx=10, pady=10)

        tk.Label(
            info_frame,
            text=f"PILAR: {pilar} - POSIÇÃO: {pos}",
            bg="#ff6f00",
            fg="white",
            font=("Arial", 12, "bold")
        ).pack(pady=5)

        tk.Label(
            info_frame,
            text=f"Bitola: ø{bitola:.1f}mm | Qtd: {qtde} | Comp: {comp:.2f}m ({int(comp * 100)}cm)",
            bg="#ff6f00",
            fg="white",
            font=("Arial", 10)
        ).pack(pady=5)

        # Tipo atual
        tipo_atual = AnalisadorGeometricoPilares.identificar_tipo_pilar(pos, bitola, comp)

        tk.Label(
            janela_edit,
            text=f"Tipo Atual: {tipo_atual}",
            bg="#f5f5f5",
            font=("Arial", 11, "bold")
        ).pack(pady=5)

        # Frame para seleção
        tk.Label(
            janela_edit,
            text="SELECIONE O TIPO:",
            bg="#f5f5f5",
            font=("Arial", 11, "bold"),
            fg="#ff6f00"
        ).pack(pady=5)

        # Variável para armazenar seleção
        var_tipo = tk.StringVar(value=tipo_atual)

        # Frame principal com scroll
        main_frame = tk.Frame(janela_edit, bg="#f5f5f5")
        main_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Canvas e Scrollbar
        canvas = tk.Canvas(main_frame, bg="#f5f5f5", height=500)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#f5f5f5")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Lista de todos os tipos disponíveis
        tipos_disponiveis = [
            ("ESTRIBO", "▭ Fechado retangular", "#ffebee", "#c62828"),
            ("NEGATIVO", "⌒ Apoios (dobras para cima)", "#ffebee", "#d32f2f"),
            ("POSITIVO", "━ Vão (ganchos para baixo)", "#e8f5e9", "#388e3c"),
            ("PORTA_ESTRIBO", "━ Superior reta", "#fff3e0", "#ff6f00"),
            ("DOBRA_DUAS_PONTAS", "┐━┌ Dobras nas duas pontas", "#fce4ec", "#e91e63"),
            ("DOBRA_UMA_PONTA", "━┐ Dobra em uma ponta", "#f3e5f5", "#673ab7"),
            ("BARRA_U", "┏━┓ Formato U", "#e3f2fd", "#1976d2"),
            ("BARRA_Z", "Z Formato Z", "#f3e5f5", "#7b1fa2"),
            ("CAVALETE", "⌃ Formato cavalete", "#e0f7fa", "#00acc1"),
            ("GRAMPO", "┌─┐ Formato grampo", "#fbe9e7", "#ff5722"),
        ]

        # Criar radiobuttons para cada tipo
        for tipo_valor, tipo_desc, cor_fundo, cor_texto in tipos_disponiveis:
            frame_tipo = tk.Frame(scrollable_frame, bg=cor_fundo, relief="ridge", bd=2)
            frame_tipo.pack(fill="x", padx=5, pady=3)

            tk.Radiobutton(
                frame_tipo,
                text=f"{tipo_valor}",
                variable=var_tipo,
                value=tipo_valor,
                bg=cor_fundo,
                font=("Arial", 10, "bold"),
                fg=cor_texto,
                width=20,
                anchor="w"
            ).pack(side="left", padx=10, pady=5)

            tk.Label(
                frame_tipo,
                text=tipo_desc,
                bg=cor_fundo,
                font=("Arial", 9),
                fg=cor_texto
            ).pack(side="left", padx=10)

        # Posicionar canvas e scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame de botões (fora do scroll)
        btn_frame = tk.Frame(janela_edit, bg="#f5f5f5")
        btn_frame.pack(fill="x", pady=10)

        def aplicar_mudanca():
            novo_tipo = var_tipo.get()

            # Salvar configuração
            if not hasattr(self, 'tipos_personalizados'):
                self.tipos_personalizados = {}

            chave = f"{pilar}_{pos}"
            self.tipos_personalizados[chave] = {
                'tipo': novo_tipo,
                'medidas': None
            }

            # Fechar janela
            janela_edit.destroy()

            # Feedback visual
            if hasattr(self, 'janela_etiq'):
                titulo_original = self.janela_etiq.title()
                self.janela_etiq.title(f"✅ {pilar}/{pos} = {novo_tipo} salvo!")
                self.janela_etiq.after(2000, lambda: self.janela_etiq.title(titulo_original))

        # Frame dos botões de ação
        botoes_acao = tk.Frame(btn_frame, bg="#f5f5f5")
        botoes_acao.pack(pady=5)

        tk.Button(
            botoes_acao,
            text="✅ Aplicar",
            command=aplicar_mudanca,
            bg="#27ae60",
            fg="white",
            font=("Arial", 11, "bold"),
            padx=20,
            pady=5
        ).pack(side="left", padx=10)

        tk.Button(
            botoes_acao,
            text="❌ Cancelar",
            command=janela_edit.destroy,
            bg="#e74c3c",
            fg="white",
            font=("Arial", 11, "bold"),
            padx=20,
            pady=5
        ).pack(side="left", padx=10)

    def atualizar_botoes_navegacao(self):
        """Atualiza estado dos botões de navegação"""
        if self.pagina_atual == 0:
            self.btn_primeira.config(state="disabled")
            self.btn_anterior.config(state="disabled")
        else:
            self.btn_primeira.config(state="normal")
            self.btn_anterior.config(state="normal")

        if self.pagina_atual >= self.total_paginas - 1:
            self.btn_proxima.config(state="disabled")
            self.btn_ultima.config(state="disabled")
        else:
            self.btn_proxima.config(state="normal")
            self.btn_ultima.config(state="normal")

        self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {self.total_paginas}")

    def primeira_pagina(self):
        self.pagina_atual = 0
        self.desenhar_etiquetas_com_selecao()
        self.atualizar_botoes_navegacao()

    def _imprimir_etiquetas_rapido(self):
        """Impressão rápida profissional - marca todas e gera direto"""
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return
        
        # Marcar todas as etiquetas
        if not hasattr(self, 'etiquetas_selecionadas'):
            self.etiquetas_selecionadas = {}
        # Inicializa etiquetas_selecionadas apenas se não existir ou tamanho diferente
        if not hasattr(self, 'etiquetas_selecionadas') or len(self.etiquetas_selecionadas) != len(self.dados_processados):
            self.etiquetas_selecionadas = {i: True for i in range(len(self.dados_processados))}
        
        # Chamar geração direta com TODAS
        self._gerar_etiquetas_direto()

    def pagina_anterior(self):
        if self.pagina_atual > 0:
            self.pagina_atual -= 1
            self.desenhar_etiquetas_com_selecao()
            self.atualizar_botoes_navegacao()

    def proxima_pagina(self):
        if self.pagina_atual < self.total_paginas - 1:
            self.pagina_atual += 1
            self.desenhar_etiquetas_com_selecao()
            self.atualizar_botoes_navegacao()

    def _imprimir_direto_do_editor(self):
        """Imprime direto do editor usando PNGs gerados do canvas"""
        # Pegar selecionadas
        selecionadas = [i for i, selecionada in self.etiquetas_selecionadas.items() if selecionada]
        if not selecionadas:
            messagebox.showwarning("Atenção", "Selecione pelo menos uma etiqueta!")
            return
        # Confirmar
        if not messagebox.askyesno(
            "Confirmar Impressão",
            f"Imprimir {len(selecionadas)} etiqueta(s) selecionada(s)?"
        ):
            return
        
        # Gerar PNGs do editor
        try:
            import tempfile
            import os
            pasta_temp = os.path.join(tempfile.gettempdir(), "etiquetas_impressao")
            os.makedirs(pasta_temp, exist_ok=True)
            
            # Gerar apenas as selecionadas
            pngs_para_imprimir = []
            for idx in selecionadas:
                dado = self.dados_processados[idx]
                pilar, pos, bitola, qtde, comp = dado[0], dado[1], dado[2], dado[3], dado[4]
                png_path = self._gerar_png_etiqueta(idx, pilar, pos, bitola, qtde, comp, pasta_temp)
                pngs_para_imprimir.append(png_path)
            
            # Imprimir os PNGs
            for png in pngs_para_imprimir:
                os.startfile(png, "print")
            
            messagebox.showinfo("Sucesso", f"{len(selecionadas)} etiqueta(s) enviada(s) para impressão!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao imprimir:\n{e}")
    
    def desenhar_etiquetas_com_selecao(self):
        """Renderiza etiquetas 100×150mm COM PICOTES (original) + CHECKBOXES para seleção"""
        try:
            yview_prev = getattr(self, '_pending_yview', None)
        except Exception:
            yview_prev = None
        if yview_prev is None:
            try:
                yview_prev = self.canvas_etiq.yview()
            except Exception:
                yview_prev = None
        
        self.canvas_etiq.delete("all")
        self._barcode_images = []
        self._desenho_images = []

        PX_MM = 4
        zf = getattr(self, 'zoom_factor', 1.0)
        MARGEM = (10 * PX_MM) * zf
        LARGURA_ETIQ = (100 * PX_MM) * zf
        ALTURA_TOPO = (93 * PX_MM) * zf
        ALTURA_MICRO = (18 * PX_MM) * zf
        ESPACO_PICOTE = (1 * PX_MM) * zf

        total_picotes_h = (3 * ALTURA_MICRO) + (2 * ESPACO_PICOTE)
        ALTURA_ETIQ = (150 * PX_MM) * zf
        altura_etiqueta = ALTURA_ETIQ

        # Centro horizontal na tela
        canvas_w = int(self.canvas_etiq.winfo_width())
        # Centralizar etiqueta no canvas, mesmo com zoom
        x_base = max(MARGEM, (canvas_w - LARGURA_ETIQ) // 2)

        # Índice inicial baseado na página
        inicio = self.pagina_atual * self.etiquetas_por_pagina
        fim = min(len(self.dados_processados), inicio + self.etiquetas_por_pagina)

        y_cursor = MARGEM

        # --- REFORÇO: Redesenhar etiquetas e garantir visualização ---
        self._checkbox_positions = {}  # REFORÇO: sempre reiniciar antes de desenhar
        for i in range(inicio, fim):
            if i < 0 or i >= len(self.dados_processados):
                print(f"[WARN] Índice {i} inválido em fase4 (0-{len(self.dados_processados)-1})")
                continue
            dado = self.dados_processados[i]
            if not dado or len(dado) < 5:
                print(f"[WARN] Dado inválido em fase4 no índice {i}: {dado}")
                continue
            try:
                pilar, pos, bitola, qtde, comp = (
                    dado[0],
                    str(dado[1]),
                    float(dado[2]),
                    int(dado[3]),
                    float(dado[4]),
                )
            except (ValueError, TypeError, IndexError) as e:
                print(f"[WARN] Erro ao extrair dados em fase4: {e}")
                continue
            chave = (pilar, pos)
            cortado_val = False
            if chave in self.medidas_customizadas:
                cortado_val = bool(self.medidas_customizadas[chave].get('cortado', False))
            
            self._desenhar_moldura_etiqueta(x_base, y_cursor, LARGURA_ETIQ, ALTURA_TOPO)
            self._desenhar_topo_etiqueta(x_base, y_cursor, LARGURA_ETIQ, ALTURA_TOPO, pilar, pos, bitola, qtde, comp)
            extra_gap = max(0, (ALTURA_ETIQ - ALTURA_TOPO - total_picotes_h) / 2)
            base_y = y_cursor + ALTURA_TOPO + (ESPACO_PICOTE / 2) + extra_gap + (2 * PX_MM * zf)
            picote_margin = 8 * zf
            for idx in range(3):
                y_sec = base_y + idx * (ALTURA_MICRO + ESPACO_PICOTE)
                self.canvas_etiq.create_rectangle(x_base + picote_margin, y_sec, x_base + LARGURA_ETIQ, y_sec + ALTURA_MICRO, outline="#cccccc", width=1)
                self._desenhar_secao_micro(
                    x_base + picote_margin + 6 * zf,
                    y_sec + 6 * zf,
                    LARGURA_ETIQ - picote_margin - 12 * zf,
                    ALTURA_MICRO - 12 * zf,
                    pilar,
                    pos,
                    bitola,
                    qtde,
                    comp
                )
                if idx < 2:
                    self._desenhar_picote(x_base, y_sec + ALTURA_MICRO + (ESPACO_PICOTE / 2), LARGURA_ETIQ)
            
            # Informações de identificação na parte superior (E numero da etiqueta)
            info_y = y_cursor + 8
            self.canvas_etiq.create_text(x_base + 8, info_y, text=f"E {i+1}", font=("Arial", int(9*zf), "bold"), fill="#333333", anchor="nw")
            
            # TAG para permitir clique em toda etiqueta para editar
            tag_ret = self.canvas_etiq.create_rectangle(x_base-1, y_cursor-1, x_base + LARGURA_ETIQ+1, 
                                             y_cursor + altura_etiqueta+1,
                                             fill="", outline="", 
                                             tags=f"etiq_{i}")
            # Colocar o retângulo de tag no topo para capturar cliques
            self.canvas_etiq.tag_raise(f"etiq_{i}")
            
            # Indicativo visual discreto do campo 'cortado' (dentro da etiqueta, canto inferior)
            if cortado_val:
                self.canvas_etiq.create_text(x_base + LARGURA_ETIQ - 15, y_cursor + altura_etiqueta - 10, text="✗", font=("Arial", 14, "bold"), fill="#e74c3c")
            # CHECKBOX DE SELEÇÃO (etiquetas_selecionadas)
            checkbox_size = 28 * zf
            x_checkbox = x_base + 8 * zf
            y_checkbox = y_cursor + 8 * zf
            if not hasattr(self, '_checkbox_positions'):
                self._checkbox_positions = {}
            self._checkbox_positions[i] = {
                'x1': x_checkbox - 5 * zf, 'y1': y_checkbox - 5 * zf,
                'x2': x_checkbox + checkbox_size + 5 * zf, 'y2': y_checkbox + checkbox_size + 5 * zf,
                'pilar': pilar, 'pos': pos, 'bitola': bitola, 'qtde': qtde, 'comp': comp
            }
            self.canvas_etiq.create_rectangle(x_checkbox-2 * zf, y_checkbox-2 * zf, x_checkbox+checkbox_size+2 * zf, y_checkbox+checkbox_size+2 * zf,
                                             fill="white", outline="white", width=1)
            if self.etiquetas_selecionadas.get(i, True):
                self.canvas_etiq.create_rectangle(x_checkbox, y_checkbox, x_checkbox+checkbox_size, y_checkbox+checkbox_size,
                                                  fill="#27ae60", outline="#1a5c3a", width=int(3*zf), tags=f"checkbox_{i}")
                self.canvas_etiq.create_line(x_checkbox+6*zf, y_checkbox+14*zf, x_checkbox+11*zf, y_checkbox+20*zf, 
                                             x_checkbox+22*zf, y_checkbox+5*zf, fill="white", width=int(3*zf), capstyle="round", joinstyle="round", tags=f"checkbox_{i}")
            else:
                self.canvas_etiq.create_rectangle(x_checkbox, y_checkbox, x_checkbox+checkbox_size, y_checkbox+checkbox_size,
                                                  outline="#333333", width=int(2*zf), fill="white", tags=f"checkbox_{i}")
            
            # Área clicável invisível sobre o checkbox
            self.canvas_etiq.create_rectangle(x_checkbox - 5 * zf, y_checkbox - 5 * zf, 
                                             x_checkbox + checkbox_size + 5 * zf, y_checkbox + checkbox_size + 5 * zf,
                                             fill="", outline="", tags=f"checkbox_{i}")
            self.canvas_etiq.tag_bind(f"checkbox_{i}", "<Button-1>", lambda e, idx=i: self._toggle_etiqueta_selecao(idx))
            if self.etiquetas_selecionadas.get(i, True):
                self.canvas_etiq.create_text(x_checkbox+checkbox_size+8*zf, y_checkbox+checkbox_size//2, 
                                            text="Selecionado", font=("Arial", int(8*zf), "bold"), fill="#27ae60", anchor="w")
            else:
                self.canvas_etiq.create_text(x_checkbox+checkbox_size+8*zf, y_checkbox+checkbox_size//2, 
                                            text="Clique para selecionar", font=("Arial", int(7*zf)), fill="#999999", anchor="w")
            y_cursor += altura_etiqueta + (ESPACO_PICOTE * 3)

        # Calcular total de páginas
        self.total_paginas = max(1, math.ceil(len(self.dados_processados) / self.etiquetas_por_pagina))

        # Ajustar área de scroll para caber todas as etiquetas desta página + margem extra
        self.canvas_etiq.configure(scrollregion=(0, 0, canvas_w, max(1188, y_cursor + MARGEM + (ALTURA_MICRO * 2))))

        # Reaplicar zoom atual sem perder posição
        self._zoom_aplicado = 1.0
        self._reaplicar_zoom()
        if 'yview_prev' in locals() and yview_prev:
            try:
                self.canvas_etiq.yview_moveto(yview_prev[0])
            except Exception:
                pass
        if hasattr(self, '_pending_yview'):
            try:
                del self._pending_yview
            except Exception:
                pass

        # Atualizar label de página
        if hasattr(self, 'label_pagina'):
            self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {self.total_paginas}")
    
    def _toggle_etiqueta_selecao(self, idx):
        """Toggle - inverte o estado da etiqueta clicada"""
        # Inverter estado APENAS da etiqueta clicada
        current_state = self.etiquetas_selecionadas.get(idx, True)
        self.etiquetas_selecionadas[idx] = not current_state
        
        # Re-renderizar
        self.desenhar_etiquetas_com_selecao()
        
        # Atualizar contador
        total_selecionadas = sum(1 for v in self.etiquetas_selecionadas.values() if v)
        self.label_selecionadas.config(text=f"Selecionadas: {total_selecionadas}/{len(self.dados_processados)}")
    
    def _handle_canvas_click(self, event):
        """Handler de clique no canvas - verifica se clicou em checkbox ou etiqueta"""
        x = self.canvas_etiq.canvasx(event.x)
        y = self.canvas_etiq.canvasy(event.y)
        
        # Primeiro: verificar se clicou em checkbox (área menor, prioridade máxima)
        if hasattr(self, '_checkbox_positions') and self._checkbox_positions:
            for idx, pos_info in self._checkbox_positions.items():
                x1, y1, x2, y2 = pos_info['x1'], pos_info['y1'], pos_info['x2'], pos_info['y2']
                if x1 <= x <= x2 and y1 <= y <= y2:
                    self._toggle_etiqueta_selecao(idx)
                    return "break"
        
        # Segundo: verificar se clicou na área da etiqueta para editar
        items_at_point = self.canvas_etiq.find_overlapping(x-5, y-5, x+5, y+5)
        for item in items_at_point:
            tags = self.canvas_etiq.gettags(item)
            for tag in tags:
                if tag.startswith('etiq_'):
                    idx_str = tag.replace('etiq_', '')
                    try:
                        idx = int(idx_str)
                        if idx < len(self.dados_processados):
                            dado = self.dados_processados[idx]
                            pilar = dado[0]
                            pos = str(dado[1])
                            bitola = float(dado[2])
                            qtde = int(dado[3])
                            comp = float(dado[4])
                            self._editar_etiqueta_dados(idx, pilar, pos, bitola, qtde, comp)
                            return "break"
                    except:
                        pass

    def _editar_etiqueta_dados(self, idx, pilar, pos, bitola, qtde, comp):
        """Abre diálogo para editar dados + FORMA/DESENHO + MEDIDAS ESPECÍFICAS por tipo"""
        import tkinter as tk
        from tkinter import ttk, messagebox
        
        dado = self.dados_processados[idx]
        chave = (pilar, pos)
        
        # Diálogo de edição
        dialog = tk.Toplevel(self.janela_editor)
        dialog.title(f"Editar Etiqueta #{idx+1} - {pilar}/{pos}")
        dialog.geometry("520x600")
        dialog.configure(bg="#0d2818")
        
        # Centralizar no editor
        dialog.update_idletasks()
        x = self.janela_editor.winfo_x() + 400
        y = self.janela_editor.winfo_y() + 200
        dialog.geometry(f"520x600+{x}+{y}")
        
        # Campos de edição
        tk.Label(dialog, text=f"Editar: V{pilar} {pos}", bg="#0d2818", fg="#ff9800", 
                font=("Arial", 11, "bold")).pack(pady=10)
        
        frame = tk.Frame(dialog, bg="#0d2818")
        frame.pack(padx=10, pady=5)
        
        # Bitola
        tk.Label(frame, text="Bitola (mm):", bg="#0d2818", fg="white").pack(side="left", padx=5)
        var_bitola = tk.DoubleVar(value=bitola)
        tk.Entry(frame, textvariable=var_bitola, width=10, font=("Arial", 10)).pack(side="left", padx=5)
        
        # Quantidade
        frame2 = tk.Frame(dialog, bg="#0d2818")
        frame2.pack(padx=10, pady=5)
        tk.Label(frame2, text="Quantidade:", bg="#0d2818", fg="white").pack(side="left", padx=5)
        var_qtde = tk.IntVar(value=qtde)
        tk.Entry(frame2, textvariable=var_qtde, width=10, font=("Arial", 10)).pack(side="left", padx=5)
        
        # Comprimento
        frame3 = tk.Frame(dialog, bg="#0d2818")
        frame3.pack(padx=10, pady=5)
        tk.Label(frame3, text="Comprimento (m):", bg="#0d2818", fg="white").pack(side="left", padx=5)
        var_comp = tk.DoubleVar(value=comp)
        tk.Entry(frame3, textvariable=var_comp, width=10, font=("Arial", 10)).pack(side="left", padx=5)
        
        # FORMA/DESENHO
        frame4 = tk.Frame(dialog, bg="#0d2818")
        frame4.pack(padx=10, pady=5)
        tk.Label(frame4, text="Forma/Desenho:", bg="#0d2818", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)
        
        # Mapeamento de formas (display -> valor interno)
        formas_map = {
            "Reta": "reta",
            "Gancho": "gancho",
            "Dobra": "dobra",
            "Dobra Dupla": "dobra_dupla",
            "Estribo": "estribo",
            "Estribo Quadrado": "estribo_quadrado",
            "Estribo Redondo": "estribo_redondo"
        }
        formas_map_inverso = {v: k for k, v in formas_map.items()}
        
        forma_atual = self.formas_customizadas.get(chave, "reta")
        forma_display = formas_map_inverso.get(forma_atual, "Reta")
        
        var_forma = tk.StringVar(value=forma_display)
        combo_forma = ttk.Combobox(frame4, textvariable=var_forma, values=list(formas_map.keys()), 
                                   state="readonly", width=20)
        combo_forma.set(forma_display)
        combo_forma.pack(side="left", padx=5)
        
        # FRAME para medidas dinâmicas
        frame_medidas = tk.Frame(dialog, bg="#0d2818")
        frame_medidas.pack(padx=10, pady=5, fill="both", expand=True)
        
        # Obter valores atuais
        medidas_atual = self.medidas_customizadas.get(chave, {})
        
        # Variáveis para medidas
        var_medida_dobra = tk.DoubleVar(value=medidas_atual.get('medida_dobra', 0.0))
        var_medida_dobra_2 = tk.DoubleVar(value=medidas_atual.get('medida_dobra_2', 0.0))
        var_lado1 = tk.DoubleVar(value=medidas_atual.get('lado1', 0.0))
        var_lado2 = tk.DoubleVar(value=medidas_atual.get('lado2', 0.0))
        var_lado3 = tk.DoubleVar(value=medidas_atual.get('lado3', 0.0))
        var_lado4 = tk.DoubleVar(value=medidas_atual.get('lado4', 0.0))
        var_raio = tk.DoubleVar(value=medidas_atual.get('raio', 0.0))
        
        def atualizar_campos_forma(*args):
            """Mostrar/ocultar campos conforme forma"""
            for widget in frame_medidas.winfo_children():
                widget.destroy()
            
            forma_display = var_forma.get()
            forma = formas_map.get(forma_display, "reta")  # Converter display -> valor interno
            
            if forma == "reta":
                tk.Label(frame_medidas, text="(Sem medidas adicionais)", bg="#0d2818", fg="#888888", 
                        font=("Arial", 9)).pack(pady=10)
            elif forma == "gancho":
                f = tk.Frame(frame_medidas, bg="#0d2818")
                f.pack(pady=3)
                tk.Label(f, text="Dobra A (cm):", bg="#0d2818", fg="#ffcc00", font=("Arial", 9)).pack(side="left", padx=5)
                tk.Entry(f, textvariable=var_lado1, width=10, font=("Arial", 10)).pack(side="left", padx=5)
                f2 = tk.Frame(frame_medidas, bg="#0d2818")
                f2.pack(pady=3)
                tk.Label(f2, text="Dobra B (cm):", bg="#0d2818", fg="#ffcc00", font=("Arial", 9)).pack(side="left", padx=5)
                tk.Entry(f2, textvariable=var_lado2, width=10, font=("Arial", 10)).pack(side="left", padx=5)
                f3 = tk.Frame(frame_medidas, bg="#0d2818")
                f3.pack(pady=3)
                tk.Label(f3, text="Corpo (cm):", bg="#0d2818", fg="#ffcc00", font=("Arial", 9)).pack(side="left", padx=5)
                tk.Entry(f3, textvariable=var_lado3, width=10, font=("Arial", 10)).pack(side="left", padx=5)
            elif forma == "dobra":
                f = tk.Frame(frame_medidas, bg="#0d2818")
                f.pack(pady=3)
                tk.Label(f, text="Dobra (cm):", bg="#0d2818", fg="#ffcc00", font=("Arial", 9)).pack(side="left", padx=5)
                tk.Entry(f, textvariable=var_medida_dobra, width=10, font=("Arial", 10)).pack(side="left", padx=5)
                f2 = tk.Frame(frame_medidas, bg="#0d2818")
                f2.pack(pady=3)
                tk.Label(f2, text="Corpo (cm):", bg="#0d2818", fg="#ffcc00", font=("Arial", 9)).pack(side="left", padx=5)
                tk.Entry(f2, textvariable=var_medida_dobra_2, width=10, font=("Arial", 10)).pack(side="left", padx=5)
            elif forma == "dobra_dupla":
                f1 = tk.Frame(frame_medidas, bg="#0d2818")
                f1.pack(pady=3)
                tk.Label(f1, text="Dobra 1 (cm):", bg="#0d2818", fg="#ffcc00", font=("Arial", 9)).pack(side="left", padx=5)
                tk.Entry(f1, textvariable=var_medida_dobra, width=10, font=("Arial", 10)).pack(side="left", padx=5)
                f2 = tk.Frame(frame_medidas, bg="#0d2818")
                f2.pack(pady=3)
                tk.Label(f2, text="Dobra 2 (cm):", bg="#0d2818", fg="#ffcc00", font=("Arial", 9)).pack(side="left", padx=5)
                tk.Entry(f2, textvariable=var_medida_dobra_2, width=10, font=("Arial", 10)).pack(side="left", padx=5)
                f3 = tk.Frame(frame_medidas, bg="#0d2818")
                f3.pack(pady=3)
                tk.Label(f3, text="Corpo (cm):", bg="#0d2818", fg="#ffcc00", font=("Arial", 9)).pack(side="left", padx=5)
                tk.Entry(f3, textvariable=var_lado1, width=10, font=("Arial", 10)).pack(side="left", padx=5)
            elif forma in ("estribo_quadrado", "estribo"):
                for i, var, label_txt in [(1, var_lado1, "Lado 1 (cm):"),
                                          (2, var_lado2, "Lado 2 (cm):"),
                                          (3, var_lado3, "Lado 3 (cm):"),
                                          (4, var_lado4, "Lado 4 (cm):")]:
                    f = tk.Frame(frame_medidas, bg="#0d2818")
                    f.pack(pady=3)
                    tk.Label(f, text=label_txt, bg="#0d2818", fg="#ffcc00", font=("Arial", 9)).pack(side="left", padx=5)
                    tk.Entry(f, textvariable=var, width=12, font=("Arial", 10)).pack(side="left", padx=5)
            elif forma == "estribo_redondo":
                f = tk.Frame(frame_medidas, bg="#0d2818")
                f.pack(pady=5)
                tk.Label(f, text="Raio (cm):", bg="#0d2818", fg="#ffcc00", font=("Arial", 9, "bold")).pack(side="left", padx=5)
                tk.Entry(f, textvariable=var_raio, width=12, font=("Arial", 10)).pack(side="left", padx=5)
        
        combo_forma.bind("<<ComboboxSelected>>", atualizar_campos_forma)
        atualizar_campos_forma()
        
        # Campo Cortado
        frame_cortado = tk.Frame(dialog, bg="#0d2818")
        frame_cortado.pack(fill="x", padx=14, pady=4)
        var_cortado = tk.BooleanVar()
        cortado_atual = False
        if chave in self.medidas_customizadas and 'cortado' in self.medidas_customizadas[chave]:
            cortado_atual = bool(self.medidas_customizadas[chave]['cortado'])
        elif hasattr(self, 'checkboxes_conf'):
            chave_check = f"{pilar}_{pos}"
            if chave_check in self.checkboxes_conf:
                cortado_atual = bool(self.checkboxes_conf[chave_check]['cortado'].get())
        var_cortado.set(cortado_atual)
        tk.Checkbutton(frame_cortado, text="Cortado", variable=var_cortado, bg="#0d2818", fg="white",
                      selectcolor="#0d2818", font=("Arial", 10)).pack(side="left")
        
        # Botões
        btn_frame = tk.Frame(dialog, bg="#0d2818")
        btn_frame.pack(pady=15)
        
        def salvar_edicao():
            try:
                peso_novo = dado[5] if len(dado) > 5 else 0
                self.dados_processados[idx] = (pilar, pos, var_bitola.get(), var_qtde.get(), var_comp.get(), peso_novo)
                
                medidas_novo = {
                    'bitola': var_bitola.get(),
                    'qtde': var_qtde.get(),
                    'comp': var_comp.get(),
                    'lado1': var_lado1.get(),
                    'lado2': var_lado2.get(),
                    'lado3': var_lado3.get(),
                    'lado4': var_lado4.get(),
                    'raio': var_raio.get(),
                    'medida_dobra': var_medida_dobra.get(),
                    'medida_dobra_2': var_medida_dobra_2.get(),
                    'cortado': var_cortado.get()
                }
                self.medidas_customizadas[chave] = medidas_novo
                
                # Converter display -> valor interno antes de salvar
                forma_display = var_forma.get()
                forma_selecionada = formas_map.get(forma_display, "reta")
                self.formas_customizadas[chave] = forma_selecionada
                
                # Salvar cortado no checklist
                if hasattr(self, 'checkboxes_conf'):
                    chave_check = f"{pilar}_{pos}"
                    if chave_check in self.checkboxes_conf:
                        self.checkboxes_conf[chave_check]['cortado'].set(var_cortado.get())
                
                dialog.destroy()
                self.desenhar_etiquetas_com_selecao()
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar: {e}")
        
        tk.Button(btn_frame, text="✅ SALVAR", command=salvar_edicao,
                 bg="#27ae60", fg="white", font=("Arial", 10), padx=15, pady=5).pack(side="left", padx=5)
        tk.Button(btn_frame, text="✕ CANCELAR", command=dialog.destroy,
                 bg="#e74c3c", fg="white", font=("Arial", 10), padx=15, pady=5).pack(side="left", padx=5)

    
    def _ir_primeira_pagina_etiquetas(self):
        """Volta para primeira página"""
        self.pagina_atual = 0
        self.canvas_etiq.delete("all")
        self.desenhar_etiquetas_com_selecao()
        self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {self.total_paginas}")
    
    def _ir_pagina_anterior_etiquetas(self):
        """Vai para página anterior"""
        if self.pagina_atual > 0:
            self.pagina_atual -= 1
            self.canvas_etiq.delete("all")
            self.desenhar_etiquetas_com_selecao()
            self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {self.total_paginas}")
    
    def _ir_proxima_pagina_etiquetas(self):
        """Vai para próxima página"""
        if self.pagina_atual < self.total_paginas - 1:
            self.pagina_atual += 1
            self.canvas_etiq.delete("all")
            self.desenhar_etiquetas_com_selecao()
            self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {self.total_paginas}")
    
    def _ir_ultima_pagina_etiquetas(self):
        """Vai para última página"""
        self.pagina_atual = self.total_paginas - 1
        self.canvas_etiq.delete("all")
        self.desenhar_etiquetas_com_selecao()
        self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {self.total_paginas}")
    
    def _mostrar_ajuda_edicao(self):
        """Mostra instruções de como editar os valores"""
        messagebox.showinfo(
            "✏️ COMO EDITAR ETIQUETAS",
            "CLIQUE SOBRE OS VALORES para editar:\n\n"
            "• Ø BITOLA (mm) - diâmetro da barra\n"
            "• COMPRIMENTO (m) - tamanho em metros\n"
            "• QUANTIDADE - número de unidades\n\n"
            "MUDANÇAS SÃO SALVAS AUTOMATICAMENTE\n\n"
            "Quando terminar, clique em:\n"
            "'✅ CONFIRMAR E IMPRIMIR'"
        )
    
    def _imprimir_etiqueta_especifica(self, idx):
        """Imprime apenas uma etiqueta específica após editar"""
        if idx < 0 or idx >= len(self.dados_processados):
            messagebox.showerror("Erro", f"Índice inválido: {idx}")
            return
        
        # Marcar apenas essa etiqueta para impressão
        self.etiquetas_selecionadas = {i: False for i in range(len(self.dados_processados))}
        self.etiquetas_selecionadas[idx] = True
        
        # Chamar função de impressão
        print(f"\n[IMPRESSÃO] Selecionada apenas etiqueta {idx}")
        self._confirmar_e_imprimir_etiquetas()
    
    def _confirmar_e_imprimir_etiquetas(self):
        """Imprime etiquetas selecionadas no padrao vigas (HTML + PNG base64)."""
        self._imprimir_etiquetas_selecionadas()

    def _imprimir_etiquetas_selecionadas(self):
        """Gera PNGs das etiquetas selecionadas e abre HTML para impressao (padrao vigas)."""
        selecionadas = [i for i, sel in self.etiquetas_selecionadas.items() if sel]
        if not selecionadas:
            messagebox.showwarning("Atenção", "Nenhuma etiqueta selecionada!")
            return

        try:
            def _to_float(val, default=0.0):
                try:
                    if isinstance(val, str):
                        v = val.replace("ø", "").replace("Ø", "")
                        v = v.replace("mm", "").replace(" ", "")
                        v = v.replace(",", ".")
                        return float(v)
                    return float(val)
                except Exception:
                    return float(default)

            import tempfile
            import base64
            import webbrowser

            pasta_temp = os.path.join(tempfile.gettempdir(), "etiquetas_pilares_editor")
            os.makedirs(pasta_temp, exist_ok=True)

            caminhos = []
            for idx in selecionadas:
                dado = self.dados_processados[idx]
                pilar, pos, bitola, qtde, comp = dado[0], dado[1], dado[2], dado[3], dado[4]
                bitola_f = _to_float(bitola)
                comp_f = _to_float(comp)
                qtde_i = int(round(_to_float(qtde)))
                caminhos.append(self._gerar_png_etiqueta_editor(idx, pilar, pos, bitola_f, qtde_i, comp_f, pasta_temp))

            imagens_base64 = []
            for caminho in caminhos:
                try:
                    with open(caminho, 'rb') as f:
                        imagens_base64.append(base64.b64encode(f.read()).decode('utf-8'))
                except Exception as e:
                    print(f"[WARN] Erro ao ler {caminho}: {e}")

            if not imagens_base64:
                messagebox.showerror("Erro", "Não foi possível processar as imagens para impressão.")
                return

            html_content = f"""
<!DOCTYPE html>
<html lang=\"pt-BR\">
<head>
    <meta charset=\"UTF-8\">
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">
    <title>Impressao de Etiquetas - Pilares</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: Arial, sans-serif; background: #f5f5f5; padding: 20px; font-weight: bold; }}
        .toolbar {{ position: sticky; top: 0; background: #ffffff; padding: 10px; border-bottom: 1px solid #ddd; display: flex; gap: 10px; z-index: 10; }}
        .btn {{ background: #27ae60; color: white; border: none; padding: 8px 14px; border-radius: 4px; font-weight: bold; cursor: pointer; }}
        .page {{ display: grid; gap: 20px; }}
        .etiqueta {{ background: white; padding: 10px; border: 2px solid #333; width: 100mm; height: 150mm; }}
        img {{ width: 100mm; height: 150mm; object-fit: contain; display: block; margin: 0 auto; image-rendering: auto; }}
        @media print {{
            @page {{ size: 100mm 150mm; margin: 0; }}
            html, body {{ width: 100%; height: 100%; }}
            body {{ background: white; padding: 0; margin: 0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
            .toolbar {{ display: none; }}
            .page {{ gap: 0; }}
            .etiqueta {{ border: none; padding: 0; width: 100mm; height: 150mm; page-break-inside: avoid; page-break-after: always; }}
            img {{ width: 100mm; height: 150mm; }}
        }}
    </style>
</head>
<body>
    <div class=\"toolbar\">
        <button class=\"btn\" onclick=\"window.print()\">Imprimir</button>
    </div>
    <div class=\"page\">
        {''.join([f'<div class="etiqueta"><img src="data:image/png;base64,{img}" /></div>' for img in imagens_base64])}
    </div>
</body>
</html>
"""

            html_temp = tempfile.mktemp(suffix="_etiquetas_pilares.html")
            with open(html_temp, 'w', encoding='utf-8') as f:
                f.write(html_content)
            try:
                webbrowser.open('file://' + html_temp)
            except Exception:
                try:
                    os.startfile(html_temp)
                except Exception:
                    pass
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Falha ao imprimir etiquetas:\n{e}")

    def _sincronizar_dados_editor(self):
        """
        Garante que as customizações e dados do editor estejam atualizados antes de imprimir.
        Gera PNGs a partir do canvas do editor.
        """
        import os
        import tempfile
        from PIL import Image, ImageDraw
        
        # Criar pasta temporária para PNGs do editor
        pasta_temp = os.path.join(tempfile.gettempdir(), "etiquetas_editor")
        os.makedirs(pasta_temp, exist_ok=True)
        
        # Gerar PNG para cada etiqueta selecionada
        self.caminhos_etiquetas_geradas = []
        selecionadas = [i for i, sel in self.etiquetas_selecionadas.items() if sel]
        
        for idx in range(len(self.dados_processados)):
            dado = self.dados_processados[idx]
            pilar, pos, bitola, qtde, comp = dado[0], dado[1], dado[2], dado[3], dado[4]
            
            # Gerar PNG usando o mesmo motor do editor
            png_path = self._gerar_png_etiqueta(idx, pilar, pos, bitola, qtde, comp, pasta_temp)
            self.caminhos_etiquetas_geradas.append(png_path)
    
    def _resolver_forma_medidas_pilar(self, pilar, pos, bitola, comp):
        """Resolve forma e medidas para desenho no PNG das etiquetas de pilares."""
        chave = (pilar, pos)
        forma_raw = self.formas_customizadas.get(chave) if hasattr(self, 'formas_customizadas') else None
        forma = None

        if forma_raw:
            forma_lower = str(forma_raw).lower()
            if 'redondo' in forma_lower or 'red.' in forma_lower:
                forma = 'estribo_redondo'
            elif 'estribo' in forma_lower or 'quadrado' in forma_lower:
                forma = 'estribo'
            elif 'dobra' in forma_lower and 'dupla' in forma_lower:
                forma = 'dobra_dupla'
            elif 'dobra' in forma_lower:
                forma = 'dobra'
            elif 'gancho' in forma_lower:
                forma = 'gancho'
            else:
                forma = 'reta'

        if not forma:
            try:
                tipo_barra = AnalisadorGeometricoPilares.identificar_tipo_pilar(pos, bitola, comp)
                tipo_upper = str(tipo_barra).upper()
                if 'ESTRIBO' in tipo_upper:
                    forma = 'estribo'
                elif 'GANCHO' in tipo_upper or 'POSITIVO' in tipo_upper:
                    forma = 'gancho'
                elif 'DOBRA_DUAS_PONTAS' in tipo_upper:
                    forma = 'dobra_dupla'
                elif 'DOBRA_UMA_PONTA' in tipo_upper:
                    forma = 'dobra'
                else:
                    forma = 'reta'
            except Exception:
                forma = 'reta'

        medidas_dict = self.medidas_customizadas.get(chave, {}) if hasattr(self, 'medidas_customizadas') else {}
        medida_dobra = None
        medidas_gancho = None
        estribo_lados = None

        if forma == 'dobra':
            m_dobra = medidas_dict.get('medida_dobra', 0.0)
            m_corpo = medidas_dict.get('medida_dobra_2', 0.0)
            medida_dobra = (m_dobra, m_corpo)
        elif forma == 'dobra_dupla':
            m1 = medidas_dict.get('medida_dobra', 0.0)
            m2 = medidas_dict.get('medida_dobra_2', 0.0)
            m_corpo = medidas_dict.get('lado1', 0.0)
            medida_dobra = (m1, m2, m_corpo)
        elif forma == 'gancho':
            lado1 = medidas_dict.get('lado1', 0.0)
            lado2 = medidas_dict.get('lado2', 0.0)
            medidas_gancho = (lado1, lado2)
        elif forma == 'estribo':
            va = medidas_dict.get('lado1', 0.0)
            vb = medidas_dict.get('lado2', 0.0)
            vc = medidas_dict.get('lado3', 0.0)
            vd = medidas_dict.get('lado4', 0.0)
            estribo_lados = (va, vb, vc, vd)
        elif forma == 'estribo_redondo':
            medida_dobra = medidas_dict.get('raio', 0.0)

        if not medida_dobra and forma == 'dobra' and hasattr(self, 'dobra_medidas') and chave in self.dobra_medidas:
            medida_dobra = self.dobra_medidas.get(chave)
        if not medida_dobra and forma == 'dobra_dupla' and hasattr(self, 'dobra_dupla_medidas') and chave in self.dobra_dupla_medidas:
            medida_dobra = self.dobra_dupla_medidas.get(chave)
        if not medidas_gancho and forma == 'gancho' and hasattr(self, 'gancho_medidas') and chave in self.gancho_medidas:
            medidas_gancho = self.gancho_medidas.get(chave)
        if forma == 'estribo' and hasattr(self, 'estribo_lados') and chave in self.estribo_lados:
            estribo_lados = self.estribo_lados.get(chave)

        return forma, medida_dobra, medidas_gancho, estribo_lados

    def _gerar_png_etiqueta_editor(self, idx, pilar, pos, bitola, qtde, comp, pasta_destino):
        """Gera PNG com o mesmo layout do editor (100x150mm, 3 picotes)."""
        from PIL import Image, ImageDraw, ImageFont
        import os

        PX_MM = 11.811  # 300 DPI

        def mm(v):
            return int(round(v * PX_MM))

        largura = mm(100)
        altura = mm(150)
        altura_topo = mm(93)
        altura_picote = mm(18)
        espaco_picote = mm(1)

        img = Image.new("RGB", (largura, altura), "white")
        draw = ImageDraw.Draw(img)

        try:
            font_small = ImageFont.truetype("arial.ttf", 32)
            font_tiny = ImageFont.truetype("arial.ttf", 28)
            font_bold = ImageFont.truetype("arialbd.ttf", 36)
            font_pos = ImageFont.truetype("arialbd.ttf", 42)
            font_picote = ImageFont.truetype("arialbd.ttf", 38)
        except Exception:
            font_small = ImageFont.load_default()
            font_tiny = font_small
            font_bold = font_small
            font_pos = font_small
            font_picote = font_small

        draw.rectangle([(0, 0), (largura, altura_topo)], outline="#666666", width=3)

        faixa_larg = mm(10)
        os_w = mm(18)
        os_h = mm(30)
        area_os_x1 = largura - os_w - faixa_larg
        area_os_y1 = 0
        area_os_x2 = largura - faixa_larg
        area_os_y2 = os_h
        draw.rectangle([(area_os_x1, area_os_y1), (area_os_x2, area_os_y2)], outline="#000000", width=2)
        draw.text((area_os_x1 + mm(2), area_os_y1 + mm(2)), "OS", font=font_small, fill="#000000")

        os_txt = f"{self.pagina_atual + 1}-{self.total_paginas}" if hasattr(self, 'total_paginas') else "-"
        linhas_os = os_txt.split("-")
        start_y = area_os_y1 + os_h // 2 - mm(3)
        espaco_linha = mm(4)
        for i, linha in enumerate(linhas_os):
            draw.text(((area_os_x1 + area_os_x2) // 2, start_y + i * espaco_linha), linha,
                      font=font_bold, fill="#000000", anchor="mm")

        faixa_x1 = largura - faixa_larg
        faixa_x2 = largura
        draw.rectangle([(faixa_x1, 0), (faixa_x2, altura_topo)], outline="#000000", width=2)
        obra_nome = self.var_obra.get() if hasattr(self, 'var_obra') else "OBRA.001"
        try:
            img_tmp = Image.new('RGBA', (mm(60), faixa_larg), (255, 255, 255, 0))
            dtmp = ImageDraw.Draw(img_tmp)
            try:
                fnt = ImageFont.truetype("arial.ttf", 28)
            except Exception:
                fnt = ImageFont.load_default()
            tw, th = dtmp.textsize(obra_nome, font=fnt)
            text_x = (img_tmp.width - tw) // 2
            text_y = (img_tmp.height - th) // 2
            dtmp.text((text_x, text_y), obra_nome, fill=(0, 0, 0, 255), font=fnt)
            img_tmp = img_tmp.rotate(90, expand=True)
            img.paste(img_tmp, (faixa_x1 + (faixa_larg - img_tmp.width) // 2, (altura_topo - img_tmp.height) // 2), img_tmp)
        except Exception:
            pass

        area_util_x1 = mm(12)
        area_util_x2 = area_os_x1 - mm(8)
        area_util_w = max(0, area_util_x2 - area_util_x1)

        step_y = mm(8)
        y_current = mm(12)
        pavimento = self.var_pavimento.get() if hasattr(self, 'var_pavimento') else "TÉRREO"

        draw.text((area_util_x1, y_current), "Sigla/Obra", font=font_bold, fill="#000000")
        draw.text((area_util_x1 + mm(28), y_current), obra_nome.replace(" ", " - "), font=font_bold, fill="#000000")
        y_current += step_y

        draw.text((area_util_x1, y_current), "Desenho", font=font_bold, fill="#000000")
        draw.text((area_util_x1 + mm(28), y_current), "PILAR", font=font_bold, fill="#e11d48")
        y_current += step_y

        draw.text((area_util_x1, y_current), "Pavimento", font=font_bold, fill="#000000")
        draw.text((area_util_x1 + mm(28), y_current), pavimento, font=font_bold, fill="#000000")
        y_current += step_y

        draw.text((area_util_x1, y_current), "Elemento", font=font_bold, fill="#000000")
        pos_block_w = mm(26)
        elem_x = area_util_x1 + mm(28)
        elem_max_w = max(0, (area_util_x2 - pos_block_w - mm(4)) - elem_x)
        elemento_txt = str(pilar)
        draw.text((elem_x, y_current), elemento_txt, font=font_bold, fill="#000000")
        draw.text((area_util_x2 - mm(10), y_current), "POS", font=font_bold, fill="#000000")
        draw.text((area_util_x2 - mm(2), y_current), f"{pos}", font=font_pos, fill="#000000", anchor="rm")

        th = mm(8)
        tr = mm(10)
        tab_y1 = mm(45)
        tab_y2 = tab_y1 + th + tr
        tab_x = area_util_x1
        cw1 = mm(16)
        cw2 = mm(34)
        cw3 = mm(22)
        cw4 = mm(18)
        tot = cw1 + cw2 + cw3 + cw4
        if tot > area_util_w and area_util_w > 0:
            escala = area_util_w / tot
            cw1 = int(cw1 * escala); cw2 = int(cw2 * escala); cw3 = int(cw3 * escala); cw4 = int(cw4 * escala)

        draw.rectangle([(tab_x, tab_y1), (tab_x + cw1 + cw2 + cw3 + cw4, tab_y2)], outline="#000000", width=2)
        draw.line([(tab_x + cw1, tab_y1), (tab_x + cw1, tab_y2)], fill="#000000", width=2)
        draw.line([(tab_x + cw1 + cw2, tab_y1), (tab_x + cw1 + cw2, tab_y2)], fill="#000000", width=2)
        draw.line([(tab_x + cw1 + cw2 + cw3, tab_y1), (tab_x + cw1 + cw2 + cw3, tab_y2)], fill="#000000", width=2)
        draw.line([(tab_x, tab_y1 + th), (tab_x + cw1 + cw2 + cw3 + cw4, tab_y1 + th)], fill="#000000", width=2)

        draw.text((tab_x + cw1 // 2, tab_y1 + th // 2), "Bitola", font=font_bold, fill="#000000", anchor="mm")
        draw.text((tab_x + cw1 + cw2 // 2, tab_y1 + th // 2), "Compr. Unit.", font=font_bold, fill="#000000", anchor="mm")
        draw.text((tab_x + cw1 + cw2 + cw3 // 2, tab_y1 + th // 2), "Peso", font=font_bold, fill="#000000", anchor="mm")
        draw.text((tab_x + cw1 + cw2 + cw3 + cw4 // 2, tab_y1 + th // 2), "Qtde", font=font_bold, fill="#000000", anchor="mm")

        chave = (pilar, pos)
        if hasattr(self, 'medidas_customizadas') and chave in self.medidas_customizadas:
            bitola = self.medidas_customizadas[chave].get('bitola', bitola)
            comp = self.medidas_customizadas[chave].get('comp', comp)
            qtde = self.medidas_customizadas[chave].get('qtde', qtde)

        peso_val = 0.0
        try:
            from core.peso import peso_linear_kg_m
            peso_val = peso_linear_kg_m(float(bitola)) * float(comp) * float(qtde)
        except Exception:
            pass

        draw.text((tab_x + cw1 // 2, tab_y1 + th + tr // 2), f"{bitola:.2f}", font=font_small, fill="#000000", anchor="mm")
        draw.text((tab_x + cw1 + cw2 // 2, tab_y1 + th + tr // 2), f"{comp:.3f}", font=font_small, fill="#000000", anchor="mm")
        draw.text((tab_x + cw1 + cw2 + cw3 // 2, tab_y1 + th + tr // 2), f"{peso_val:.2f}", font=font_small, fill="#000000", anchor="mm")
        draw.text((tab_x + cw1 + cw2 + cw3 + cw4 // 2, tab_y1 + th + tr // 2), f"{qtde}", font=font_small, fill="#000000", anchor="mm")

        draw_area_x1 = area_util_x1 + mm(3)
        draw_area_x2 = area_util_x2 - mm(3)
        draw_area_y1 = tab_y2 + mm(3)
        draw_area_y2 = altura_topo - mm(3)
        if draw_area_x2 > draw_area_x1 and draw_area_y2 > draw_area_y1:
            draw.rectangle([(draw_area_x1, draw_area_y1), (draw_area_x2, draw_area_y2)], outline="#e11d48", width=3)
            forma, medida_dobra, medidas_gancho, estribo_lados = self._resolver_forma_medidas_pilar(pilar, pos, bitola, comp)
            self._desenhar_forma_pil(
                draw,
                draw_area_x1,
                draw_area_y1,
                draw_area_x2 - draw_area_x1,
                draw_area_y2 - draw_area_y1,
                forma,
                medida_dobra,
                medidas_gancho,
                estribo_lados,
                font=font_tiny,
                comprimento=comp
            )

        total_picotes_h = (3 * altura_picote) + (2 * espaco_picote)
        y_picote_base = altura_topo + max(0, (altura - altura_topo - total_picotes_h) // 2) + (espaco_picote // 2) + mm(4)
        obra_nome = self.var_obra.get() if hasattr(self, 'var_obra') else "OBRA 001"
        pavimento = self.var_pavimento.get() if hasattr(self, 'var_pavimento') else ""
        picote_x1 = mm(2)
        picote_x2 = largura

        try:
            pilar_index = sum(1 for i in range(idx) if self.dados_processados[i][0] == pilar)
            pilar_total = sum(1 for d in self.dados_processados if d[0] == pilar)
            os_num = f"{pilar_index + 1}-{pilar_total}"
        except Exception:
            os_num = "-"

        for i in range(3):
            y_picote = y_picote_base + i * (altura_picote + espaco_picote)
            draw.rectangle([(picote_x1, y_picote), (picote_x2, y_picote + altura_picote)], outline="#cccccc", width=2)
            texto = f"{obra_nome} | {pilar} | {pos} | ø{bitola:.1f} | {qtde}x{comp:.2f}m"
            draw.text(((picote_x1 + picote_x2) // 2, y_picote + altura_picote // 2), texto,
                      font=font_picote, fill="#000000", anchor="mm")
            if i < 2:
                y_linha = y_picote + altura_picote + espaco_picote // 2
                for x in range(0, largura, 12):
                    draw.line([(x, y_linha), (x + 6, y_linha)], fill="#999999", width=2)

        rotacao = getattr(self, '_rotacao_impressao', 0)
        if rotacao == 180:
            img = img.rotate(180, expand=False)
        elif rotacao == 90:
            img = img.rotate(90, expand=True)
        elif rotacao == 270:
            img = img.rotate(270, expand=True)

        nome_arquivo = f"{pilar}_{pos}_Ø{bitola}_Q{qtde}.png"
        caminho_final = os.path.join(pasta_destino, nome_arquivo)
        img.save(caminho_final, dpi=(300, 300))
        return caminho_final

    def _gerar_png_etiqueta(self, idx, pilar, pos, bitola, qtde, comp, pasta_destino):
        """Encaminha para o gerador com layout do editor."""
        try:
            return self._gerar_png_etiqueta_editor(idx, pilar, pos, bitola, qtde, comp, pasta_destino)
        except Exception as e:
            print(f"[WARN] Falha no PNG do editor, usando legado: {e}")
            return self._gerar_png_etiqueta_legacy(idx, pilar, pos, bitola, qtde, comp, pasta_destino)

    def _gerar_png_etiqueta_legacy(self, idx, pilar, pos, bitola, qtde, comp, pasta_destino):
        """
        Gera PNG de uma etiqueta 100x150mm com 3 picotes embaixo
        Idêntico ao editor visual
        """
        from PIL import Image, ImageDraw, ImageFont
        import os
        import barcode
        from barcode.writer import ImageWriter
        
        # Dimensões em pixels (300 DPI)
        PX_MM = 11.811  # 300 DPI
        LARGURA = int(100 * PX_MM)  # 100mm
        ALTURA = int(150 * PX_MM)   # 150mm
        
        # Criar imagem
        img = Image.new('RGB', (LARGURA, ALTURA), 'white')
        draw = ImageDraw.Draw(img)
        
        try:
            font_title = ImageFont.truetype("arialbd.ttf", 52)
            font_large = ImageFont.truetype("arialbd.ttf", 72)
            font_medium = ImageFont.truetype("arialbd.ttf", 38)
            font_small = ImageFont.truetype("arial.ttf", 28)
            font_tiny = ImageFont.truetype("arial.ttf", 20)
        except:
            font_title = ImageFont.load_default()
            font_large = font_title
            font_medium = font_title
            font_small = font_title
            font_tiny = font_title
        
        # SEÇÃO TOPO (93mm)
        altura_topo = int(93 * PX_MM)
        y = 40  # Aumentado de 15 para descer o cabeçalho
        
        # Bordas da seção topo
        draw.rectangle([(0, 0), (LARGURA, altura_topo)], outline='#666666', width=3)
        
        # Calcular OS por pilar (índice/total)
        try:
            pilar_index = sum(1 for i in range(idx) if self.dados_processados[i][0] == pilar)
            pilar_total = sum(1 for d in self.dados_processados if d[0] == pilar)
            os_num = f"{pilar_index + 1}-{pilar_total}"
        except Exception:
            os_num = ""

        # Peso (se disponível) para exibir no topo
        peso = self.dados_processados[idx][5] if len(self.dados_processados[idx]) > 5 else 0

        # Linha de identificação (apenas OS - sem E)
        draw.text((LARGURA//2, y), f"OS: {pilar}/{pos} {os_num}".strip(), font=font_medium, fill='#000000', anchor='mm')
        y += 38  # Espaçamento aumentado para melhor legibilidade
        draw.text((LARGURA//2, y), f"Ø {bitola:.1f}mm | Q: {qtde} | C: {comp:.2f}m | P: {peso:.2f}kg", font=font_medium, fill='#000000', anchor='mm')
        y += 45

        # Título obra/pavimento
        obra = self.var_obra.get() if hasattr(self, 'var_obra') else "OBRA"
        pavimento = self.var_pavimento.get() if hasattr(self, 'var_pavimento') else "PAV"
        draw.text((LARGURA//2, y), f"{obra} - {pavimento}", font=font_title, fill='#000000', anchor='mm')
        y += 65
        
        # Pilar / Posição (DESTACADO)
        draw.text((LARGURA//2, y), f"{pilar} / {pos}", font=font_large, fill='#000080', anchor='mm')
        y += 78
        
        # Bitola e Qtde
        draw.text((LARGURA//2, y), f"Ø{bitola}mm - QTD: {qtde}", font=font_medium, fill='black', anchor='mm')
        y += 52
        
        # Comprimento
        draw.text((LARGURA//2, y), f"Comprimento: {comp:.2f}m", font=font_small, fill='black', anchor='mm')
        y += 42
        
        # DESENHO DA FORMA (customizado)
        chave = (pilar, pos)
        forma = self.formas_customizadas.get(chave, "reta")
        medidas = self.medidas_customizadas.get(chave, {})
        
        # Mapear nomes de display para internos
        if isinstance(forma, str):
            forma_map_display_to_internal = {
                "Reta": "reta",
                "Gancho": "gancho",
                "Dobra": "dobra",
                "Dobra Dupla": "dobra_dupla",
                "Dobra Única": "dobra",
                "Estribo": "estribo",
                "Estribo Quadrado": "estribo_quadrado",
                "Estribo Retângulo": "estribo",
                "Estribo Redondo": "estribo_redondo"
            }
            forma = forma_map_display_to_internal.get(forma, forma.lower().replace(" ", "_"))
        
        # Desenhar forma
        desenho_y = y - 20  # Começar mais cedo para usar mais espaço
        desenho_h = max(200, altura_topo - desenho_y - 20)  # Usar mais altura
        
        # Criar fonte grande para medidas
        try:
            font_medidas = ImageFont.truetype("arialbd.ttf", 35)  # FONTE AJUSTADA
        except:
            font_medidas = font_large
        
        self._desenhar_forma_png(draw, forma, medidas, LARGURA, desenho_y, desenho_h, font_medidas, comp)
        
        y = altura_topo + 5
        
        # Peso (se disponível)
        if peso > 0:
            draw.text((LARGURA//2, y), f"Peso: {peso:.2f} kg/un", font=font_tiny, fill='#666666', anchor='mm')
        
        # 3 PICOTES (seções de 19mm cada)
        altura_picote = int(19 * PX_MM)
        espaco_picote = int(2 * PX_MM)
        y_picote_base = altura_topo
        
        for i in range(3):
            y_picote = y_picote_base + i * (altura_picote + espaco_picote)
            
            # Bordas do picote
            draw.rectangle([(0, y_picote), (LARGURA, y_picote + altura_picote)], outline='#cccccc', width=2)
            
            # Conteúdo repetido em cada picote - SEM BARCODE, 2 LINHAS
            
            # Centro horizontal
            texto_center_x = LARGURA // 2
            
            # Linha 1: Obra/Pav • Pilar/Pos • OS
            y_linha1 = y_picote + altura_picote * 0.3
            texto_linha1 = f"{obra}|{pavimento} • {pilar}/{pos} • OS:{os_num}"
            draw.text((texto_center_x, y_linha1), texto_linha1, font=font_medium, fill='black', anchor='mm')
            
            # Linha 2: Bitola • Qtd • Comp
            y_linha2 = y_picote + altura_picote * 0.6
            texto_linha2 = f"Ø{bitola:.1f}mm • Q:{qtde} • C:{comp:.2f}m"
            draw.text((texto_center_x, y_linha2), texto_linha2, font=font_medium, fill='black', anchor='mm')
            
            # Linha pontilhada de picote (entre seções)
            if i < 2:
                y_linha = y_picote + altura_picote + espaco_picote // 2
                for x in range(0, LARGURA, 10):
                    draw.line([(x, y_linha), (x + 5, y_linha)], fill='#999999', width=1)
        
        # Aplicar rotação se solicitado (pegar do self se disponível)
        rotacao = getattr(self, '_rotacao_impressao', 0)
        if rotacao == 180:
            img = img.rotate(180, expand=False)
        elif rotacao == 90:
            img = img.rotate(90, expand=True)
        elif rotacao == 270:
            img = img.rotate(270, expand=True)
        
        # Salvar PNG
        nome_arquivo = f"{pilar}_{pos}_Ø{bitola}_Q{qtde}.png"
        caminho_final = os.path.join(pasta_destino, nome_arquivo)
        img.save(caminho_final, dpi=(300, 300))
        
        return caminho_final
    
    def _desenhar_forma_png(self, draw, forma, medidas, largura, y_start, altura, font_medidas=None, comprimento=None):
        """Desenha a forma da armadura no PNG - COMPATÍVEL COM TUPLAS"""
        x_center = largura // 2
        y_center = y_start + altura // 2
        
        if forma in ["reta", "Reta"]:
            draw.line([(x_center - 400, y_center), (x_center + 400, y_center)], fill='black', width=16)
            # Adicionar medida do comprimento
            if comprimento and comprimento > 0 and font_medidas:
                draw.text((x_center, y_center - 50), f"{comprimento:.2f}m", fill='black', font=font_medidas, anchor='ms')
            
        elif forma in ["dobra", "Dobra", "Dobra Única", "dobra_unica"]:
            # Dobra simples: L invertido com dobra (vertical) e corpo (horizontal)
            m_dobra = medidas.get('medida_dobra', 30)
            m_corpo = medidas.get('medida_dobra_2', 0)
            
            x1, x2 = x_center - 300, x_center + 300
            y1, yb = y_center + 80, y_center - 80
            
            draw.line([(x1, y1), (x2, y1)], fill='black', width=16)  # Horizontal (corpo)
            draw.line([(x2, y1), (x2, yb)], fill='black', width=16)  # Vertical (dobra)
            
            # Medidas
            if m_dobra > 0:
                draw.text((x2 + 60, (y1 + yb)//2), f"{m_dobra:.0f}cm", fill='black', font=font_medidas, anchor='lm')
            if m_corpo > 0:
                draw.text(((x1 + x2)//2, y1 + 50), f"{m_corpo:.0f}cm", fill='black', font=font_medidas, anchor='mt')
            
        elif forma in ["dobra_dupla", "Dobra Dupla"]:
            # Dobra dupla: U invertido com 2 dobras (verticais) e corpo (horizontal)
            m_dobra1 = medidas.get('medida_dobra', 25)
            m_dobra2 = medidas.get('medida_dobra_2', 25)
            m_corpo = medidas.get('lado1', 0)
            
            x_esq, x_dir = x_center - 300, x_center + 300
            y_base, y_top = y_center + 80, y_center - 80
            
            draw.line([(x_esq, y_base), (x_dir, y_base)], fill='black', width=16)  # Horizontal (corpo)
            draw.line([(x_esq, y_base), (x_esq, y_top)], fill='black', width=16)   # Vertical esq
            draw.line([(x_dir, y_base), (x_dir, y_top)], fill='black', width=16)   # Vertical dir
            
            # Medidas
            if m_dobra1 > 0:
                draw.text((x_esq - 60, (y_base + y_top)//2), f"{m_dobra1:.0f}cm", fill='black', font=font_medidas, anchor='rm')
            if m_dobra2 > 0:
                draw.text((x_dir + 60, (y_base + y_top)//2), f"{m_dobra2:.0f}cm", fill='black', font=font_medidas, anchor='lm')
            if m_corpo > 0:
                draw.text(((x_esq + x_dir)//2, y_base + 50), f"{m_corpo:.0f}cm", fill='black', font=font_medidas, anchor='mt')
            
        elif forma in ["gancho", "Gancho"]:
            # Gancho: formato Z com 3 medidas
            lado1 = medidas.get('lado1', 20)  # Dobra A
            lado2 = medidas.get('lado2', 20)  # Dobra B
            lado3 = medidas.get('lado3', 0)   # Corpo
            
            x1, x2 = x_center - 280, x_center + 280
            y1, y2, yb = y_center + 80, y_center + 80, y_center - 40
            
            draw.line([(x1, y1), (x1, yb)], fill='black', width=16)
            draw.line([(x1, yb), (x2, yb)], fill='black', width=16)
            draw.line([(x2, yb), (x2, y2)], fill='black', width=16)
            
            if lado1 > 0:
                draw.text((x1 - 60, (y1 + yb)//2), f"{lado1:.0f}cm", fill='black', font=font_medidas, anchor='rm')
            if lado2 > 0:
                draw.text((x2 + 60, (y2 + yb)//2), f"{lado2:.0f}cm", fill='black', font=font_medidas, anchor='lm')
            if lado3 > 0:
                draw.text(((x1 + x2)//2, yb - 50), f"{lado3:.0f}cm", fill='black', font=font_medidas, anchor='mb')
            
        elif forma in ["estribo", "estribo_quadrado", "Estribo", "Estribo Quadrado", "Estribo Retângulo"]:
            lado1 = medidas.get('lado1', 20)
            lado2 = medidas.get('lado2', 30)
            w = min(400, lado1 * 8)
            h = min(300, lado2 * 8)
            draw.rectangle([(x_center - w//2, y_center - h//2), (x_center + w//2, y_center + h//2)], 
                          outline='black', width=16)
            draw.text((x_center, y_center - h//2 - 50), f"{lado1:.0f}cm", fill='black', font=font_medidas, anchor='mm')
            draw.text((x_center + w//2 + 80, y_center), f"{lado2:.0f}cm", fill='black', font=font_medidas, anchor='lm')
            
        elif forma in ["estribo_redondo", "Estribo Redondo"]:
            raio = medidas.get('raio', 15)
            r_px = min(240, raio * 12)
            draw.ellipse([(x_center - r_px, y_center - r_px), (x_center + r_px, y_center + r_px)], 
                        outline='black', width=16)
            draw.text((x_center, y_center + r_px + 60), f"R={raio:.0f}cm", fill='black', font=font_medidas, anchor='mm')
        # Se necessário, pode-se salvar temporariamente as edições em disco ou atualizar variáveis globais
        # Por padrão, apenas garante que as estruturas estejam em memória
    
    def _dialogo_selecionar_impressora(self, impressoras, impressora_padrao):
        """
        Abre diálogo para o usuário selecionar a impressora
        
        Args:
            impressoras: Lista de impressoras disponíveis
            impressora_padrao: Impressora padrão do sistema
            
        Returns:
            Nome da impressora escolhida ou None se cancelado
        """
        from tkinter import simpledialog, Toplevel, Listbox, Button, Frame, Label, SINGLE
        
        # Criar janela customizada
        janela = Toplevel(self)
        janela.title("Selecionar Impressora")
        janela.geometry("500x350")
        janela.transient(self)
        janela.grab_set()
        
        resultado = [None]
        
        # Frame principal
        frame = Frame(janela, padx=20, pady=20)
        frame.pack(fill='both', expand=True)
        
        # Título
        Label(frame, text="🖨️ Selecione a impressora para etiquetas:", 
              font=('Arial', 12, 'bold')).pack(pady=(0, 10))
        
        # Info da padrão
        Label(frame, text=f"Impressora padrão do sistema: {impressora_padrao}",
              font=('Arial', 9), fg='#666').pack(pady=(0, 15))
        
        # Listbox com impressoras
        listbox = Listbox(frame, selectmode=SINGLE, font=('Arial', 10), height=10)
        listbox.pack(fill='both', expand=True, pady=(0, 15))
        
        # Adicionar impressoras
        idx_padrao = 0
        for i, imp in enumerate(impressoras):
            listbox.insert('end', imp)
            if imp == impressora_padrao:
                idx_padrao = i
        
        # Selecionar padrão
        listbox.selection_set(idx_padrao)
        listbox.see(idx_padrao)
        
        # Frame botões
        frame_btns = Frame(frame)
        frame_btns.pack()
        
        def confirmar():
            sel = listbox.curselection()
            if sel:
                resultado[0] = listbox.get(sel[0])
                janela.destroy()
        
        def cancelar():
            resultado[0] = None
            janela.destroy()
        
        Button(frame_btns, text="✅ Confirmar", command=confirmar, 
               bg='#28a745', fg='white', font=('Arial', 10, 'bold'),
               padx=20, pady=8).pack(side='left', padx=5)
        
        Button(frame_btns, text="❌ Cancelar", command=cancelar,
               bg='#dc3545', fg='white', font=('Arial', 10, 'bold'),
               padx=20, pady=8).pack(side='left', padx=5)
        
        # Bind duplo-clique
        listbox.bind('<Double-Button-1>', lambda e: confirmar())
        
        # Centralizar
        janela.update_idletasks()
        x = (janela.winfo_screenwidth() // 2) - (janela.winfo_width() // 2)
        y = (janela.winfo_screenheight() // 2) - (janela.winfo_height() // 2)
        janela.geometry(f"+{x}+{y}")
        
        janela.wait_window()
        return resultado[0]
            
    def _gerar_etiquetas_direto(self, indices_selecionados=None):
        """
        Método profissional DIRETO - IMPRIME INSTANTANEAMENTE
        Gera etiquetas usando GeradorEtiquetasDinamico.gerar_e_imprimir_direto()
        Se indices_selecionados=None, usa todas
        """
        if not self.dados_processados:
            messagebox.showerror("Erro", "Nenhum dado processado!")
            return
        
        if not ETIQUETAS_GERADOR_DISPONIVEL:
            messagebox.showerror("Erro", "GeradorEtiquetasDinamico não disponível!")
            return
        
        # Se não passou índices, usa todas
        if indices_selecionados is None:
            indices_selecionados = list(range(len(self.dados_processados)))
        
        # SELECIONAR IMPRESSORA
        try:
            from core.etiquetas_generator import GeradorEtiquetasDinamico
            impressoras = GeradorEtiquetasDinamico.listar_impressoras_disponiveis()
            impressora_padrao = GeradorEtiquetasDinamico.obter_impressora_padrao()
        except Exception as e:
            print(f"[ERRO] Não foi possível listar impressoras: {e}")
            impressoras = ["Argox OS-214 Plus"]
            impressora_padrao = "Argox OS-214 Plus"
        
        # Diálogo para escolher impressora
        impressora_escolhida = self._dialogo_selecionar_impressora(impressoras, impressora_padrao)
        if impressora_escolhida is None:
            print("[INFO] Impressão cancelada pelo usuário")
            return
        
        print(f"\n[IMPRESSÃO DIRETA] Gerando e imprimindo {len(indices_selecionados)} etiqueta(s)...")
        
        try:
            # Obter arquivos DXF
            arquivos_dxf = []
            if hasattr(self, 'arquivos_selecionados') and self.arquivos_selecionados:
                arquivos_dxf = self.arquivos_selecionados
            elif hasattr(self, 'lista_arquivos'):
                # Tentar obter da lista de arquivos processados
                arquivos_dxf = [item[0] for item in self.lista_arquivos if item[0].lower().endswith('.dxf')]
            
            # Criar gerador
            gerador = GeradorEtiquetasDinamico(
                arquivos_dxf=arquivos_dxf,
                pasta_etiquetas="output/etiquetas",
                obra=self.var_obra.get() if hasattr(self, 'var_obra') else "OBRA 001",
                pavimento=self.var_pavimento.get() if hasattr(self, 'var_pavimento') else "TÉRREO"
            )
            
            # Filtrar apenas os selecionados
            gerador.dados = [self.dados_processados[i] for i in indices_selecionados]
            
            # Transferir customizações se existirem
            if hasattr(self, 'medidas_customizadas'):
                gerador.medidas_customizadas = self.medidas_customizadas
            if hasattr(self, 'formas_customizadas'):
                gerador.formas_customizadas = self.formas_customizadas
            
            # Fechar editor antes de imprimir
            try:
                if hasattr(self, 'janela_editor') and self.janela_editor.winfo_exists():
                    self.janela_editor.destroy()
            except Exception:
                pass
            
            # USAR MÉTODO DE IMPRESSÃO DIRETA com impressora escolhida
            sucesso = gerador.gerar_e_imprimir_direto(
                impressora=impressora_escolhida,
                dpi_x=300,
                dpi_y=300
            )
            
            if not sucesso:
                messagebox.showerror("Erro", "Falha ao enviar etiquetas para impressão!")
            else:
                print(f"[SUCESSO] {len(indices_selecionados)} etiqueta(s) enviada(s) para {impressora_escolhida}")
            
        except Exception as e:
            print(f"[ERRO] {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao gerar/imprimir:\n{str(e)}")
            
    def _fechar_editor_etiquetas(self):
        """Fecha o editor sem gerar"""
        try:
            if hasattr(self, 'janela_editor') and self.janela_editor.winfo_exists():
                self.janela_editor.destroy()
        except Exception:
            pass
    
    def _dialogo_impressao_profissional(self):
        """
        Abre diálogo profissional de impressão com opções
        Usa ImpressaoProfissionalEtiquetas para gerar PDF
        """
        from tkinter import Toplevel, Frame, Label, Radiobutton, Button, messagebox, StringVar, Listbox, SINGLE, Scrollbar
        from core.impressao_etiquetas import ImpressaoProfissionalEtiquetas
        
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Nenhum dado para imprimir!")
            return
        
        # Contar selecionadas
        selecionadas = [i for i, sel in self.etiquetas_selecionadas.items() if sel]
        if not selecionadas:
            messagebox.showwarning("⚠️ Nenhuma seleção", "Marque pelo menos uma etiqueta!")
            return
        
        # Garantir que PNGs existem antes de validar índices
        if not hasattr(self, 'caminhos_etiquetas_geradas') or not self.caminhos_etiquetas_geradas:
            messagebox.showerror("Erro", "Etiquetas PNG não foram geradas!\n\nClique em 'Gerar Etiquetas' primeiro.")
            return
        
        print(f"[DEBUG IMPRESSÃO] Índices selecionados: {selecionadas}")
        print(f"[DEBUG IMPRESSÃO] Total dados_processados: {len(self.dados_processados)}")
        print(f"[DEBUG IMPRESSÃO] Total caminhos_etiquetas_geradas: {len(self.caminhos_etiquetas_geradas) if hasattr(self, 'caminhos_etiquetas_geradas') else 0}")
        
        # Verificar se índices são válidos e mostrar correspondência (PASSO 4)
        for idx in selecionadas:
            # CHECAGEM CRÍTICA: PNG faltando para índice selecionado
            if idx >= len(self.caminhos_etiquetas_geradas):
                raise ValueError(f"PNG faltando para índice {idx}. Total PNGs={len(self.caminhos_etiquetas_geradas)}, solicitado={idx}")
            
            if idx < len(self.dados_processados):
                dado = self.dados_processados[idx]
                if idx < len(self.caminhos_etiquetas_geradas):
                    png = self.caminhos_etiquetas_geradas[idx]
                    print(f"[OK] Índice {idx}: {dado[0]} pos {dado[1]} → PNG: {os.path.basename(png)}")
                else:
                    print(f"[ERRO] Índice {idx}: {dado[0]} pos {dado[1]} → PNG NÃO EXISTE (fora do range)")
            else:
                print(f"[ERRO] Índice {idx}: INVÁLIDO (fora do range de dados)")

        
        # Criar janela
        janela = Toplevel(self)
        janela.title("Impressão Profissional de Etiquetas")
        janela.geometry("650x700")
        janela.transient(self)
        janela.grab_set()
        
        # ========== TÍTULO ==========
        frame_title = Frame(janela, bg="#ff6f00", height=50)
        frame_title.pack(fill='x')
        Label(frame_title, text="🖨️ IMPRESSÃO PROFISSIONAL DE ETIQUETAS",
              font=("Arial", 14, "bold"), bg="#ff6f00", fg="white").pack(pady=10)
        
        # ========== CONTEÚDO COM SCROLL ==========
        from tkinter import Canvas
        canvas_scroll = Canvas(janela, bg="white", highlightthickness=0)
        scrollbar = Scrollbar(janela, orient="vertical", command=canvas_scroll.yview)
        frame_content = Frame(canvas_scroll, bg="white")
        
        frame_content.bind(
            "<Configure>",
            lambda e: canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all"))
        )
        
        canvas_scroll.create_window((0, 0), window=frame_content, anchor="nw")
        canvas_scroll.configure(yscrollcommand=scrollbar.set)
        
        canvas_scroll.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Info
        Label(frame_content, text=f"Etiquetas selecionadas: {len(selecionadas)} de {len(self.dados_processados)}",
              font=("Arial", 10, "bold"), fg="#333", bg="white").pack(anchor='w', pady=(10, 15), padx=20)
        
        # ===== DISPOSIÇÃO =====
        Label(frame_content, text="📐 Disposição de Etiquetas:",
              font=("Arial", 10, "bold"), bg="white").pack(anchor='w', pady=(10, 5), padx=20)
        
        var_disposicao = StringVar(value="uma_por_pagina")
        for opt, label in [("uma_por_pagina", "1 etiqueta por página (100×150mm)"),
                           ("duas_por_pagina", "2 etiquetas por página"),
                           ("tres_por_pagina", "3 etiquetas por página")]:
            Radiobutton(frame_content, text=label, variable=var_disposicao, value=opt,
                       font=("Arial", 9), bg="white").pack(anchor='w', padx=40)
        
        # ===== ORIENTAÇÃO =====
        Label(frame_content, text="🔄 Orientação da Página:",
              font=("Arial", 10, "bold"), bg="white").pack(anchor='w', pady=(15, 5), padx=20)
        
        var_orientacao = StringVar(value="portrait")
        for opt, label in [("portrait", "Retrato (P)"),
                           ("landscape", "Paisagem (L)")]:
            Radiobutton(frame_content, text=label, variable=var_orientacao, value=opt,
                       font=("Arial", 9), bg="white").pack(anchor='w', padx=40)
        
        # ===== ROTAÇÃO =====
        Label(frame_content, text="🔃 Rotação da Etiqueta:",
              font=("Arial", 10, "bold"), bg="white").pack(anchor='w', pady=(15, 5), padx=20)
        
        var_rotacao = StringVar(value="0")
        for opt, label in [("0", "0° - Normal"),
                           ("90", "90° - Rotação (landscape)"),
                           ("180", "180° - Invertida"),
                           ("270", "270° - Rotação inversa")]:
            Radiobutton(frame_content, text=label, variable=var_rotacao, value=opt,
                       font=("Arial", 9), bg="white").pack(anchor='w', padx=40)
        
        # ===== MARGEM =====
        Label(frame_content, text="📏 Margem: 5mm",
              font=("Arial", 10, "bold"), bg="white").pack(anchor='w', pady=(15, 0), padx=20)
        
        # ===== IMPRESSORA =====
        Label(frame_content, text="🖨️ Selecione a Impressora:",
              font=("Arial", 10, "bold"), bg="white").pack(anchor='w', pady=(15, 5), padx=20)
        
        # Listar impressoras
        try:
            impressora_obj = ImpressaoProfissionalEtiquetas([], self.var_obra.get(), self.var_pavimento.get())
            impressoras = impressora_obj.obter_impressoras_disponiveis()
        except Exception as e:
            print(f"[ERRO] Erro ao listar impressoras: {e}")
            impressoras = ["Argox OS-214 Plus"]
        
        frame_listbox = Frame(frame_content, bg="white")
        frame_listbox.pack(fill='x', padx=20, pady=(0, 20))
        
        listbox_impressoras = Listbox(frame_listbox, height=4, font=("Arial", 9))
        scrollbar_imp = Scrollbar(frame_listbox, orient="vertical", command=listbox_impressoras.yview)
        listbox_impressoras.config(yscrollcommand=scrollbar_imp.set)
        
        listbox_impressoras.pack(side="left", fill='both', expand=True)
        scrollbar_imp.pack(side="right", fill="y")
        
        impressora_padrao_idx = 0
        for i, imp in enumerate(impressoras):
            listbox_impressoras.insert('end', imp)
            if "Argox" in imp or i == 0:
                impressora_padrao_idx = i
        
        listbox_impressoras.selection_set(impressora_padrao_idx)
        listbox_impressoras.see(impressora_padrao_idx)
        
        # ========== BOTÕES NO RODAPÉ ==========
        frame_btns = Frame(janela, bg="white", padx=20, pady=15)
        frame_btns.pack(fill='x', side='bottom')
        
        def mostrar_preview():
            """Mostra preview de TODAS as etiquetas selecionadas"""
            try:
                # REGERAR PNGs do editor com customizações
                print("[PREVIEW] Gerando PNGs do editor...")
                
                import tempfile
                import os
                pasta_temp = os.path.join(tempfile.gettempdir(), "etiquetas_preview")
                os.makedirs(pasta_temp, exist_ok=True)
                
                # Gerar PNGs apenas das selecionadas usando a mesma função do editor
                pngs_preview = []
                for idx in selecionadas:
                    dado = self.dados_processados[idx]
                    pilar, pos, bitola, qtde, comp = dado[0], dado[1], dado[2], dado[3], dado[4]
                    png_path = self._gerar_png_etiqueta(idx, pilar, pos, bitola, qtde, comp, pasta_temp)
                    pngs_preview.append(png_path)
                
                print(f"[PREVIEW] {len(pngs_preview)} PNG(s) gerado(s) do editor!")
                
                # Atualizar caminhos
                self.caminhos_etiquetas_geradas = []
                for i in range(len(self.dados_processados)):
                    if i in selecionadas:
                        idx_sel = selecionadas.index(i)
                        self.caminhos_etiquetas_geradas.append(pngs_preview[idx_sel])
                    else:
                        # Para não selecionadas, gerar também para manter sincronia
                        dado = self.dados_processados[i]
                        pilar, pos, bitola, qtde, comp = dado[0], dado[1], dado[2], dado[3], dado[4]
                        png_path = self._gerar_png_etiqueta(i, pilar, pos, bitola, qtde, comp, pasta_temp)
                        self.caminhos_etiquetas_geradas.append(png_path)
                
                from PIL import ImageTk, Image
                import tkinter.ttk as ttk
                
                print(f"[PREVIEW] ✅ {len(pngs_preview)} PNG(s) gerado(s) do editor!")
                
                # Pegar PNGs das selecionadas
                pngs_para_preview = []
                for idx in selecionadas:
                    idx_lista = selecionadas.index(idx)
                    if idx_lista < len(pngs_preview):
                        png_path = pngs_preview[idx_lista]
                        if os.path.exists(png_path):
                            dado = self.dados_processados[idx]
                            pilar_esperada = dado[0]
                            pos_esperada = dado[1]
                            pngs_para_preview.append((idx, png_path, pilar_esperada, pos_esperada))
                            print(f"[PREVIEW] 📄 PNG #{idx+1}: {os.path.basename(png_path)}")
                
                if not pngs_para_preview:
                    messagebox.showerror("Erro", "Nenhum PNG válido encontrado!")
                    return
                
                # Janela com scroll
                janela_preview = Toplevel(janela)
                janela_preview.title(f"🔍 Preview - {len(pngs_para_preview)} Etiqueta(s)")
                janela_preview.geometry("650x900")
                
                Label(janela_preview, text=f"PREVIEW DE {len(pngs_para_preview)} ETIQUETA(S) - Como vai sair na impressora",
                      font=("Arial", 12, "bold"), bg="#007bff", fg="white", pady=10).pack(fill='x')
                
                # Canvas com scrollbar
                canvas_scroll = Canvas(janela_preview, bg="white")
                scrollbar_y = ttk.Scrollbar(janela_preview, orient="vertical", command=canvas_scroll.yview)
                frame_scroll = Frame(canvas_scroll, bg="white")
                
                frame_scroll.bind("<Configure>", lambda e: canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all")))
                canvas_scroll.create_window((0, 0), window=frame_scroll, anchor="nw")
                canvas_scroll.configure(yscrollcommand=scrollbar_y.set)
                
                canvas_scroll.pack(side="left", fill="both", expand=True, padx=5, pady=5)
                scrollbar_y.pack(side="right", fill="y")
                
                # Adicionar todas as imagens com info de correspondência
                janela_preview.imagens_refs = []
                for idx, png_path, pilar_esperada, pos_esperada in pngs_para_preview:
                    img_pil = Image.open(png_path)
                    img_pil = img_pil.convert("RGBA")
                    # Resize com LANCZOS para melhor qualidade
                    img_pil = img_pil.resize((600, 900), Image.Resampling.LANCZOS)
                    img_tk = ImageTk.PhotoImage(img_pil)
                    janela_preview.imagens_refs.append(img_tk)
                    
                    Label(frame_scroll, text=f"Etiqueta #{idx+1} - {pilar_esperada} pos {pos_esperada}", 
                          font=("Arial", 10, "bold"), bg="#ffc107", fg="black", pady=5).pack(fill='x', padx=5, pady=(10, 0))
                    Label(frame_scroll, text=f"PNG: {os.path.basename(png_path)}", 
                          font=("Arial", 8), bg="#e0e0e0", fg="black", pady=2).pack(fill='x', padx=5)
                    Label(frame_scroll, image=img_tk, bg="white").pack(padx=5, pady=5)
                    ttk.Separator(frame_scroll, orient='horizontal').pack(fill='x', pady=10)
                
                # Frame de botões no preview
                frame_btns_preview = Frame(janela_preview, bg="#f0f0f0", pady=15)
                frame_btns_preview.pack(fill='x', side='bottom')
                
                def imprimir_do_preview():
                    """Gera HTML com as etiquetas e abre no navegador para imprimir"""
                    try:
                        import os
                        import base64
                        import tempfile
                        import webbrowser
                        
                        # Coletar todos os PNGs
                        pngs_validos = []
                        for idx, png_path, pilar, pos in pngs_para_preview:
                            if os.path.exists(png_path):
                                pngs_validos.append(png_path)
                        
                        if not pngs_validos:
                            messagebox.showerror("Erro", "Nenhuma imagem válida para imprimir!")
                            return
                        
                        print(f"[IMPRESSÃO HTML] Gerando HTML com {len(pngs_validos)} etiquetas...")
                        
                        # Converter PNGs para base64
                        imagens_base64 = []
                        for png_path in pngs_validos:
                            try:
                                with open(png_path, 'rb') as f:
                                    img_data = base64.b64encode(f.read()).decode('utf-8')
                                    imagens_base64.append(img_data)
                            except Exception as e:
                                print(f"Erro ao ler {png_path}: {e}")
                                continue
                        
                        if not imagens_base64:
                            messagebox.showerror("Erro", "Não foi possível processar as imagens!")
                            return
                        
                        # Gerar HTML
                        html_content = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Impressão de Etiquetas - EngenhariaPlanPro</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: Arial, sans-serif;
            background-color: #f5f5f5;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1000px;
            margin: 0 auto;
        }}
        
        .header {{
            background-color: #2c3e50;
            color: white;
            padding: 20px;
            margin-bottom: 20px;
            border-radius: 5px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 24px;
            margin-bottom: 5px;
        }}
        
        .header p {{
            font-size: 14px;
            opacity: 0.9;
        }}
        
        .info {{
            background-color: #e8f4f8;
            border-left: 4px solid #3498db;
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 3px;
            font-size: 14px;
        }}
        
        .gallery {{
            display: grid;
            grid-template-columns: 1fr;
            gap: 30px;
            margin-bottom: 30px;
            max-width: 500px;
            margin-left: auto;
            margin-right: auto;
        }}
        
        .etiqueta-item {{
            background-color: white;
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            text-align: center;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            width: 100%;
            aspect-ratio: 100 / 150;
            position: relative;
        }}
        
        .etiqueta-item img {{
            width: 100%;
            height: 100%;
            margin-bottom: 0;
            display: block;
            object-fit: contain;
            max-height: none;
        }}
        
        .etiqueta-numero {{
            font-size: 10px;
            color: #666;
            background-color: transparent;
            padding: 2px;
            border-radius: 0;
            width: 100%;
            margin-top: 0;
            position: absolute;
            bottom: 0;
            left: 0;
            right: 0;
        }}
        
        .print-info {{
            background-color: #fff3cd;
            border-left: 4px solid #ffc107;
            padding: 15px;
            margin-top: 20px;
            border-radius: 3px;
            font-size: 14px;
            text-align: center;
        }}
        
        .btn-print {{
            background-color: #28a745;
            color: white;
            border: none;
            padding: 12px 30px;
            font-size: 16px;
            border-radius: 5px;
            cursor: pointer;
            margin: 10px;
        }}
        
        .btn-print:hover {{
            background-color: #218838;
        }}
        
        .btn-container {{
            text-align: center;
            margin: 20px 0;
        }}
        
        @media print {{
            body {{
                background-color: white;
                padding: 0;
            }}
            
            .header {{
                display: none;
            }}
            
            .info {{
                display: none;
            }}
            
            .print-info {{
                display: none;
            }}
            
            .btn-container {{
                display: none;
            }}
            
            .gallery {{
                grid-template-columns: 1fr;
                gap: 15px;
            }}
            
            .etiqueta-item {{
                border: none;
                box-shadow: none;
                padding: 10px;
                page-break-inside: avoid;
            }}
            
            .etiqueta-numero {{
                display: block;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📋 Impressão de Etiquetas</h1>
            <p>EngenhariaPlanPro - Sistema de Gestão de Projetos</p>
        </div>
        
        <div class="info">
            <strong>Total de etiquetas:</strong> {len(imagens_base64)} | 
            <strong>Data:</strong> {__import__('datetime').datetime.now().strftime('%d/%m/%Y %H:%M')}
        </div>
        
        <div class="btn-container">
            <button class="btn-print" onclick="window.print();">🖨️ IMPRIMIR AGORA</button>
        </div>
        
        <div class="gallery">
"""
                        
                        # Adicionar imagens ao HTML
                        for i, img_b64 in enumerate(imagens_base64, 1):
                            html_content += f"""
            <div class="etiqueta-item">
                <img src="data:image/png;base64,{img_b64}" alt="Etiqueta {i}">
                <div class="etiqueta-numero">Etiqueta #{i}</div>
            </div>
"""
                        
                        html_content += """
        </div>
        
        <div class="print-info">
            <strong>💡 Dica:</strong> Use o botão acima, Ctrl+P ou o botão de impressão do navegador
        </div>
    </div>
    
    <script>
        // Auto-print ao carregar (comentado por padrão)
        // window.print();
        
        // Função para imprimir manualmente
        function imprimir() {
            window.print();
        }
    </script>
</body>
</html>
"""
                        
                        # Criar arquivo HTML temporário
                        html_temp = tempfile.mktemp(suffix=".html")
                        with open(html_temp, 'w', encoding='utf-8') as f:
                            f.write(html_content)
                        
                        print(f"[IMPRESSÃO HTML] HTML criado: {html_temp}")
                        
                        # Abrir no navegador
                        webbrowser.open('file://' + html_temp)
                        print(f"[IMPRESSÃO HTML] ✓ Abrindo no navegador...")
                        print(f"[IMPRESSÃO HTML] ✅ {len(imagens_base64)} etiqueta(s) carregada(s)!")
                        print(f"[IMPRESSÃO HTML] Use Ctrl+P para abrir o dialog de impressão")
                            
                    except Exception as e:
                        messagebox.showerror("Erro", f"Erro ao gerar HTML:\n{e}")
                        import traceback
                        traceback.print_exc()
                
                Button(frame_btns_preview, text="🖨️ IMPRIMIR AGORA", command=imprimir_do_preview,
                       font=("Arial", 12, "bold"), bg="#28a745", fg="white",
                       padx=30, pady=12).pack(side='left', padx=10)
                
                Button(frame_btns_preview, text="✅ FECHAR", command=janela_preview.destroy,
                       font=("Arial", 12, "bold"), bg="#6c757d", fg="white",
                       padx=30, pady=12).pack(side='left', padx=10)
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao mostrar preview:\n{e}")
                import traceback
                traceback.print_exc()
        
        def executar_impressao():
            # Obter seleções
            disposicao = var_disposicao.get()
            orientacao = var_orientacao.get()
            rotacao_graus = int(var_rotacao.get())
            sel_imp = listbox_impressoras.curselection()
            
            if not sel_imp:
                messagebox.showwarning("Atenção", "Selecione uma impressora!")
                return
            
            impressora = listbox_impressoras.get(sel_imp[0])
            
            print(f"\n[IMPRESSÃO] ========== INICIANDO IMPRESSÃO ==========")
            print(f"[IMPRESSÃO] Impressora: {impressora}")
            print(f"[IMPRESSÃO] Disposição: {disposicao}")
            print(f"[IMPRESSÃO] Orientação: {orientacao}")
            print(f"[IMPRESSÃO] Rotação: {rotacao_graus}°")
            print(f"[IMPRESSÃO] Etiquetas: {len(selecionadas)}")
            
            # USAR PNGs JÁ GERADOS pelo botão "Gerar Etiquetas"
            try:
                if not hasattr(self, 'caminhos_etiquetas_geradas') or not self.caminhos_etiquetas_geradas:
                    messagebox.showerror("Erro", "Etiquetas não foram geradas!\n\nClique em 'Gerar Etiquetas' primeiro.")
                    return
                
                # CRÍTICO: Verificar se arrays correspondem
                if len(self.caminhos_etiquetas_geradas) != len(self.dados_processados):
                    messagebox.showerror("ERRO CRÍTICO", 
                        f"Incompatibilidade de dados!\n\n"
                        f"PNGs gerados: {len(self.caminhos_etiquetas_geradas)}\n"
                        f"Dados processados: {len(self.dados_processados)}\n\n"
                        f"Clique em 'Gerar Etiquetas' novamente!")
                    print(f"[ERRO CRÍTICO] PNGs: {len(self.caminhos_etiquetas_geradas)}, Dados: {len(self.dados_processados)}")
                    return
                
                # Filtrar PNGs das etiquetas selecionadas
                caminhos_png = [self.caminhos_etiquetas_geradas[i] for i in selecionadas 
                               if i < len(self.caminhos_etiquetas_geradas)]
                
                if not caminhos_png:
                    messagebox.showerror("Erro", "Nenhuma etiqueta válida selecionada!")
                    return
                
                # Verificar se PNGs existem
                pngs_faltando = [p for p in caminhos_png if not os.path.exists(p)]
                if pngs_faltando:
                    messagebox.showerror("Erro", f"PNG(s) não encontrado(s):\n{pngs_faltando[0]}\n\nGere as etiquetas novamente!")
                    return
                
                print(f"[IMPRESSÃO] Usando {len(caminhos_png)} PNG(s) já gerado(s)")
                
                # Imprimir PNGs existentes
                janela.update()
                
                if impressora.strip().lower().startswith("microsoft print to pdf"):
                    # Para PDF, criar um PDF com os PNGs
                    from reportlab.pdfgen import canvas as pdf_canvas
                    from reportlab.lib.units import mm
                    
                    pdf_temp = os.path.join("output", "impressao", "temp_etiquetas.pdf")
                    os.makedirs(os.path.dirname(pdf_temp), exist_ok=True)
                    
                    c = pdf_canvas.Canvas(pdf_temp, pagesize=(100*mm, 150*mm))
                    for png_path in caminhos_png:
                        if os.path.exists(png_path):
                            c.drawImage(png_path, 0, 0, width=100*mm, height=150*mm)
                            c.showPage()
                    c.save()
                    
                    os.startfile(pdf_temp, "print")
                    sucesso = True
                else:
                    # Imprimir PNG direto - CONSOLIDADO EM PDF ÚNICO
                    print(f"[IMPRESSÃO] Consolidando {len(caminhos_png)} PNG(s) em PDF único para {impressora}")
                    
                    try:
                        from PIL import Image
                        import tempfile
                        
                        # Criar PDF único com todos os PNGs
                        pdf_consolidado = os.path.join(tempfile.gettempdir(), "etiquetas_impressao_consolidado.pdf")
                        
                        imagens = []
                        for png_path in caminhos_png:
                            if not os.path.exists(png_path):
                                print(f"[WARN] PNG não encontrado: {png_path}")
                                continue
                            
                            img = Image.open(png_path)
                            if img.mode != 'RGB':
                                img = img.convert('RGB')
                            imagens.append(img)
                        
                        if imagens:
                            # Salvar tudo em um único PDF
                            imagens[0].save(pdf_consolidado, save_all=True, append_images=imagens[1:] if len(imagens) > 1 else [])
                            print(f"[✅] PDF consolidado criado: {pdf_consolidado}")
                            
                            # Imprimir PDF único
                            import win32api
                            try:
                                win32api.ShellExecute(0, "printto", pdf_consolidado, f'"{impressora}"', "", 0)
                                print(f"[✅] PDF enviado para {impressora}")
                                sucesso = True
                            except:
                                # Fallback: abrir PDF
                                os.startfile(pdf_consolidado, "print")
                                sucesso = True
                        else:
                            print(f"[ERRO] Nenhuma imagem válida para consolidar")
                            sucesso = False
                        
                        print(f"[✅] {len(imagens)} etiqueta(s) enviada(s) em arquivo único")
                    except Exception as e:
                        print(f"[ERRO] Erro ao consolidar/imprimir: {e}")
                        sucesso = False
                
                if sucesso:
                    messagebox.showinfo("✅ Sucesso", 
                        f"{len(selecionadas)} etiqueta(s) enviada(s) para:\n{impressora}\n\nRotação: {rotacao_graus}°")
                    janela.destroy()
                    # Fechar editor se existir
                    try:
                        if hasattr(self, 'janela_editor') and self.janela_editor.winfo_exists():
                            self.janela_editor.destroy()
                    except Exception:
                        pass
                else:
                    messagebox.showerror("Erro", "Falha ao enviar para impressão!\n\nVerifique o console para detalhes.")
                    
            except Exception as e:
                print(f"[ERRO] {e}")
                import traceback
                traceback.print_exc()
                messagebox.showerror("Erro", f"Erro na impressão:\n{str(e)}\n\nVerifique o console para detalhes.")
        
        def cancelar():
            janela.destroy()
        
        Button(frame_btns, text="🔍 VER PREVIEW", command=mostrar_preview,
               bg="#007bff", fg="white", font=("Arial", 12, "bold"),
               padx=20, pady=12, width=15).pack(side='left', padx=5)
        
        Button(frame_btns, text="✅ IMPRIMIR AGORA", command=executar_impressao,
               bg="#28a745", fg="white", font=("Arial", 12, "bold"),
               padx=40, pady=12, width=20).pack(side='left', padx=5, fill='x', expand=True)
        
        Button(frame_btns, text="❌ CANCELAR", command=cancelar,
               bg="#dc3545", fg="white", font=("Arial", 12, "bold"),
               padx=40, pady=12, width=15).pack(side='left', padx=5)
        
        # Centralizar
        janela.update_idletasks()
        x = (janela.winfo_screenwidth() // 2) - (janela.winfo_width() // 2)
        y = (janela.winfo_screenheight() // 2) - (janela.winfo_height() // 2)
        janela.geometry(f"+{x}+{y}")

    def ultima_pagina(self):
        self.pagina_atual = self.total_paginas - 1
        self.desenhar_etiquetas_com_selecao()
        self.atualizar_botoes_navegacao()

    def desenhar_preview_com_pngs_gerados(self):
        """
        NOVO: Carrega e exibe os PNGs já gerados pelo GeradorEtiquetasDinamico
        Garante que PREVIEW = IMPRESSÃO (100% idêntico)
        """
        print("\n[PREVIEW] Carregando PNGs gerados...")
        
        # Limpar canvas
        self.canvas_etiq.delete("all")
        self._barcode_images = []
        self._desenho_images = []
        
        # Verificar se há PNGs gerados
        if not hasattr(self, 'caminhos_etiquetas_geradas') or not self.caminhos_etiquetas_geradas:
            # Fallback: tentar localizar PNGs na pasta
            pasta_etiquetas = r"c:\EngenhariaPlanPro\etiquetas"
            print(f"[PREVIEW] Procurando PNGs em: {pasta_etiquetas}")
            
            if os.path.exists(pasta_etiquetas):
                pngs = sorted([os.path.join(pasta_etiquetas, f) 
                              for f in os.listdir(pasta_etiquetas) 
                              if f.startswith('ETIQUETA_') and f.endswith('.png')])
                self.caminhos_etiquetas_geradas = pngs
                print(f"[PREVIEW] Encontrados {len(pngs)} PNG(s)")
            else:
                print(f"[PREVIEW] Pasta não existe: {pasta_etiquetas}")
        
        if not self.caminhos_etiquetas_geradas:
            messagebox.showerror(
                "Erro - Etiquetas não geradas", 
                "Nenhuma etiqueta PNG foi gerada!\n\n"
                "Por favor, processe os arquivos DXF primeiro.\n"
                "O sistema irá gerar as etiquetas automaticamente."
            )
            return
        
        # Calcular índices da página atual
        inicio = self.pagina_atual * self.etiquetas_por_pagina
        fim = min(inicio + self.etiquetas_por_pagina, len(self.caminhos_etiquetas_geradas))
        
        print(f"[PREVIEW] Página {self.pagina_atual + 1}: exibindo etiquetas {inicio+1} a {fim}")
        
        # Dimensões e posicionamento
        canvas_w = int(self.canvas_etiq.cget('width'))
        margem = 20
        y_cursor = margem
        
        # Carregar e exibir cada PNG
        for idx in range(inicio, fim):
            caminho_png = self.caminhos_etiquetas_geradas[idx]
            
            try:
                # Carregar PNG gerado
                img = Image.open(caminho_png)
                print(f"[PREVIEW] Carregando: {os.path.basename(caminho_png)} ({img.size[0]}x{img.size[1]}px)")
                
                # Redimensionar para caber no canvas (mantendo proporção)
                max_width = canvas_w - (2 * margem)
                max_height = 600
                
                # Calcular escala mantendo proporção
                scale_w = max_width / img.size[0] if img.size[0] > max_width else 1.0
                scale_h = max_height / img.size[1] if img.size[1] > max_height else 1.0
                scale = min(scale_w, scale_h)
                
                if scale < 1.0:
                    new_size = (int(img.size[0] * scale), int(img.size[1] * scale))
                    img = img.resize(new_size, Image.Resampling.LANCZOS)
                    print(f"[PREVIEW] Redimensionado para: {new_size[0]}x{new_size[1]}px")
                
                # Converter para PhotoImage
                photo = ImageTk.PhotoImage(img)
                self._desenho_images.append(photo)
                
                # Centralizar horizontalmente
                x_center = canvas_w // 2
                
                # Desenhar no canvas
                self.canvas_etiq.create_image(x_center, y_cursor, image=photo, anchor="n")
                
                # Adicionar label com número da etiqueta
                self.canvas_etiq.create_text(
                    x_center, y_cursor + img.height + 10,
                    text=f"Etiqueta #{idx + 1} de {len(self.caminhos_etiquetas_geradas)}",
                    font=("Arial", 9, "bold"),
                    fill="#ff6f00"
                )
                
                # Adicionar nome do arquivo (para debug)
                self.canvas_etiq.create_text(
                    x_center, y_cursor + img.height + 25,
                    text=os.path.basename(caminho_png),
                    font=("Arial", 7),
                    fill="gray"
                )
                
                # Atualizar cursor
                y_cursor += img.height + 50
                
            except Exception as e:
                print(f"[ERRO] Falha ao carregar PNG {caminho_png}: {e}")
                import traceback
                traceback.print_exc()
                
                # Desenhar placeholder de erro
                self.canvas_etiq.create_rectangle(
                    margem, y_cursor, canvas_w - margem, y_cursor + 200,
                    outline="red", width=2, fill="#ffe6e6"
                )
                self.canvas_etiq.create_text(
                    canvas_w // 2, y_cursor + 100,
                    text=f"❌ ERRO ao carregar\n{os.path.basename(caminho_png)}\n\n{str(e)}",
                    font=("Arial", 10),
                    fill="red",
                    justify="center"
                )
                y_cursor += 220
        
        # Atualizar scroll region
        self.canvas_etiq.configure(scrollregion=(0, 0, canvas_w, y_cursor + margem))
        
        # Atualizar label de página
        if hasattr(self, 'label_pagina'):
            self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {self.total_paginas}")
        
        print(f"[PREVIEW] ✅ Preview atualizado com sucesso!\n")

    def aplicar_etq_pagina(self):
        try:
            val = int(self.var_etq_pag.get())
            if val < 1:
                val = 1
        except Exception:
            val = 1
        self.etiquetas_por_pagina = val
        import math
        self.total_paginas = max(1, math.ceil(len(self.dados_processados) / self.etiquetas_por_pagina))
        self.pagina_atual = min(self.pagina_atual, self.total_paginas - 1)
        self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {self.total_paginas}")
        self.desenhar_etiquetas_com_selecao()
        self.atualizar_botoes_navegacao()

    def aplicar_zoom(self):
        try:
            z = float(self.var_zoom.get()) / 100.0
            if z <= 0:
                z = 1.0
        except Exception:
            z = 1.0
        self.zoom_factor = z
        self._reaplicar_zoom()

    def _reaplicar_zoom(self):
        if not hasattr(self, 'canvas_etiq'):
            return
        desejado = getattr(self, 'zoom_factor', 1.0)
        atual = getattr(self, '_zoom_aplicado', 1.0)
        if desejado <= 0:
            desejado = 1.0
        fator = desejado / atual if atual != 0 else desejado
        if abs(fator - 1.0) < 1e-3:
            self._zoom_aplicado = desejado
            return
        try:
            yview = self.canvas_etiq.yview()
        except Exception:
            yview = (0, 0)
        try:
            self.canvas_etiq.scale("all", 0, 0, fator, fator)
            sr = self.canvas_etiq.cget('scrollregion')
            if sr:
                vals = [float(v) for v in sr.split()]
                if len(vals) == 4:
                    vals = [v * fator for v in vals]
                    self.canvas_etiq.configure(scrollregion=tuple(vals))
            self._zoom_aplicado = desejado
            if yview:
                self.canvas_etiq.yview_moveto(yview[0])
        except Exception as e:
            print(f"[zoom] falha ao aplicar: {e}")

    def exportar_excel(self):
        """Exporta dados para Excel com formatação profissional"""
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # Aba 1: Dados Gerais
            ws1 = wb.active
            ws1.title = "Pilares - Geral"

            # Cabeçalho
            ws1.merge_cells('A1:F1')
            ws1['A1'] = 'RELATÓRIO DE PILARES - DETALHADO'
            ws1['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws1['A1'].fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
            ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # Informações do projeto
            ws1['A3'] = 'Obra:'
            ws1['B3'] = self.var_obra.get()
            ws1['A4'] = 'Pavimento:'
            ws1['B4'] = self.var_pavimento.get()
            ws1['A5'] = 'Data:'
            ws1['B5'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            ws1['D3'] = 'Total de Barras:'
            ws1['E3'] = self.total_barras
            ws1['D4'] = 'Peso Total (kg):'
            ws1['E4'] = f"{self.total_kg:.2f}"

            # Cabeçalhos da tabela
            headers = ['PILAR', 'POSIÇÃO', 'BITOLA (mm)', 'QTD', 'COMP. (m)', 'PESO (kg)']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=7, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Dados
            for row_idx, dado in enumerate(self.dados_processados, 8):
                for col_idx, valor in enumerate(dado, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx)
                    cell.value = valor
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # Alternar cores das linhas
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

            # Total
            row_total = len(self.dados_processados) + 9
            ws1.merge_cells(f'A{row_total}:C{row_total}')
            ws1[f'A{row_total}'] = 'TOTAL GERAL'
            ws1[f'A{row_total}'].font = Font(bold=True, size=12)
            ws1[f'A{row_total}'].alignment = Alignment(horizontal='right')
            ws1[f'A{row_total}'].fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

            ws1[f'D{row_total}'] = self.total_barras
            ws1[f'D{row_total}'].font = Font(bold=True)
            ws1[f'D{row_total}'].fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

            ws1[f'F{row_total}'] = f"{self.total_kg:.2f}"
            ws1[f'F{row_total}'].font = Font(bold=True)
            ws1[f'F{row_total}'].fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

            # Aba 2: Resumo por Bitola
            ws2 = wb.create_sheet("Resumo por Bitola")

            ws2['A1'] = 'RESUMO POR BITOLA'
            ws2['A1'].font = Font(size=14, bold=True)

            # Calcular resumo
            resumo_bitolas = {}
            for dado in self.dados_processados:
                bitola = dado[2]
                peso = dado[5]
                qtde = dado[3]
                if bitola not in resumo_bitolas:
                    resumo_bitolas[bitola] = {'peso': 0, 'qtde': 0}
                resumo_bitolas[bitola]['peso'] += peso
                resumo_bitolas[bitola]['qtde'] += qtde

            # Cabeçalhos
            ws2['A3'] = 'BITOLA (mm)'
            ws2['B3'] = 'QUANTIDADE'
            ws2['C3'] = 'PESO (kg)'
            ws2['D3'] = '% DO TOTAL'

            for cell in ['A3', 'B3', 'C3', 'D3']:
                ws2[cell].font = Font(bold=True, color="FFFFFF")
                ws2[cell].fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
                ws2[cell].alignment = Alignment(horizontal='center')

            # Dados do resumo
            row = 4
            for bitola in sorted(resumo_bitolas.keys()):
                ws2[f'A{row}'] = f"ø {bitola}"
                ws2[f'B{row}'] = resumo_bitolas[bitola]['qtde']
                ws2[f'C{row}'] = f"{resumo_bitolas[bitola]['peso']:.2f}"
                ws2[f'D{row}'] = f"{(resumo_bitolas[bitola]['peso']/self.total_kg*100):.1f}%"
                row += 1

            # Aba 3: Resumo por Tipo
            ws3 = wb.create_sheet("Resumo por Tipo")

            ws3['A1'] = 'RESUMO POR TIPO DE BARRA'
            ws3['A1'].font = Font(size=14, bold=True)

            # Calcular resumo por tipo
            resumo_tipos = {}
            for dado in self.dados_processados:
                pos = dado[1]
                bitola = dado[2]
                comp = dado[4]
                peso = dado[5]
                qtde = dado[3]

                tipo = AnalisadorGeometricoPilares.identificar_tipo_pilar(pos, bitola, comp)

                if tipo not in resumo_tipos:
                    resumo_tipos[tipo] = {'peso': 0, 'qtde': 0}
                resumo_tipos[tipo]['peso'] += peso
                resumo_tipos[tipo]['qtde'] += qtde

            # Cabeçalhos
            ws3['A3'] = 'TIPO'
            ws3['B3'] = 'QUANTIDADE'
            ws3['C3'] = 'PESO (kg)'
            ws3['D3'] = '% DO TOTAL'

            for cell in ['A3', 'B3', 'C3', 'D3']:
                ws3[cell].font = Font(bold=True, color="FFFFFF")
                ws3[cell].fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
                ws3[cell].alignment = Alignment(horizontal='center')

            # Dados do resumo por tipo
            row = 4
            for tipo in sorted(resumo_tipos.keys()):
                ws3[f'A{row}'] = tipo
                ws3[f'B{row}'] = resumo_tipos[tipo]['qtde']
                ws3[f'C{row}'] = f"{resumo_tipos[tipo]['peso']:.2f}"
                ws3[f'D{row}'] = f"{(resumo_tipos[tipo]['peso']/self.total_kg*100):.1f}%"
                row += 1

            # Ajustar largura das colunas em todas as abas
            for ws in [ws1, ws2, ws3]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Salvar arquivo
            arquivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=f"pilares_{self.var_pavimento.get()}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )

            if arquivo:
                wb.save(arquivo)
                messagebox.showinfo("Sucesso", f"Arquivo Excel salvo com sucesso!\n{arquivo}")

                if messagebox.askyesno("Abrir arquivo", "Deseja abrir o arquivo Excel agora?"):
                    os.startfile(arquivo)

        except ImportError:
            messagebox.showerror("Erro", "Biblioteca openpyxl não instalada!\n\nExecute: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar Excel:\n{str(e)}")

    def imprimir_direto(self):
        """Imprime relatório em HTML"""
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return

        try:
            # Montar conteúdo HTML
            linhas_html = []
            
            pilar_atual = None
            for dado in self.dados_processados:
                if dado[0] != pilar_atual:
                    pilar_atual = dado[0]
                    linhas_html.append(f'<tr class="pilar-header"><td colspan="5">{pilar_atual}</td></tr>')

                linhas_html.append(
                    f'<tr>'
                    f'<td>{dado[1]}</td>'
                    f'<td>Ø{dado[2]:.1f}mm</td>'
                    f'<td>{dado[3]}</td>'
                    f'<td>{dado[4]:.2f}m</td>'
                    f'<td>{dado[5]:.2f}kg</td>'
                    f'</tr>'
                )

            html_content = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <title>Relatório de Pilares</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: white;
        }}
        .header {{
            text-align: center;
            margin-bottom: 30px;
            border-bottom: 3px solid #2c3e50;
            padding-bottom: 15px;
        }}
        .header h1 {{
            color: #2c3e50;
            margin: 10px 0;
        }}
        .info {{
            background-color: #f8f9fa;
            padding: 15px;
            margin-bottom: 20px;
            border-left: 4px solid #3498db;
        }}
        .btn-container {{
            text-align: center;
            margin: 20px 0;
        }}
        .btn-print {{
            background-color: #27ae60;
            color: white;
            border: none;
            padding: 12px 30px;
            font-size: 16px;
            border-radius: 5px;
            cursor: pointer;
            margin: 10px;
        }}
        .btn-print:hover {{
            background-color: #229954;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
        }}
        th {{
            background-color: #34495e;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .pilar-header {{
            background-color: #3498db !important;
            color: white;
            font-weight: bold;
            font-size: 14px;
        }}
        .pilar-header td {{
            padding: 8px 12px;
        }}
        .totais {{
            background-color: #2c3e50;
            color: white;
            padding: 20px;
            margin-top: 20px;
            border-radius: 5px;
        }}
        .totais h3 {{
            margin: 0 0 10px 0;
        }}
        @media print {{
            body {{ margin: 0; }}
            .btn-container {{ display: none; }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📋 RELATÓRIO DE PILARES</h1>
        <p>EngenhariaPlanPro</p>
    </div>
    
    <div class="info">
        <strong>Obra:</strong> {self.var_obra.get()}<br>
        <strong>Pavimento:</strong> {self.var_pavimento.get()}<br>
        <strong>Data:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}
    </div>
    
    <div class="btn-container">
        <button class="btn-print" onclick="window.print();">🖨️ IMPRIMIR AGORA</button>
    </div>
    
    <table>
        <thead>
            <tr>
                <th>Posição</th>
                <th>Bitola</th>
                <th>Quantidade</th>
                <th>Comprimento</th>
                <th>Peso</th>
            </tr>
        </thead>
        <tbody>
            {''.join(linhas_html)}
        </tbody>
    </table>
    
    <div class="totais">
        <h3>📊 TOTAIS GERAIS</h3>
        <p><strong>Quantidade de barras:</strong> {self.total_barras}</p>
        <p><strong>Peso total:</strong> {self.total_kg:.2f} kg</p>
    </div>
</body>
</html>
"""

            # Criar arquivo HTML temporário
            html_temp = tempfile.mktemp(suffix=".html")
            with open(html_temp, 'w', encoding='utf-8') as f:
                f.write(html_content)

            # Abrir no navegador
            import webbrowser
            webbrowser.open('file://' + html_temp)
            print(f"[IMPRESSÃO HTML] Relatório aberto no navegador.")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório:\n{str(e)}")

    def limpar(self):
        """Limpa todos os dados"""
        self.arquivos_selecionados = []
        self.dados_processados = []
        self.total_kg = 0.0
        self.total_barras = 0
        self.tipos_personalizados = {}

        for item in self.tree.get_children():
            self.tree.delete(item)

        self.status_label.config(text="✅ Pronto para processar")
        self.info_label.config(text="")

    def imprimir_com_preview(self, conteudo):
        """Gera HTML do conteúdo texto e abre no navegador"""
        try:
            # Converter texto para HTML com formatação
            linhas = conteudo.split('\n')
            linhas_html = []
            
            for linha in linhas:
                linha = linha.replace(' ', '&nbsp;')
                if '===' in linha or '---' in linha:
                    linhas_html.append(f'<div class="separador">{linha}</div>')
                elif linha.strip().startswith('>>>'):
                    linhas_html.append(f'<div class="titulo-secao">{linha}</div>')
                elif 'TOTAL' in linha.upper() or 'ROMANEIO' in linha.upper():
                    linhas_html.append(f'<div class="destaque">{linha}</div>')
                else:
                    linhas_html.append(f'<div>{linha}</div>')

            html_content = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <title>Romaneio - Impressão</title>
    <style>
        body {{
            font-family: 'Courier New', monospace;
            margin: 20px;
            background-color: white;
            color: #333;
            line-height: 1.4;
        }}
        .header {{
            text-align: center;
            margin-bottom: 20px;
            padding: 15px;
            background-color: #2c3e50;
            color: white;
            border-radius: 5px;
        }}
        .btn-container {{
            text-align: center;
            margin: 20px 0;
        }}
        .btn-print {{
            background-color: #27ae60;
            color: white;
            border: none;
            padding: 12px 30px;
            font-size: 16px;
            border-radius: 5px;
            cursor: pointer;
            margin: 10px;
        }}
        .btn-print:hover {{
            background-color: #229954;
        }}
        .conteudo {{
            white-space: pre-wrap;
            font-size: 12px;
        }}
        .separador {{
            color: #3498db;
            font-weight: bold;
            margin: 5px 0;
        }}
        .titulo-secao {{
            color: #e74c3c;
            font-weight: bold;
            margin: 10px 0;
        }}
        .destaque {{
            background-color: #f39c12;
            color: white;
            padding: 5px;
            font-weight: bold;
            margin: 5px 0;
        }}
        @media print {{
            body {{ margin: 0; }}
            .header {{ background-color: #2c3e50; -webkit-print-color-adjust: exact; }}
            .destaque {{ background-color: #f39c12; -webkit-print-color-adjust: exact; }}
            .btn-container {{ display: none; }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h2>📄 ROMANEIO DE PILARES</h2>
    </div>
    
    <div class="btn-container">
        <button class="btn-print" onclick="window.print();">🖨️ IMPRIMIR AGORA</button>
    </div>
    
    <div class="conteudo">
        {''.join(linhas_html)}
    </div>
</body>
</html>
"""

            # Criar arquivo HTML temporário
            html_temp = tempfile.mktemp(suffix=".html")
            with open(html_temp, 'w', encoding='utf-8') as f:
                f.write(html_content)

            # Abrir no navegador
            import webbrowser
            webbrowser.open('file://' + html_temp)
            print(f"[IMPRESSÃO HTML] Romaneio aberto no navegador.")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar HTML:\n{str(e)}")

    def confirmar_impressao(self, conteudo, janela_preview):
        """Redireciona para imprimir_com_preview (HTML)"""
        janela_preview.destroy()
        self.imprimir_com_preview(conteudo)

    def salvar_txt(self, conteudo):
        """Salva conteúdo em arquivo texto"""
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Texto", "*.txt")],
            initialfile=f"romaneio_pilares_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
        )
        if arquivo:
            with open(arquivo, "w", encoding="utf-8") as f:
                f.write(conteudo)
            messagebox.showinfo("Sucesso", "Arquivo salvo com sucesso!")


# Executar aplicação
if __name__ == "__main__":
    app = PilaresApp()
    app.mainloop()