# lajes_app.py - Sistema Completo de Processamento de Lajes
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
from datetime import datetime
import os
import sys
import math

# Adicionar o diretório ao path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Importação condicional para evitar erro se o arquivo não existir
try:
    from core.lajes_motor import processar_lajes
except ImportError:
    # Função temporária caso o arquivo não exista
    def processar_lajes(arquivos):
        """Função temporária para processar lajes"""
        dados = []
        total_kg = 0.0
        total_barras = 0
        
        # Dados de exemplo para teste
        lajes_exemplo = [
            # (laje, tipo, especificacao, qtde, area/comp, peso)
            ("L1", "TELA", "Q196", 1, 25.50, 125.45),
            ("L1", "NEG", "N1 ø10.0", 15, 3.20, 28.42),
            ("L1", "NEG", "N2 ø10.0", 12, 2.80, 21.93),
            ("L1", "POS", "N3 ø8.0", 18, 4.50, 18.35),
            ("L2", "TELA", "Q283", 1, 32.40, 183.47),
            ("L2", "NEG", "N1 ø12.5", 20, 3.50, 48.85),
            ("L2", "NEG", "N2 ø12.5", 16, 3.20, 39.39),
            ("L2", "DIST", "N4 ø6.3", 25, 2.50, 12.25),
            ("L3", "TELA", "Q196", 1, 18.90, 92.98),
            ("L3", "NEG", "N1 ø10.0", 10, 2.80, 17.51),
            ("L3", "REFORCO", "N5 ø12.5", 8, 1.50, 11.82),
            ("L4", "TELA", "Q138", 1, 42.30, 145.67),
            ("L4", "NEG", "N1 ø8.0", 30, 3.00, 28.42),
            ("L4", "POS", "N2 ø10.0", 22, 4.20, 36.25),
        ]
        
        for laje in lajes_exemplo:
            dados.append(laje)
            total_kg += laje[5]
            if laje[1] != "TELA":
                total_barras += laje[3]
            else:
                total_barras += 1  # Tela conta como 1
        
        return dados, total_kg, total_barras

class AnalisadorGeometricoLajes:
    """Analisa e identifica tipos de armaduras para LAJES"""
    
    @staticmethod
    def identificar_tipo_laje(tipo_str, especificacao):
        """
        Identifica o tipo de armadura da laje baseado nos dados
        """
        tipo_upper = str(tipo_str).upper()
        espec_upper = str(especificacao).upper()
        
        # TELAS SOLDADAS
        if tipo_upper == "TELA" or "Q" in espec_upper:
            # Identificar tipo de tela
            if "Q92" in espec_upper or "Q-92" in espec_upper:
                return "TELA_Q92"
            elif "Q138" in espec_upper or "Q-138" in espec_upper:
                return "TELA_Q138"
            elif "Q196" in espec_upper or "Q-196" in espec_upper:
                return "TELA_Q196"
            elif "Q283" in espec_upper or "Q-283" in espec_upper:
                return "TELA_Q283"
            elif "Q396" in espec_upper or "Q-396" in espec_upper:
                return "TELA_Q396"
            else:
                return "TELA_ESPECIAL"
        
        # BARRAS NEGATIVAS
        elif tipo_upper in ["NEG", "NEGATIVO", "NEGATIVA"]:
            return "NEGATIVO"
        
        # BARRAS POSITIVAS
        elif tipo_upper in ["POS", "POSITIVO", "POSITIVA"]:
            return "POSITIVO"
        
        # ARMADURA DE DISTRIBUIÇÃO
        elif tipo_upper in ["DIST", "DISTRIBUICAO", "DISTRIBUIÇÃO"]:
            return "DISTRIBUICAO"
        
        # REFORÇOS
        elif tipo_upper in ["REFORCO", "REFORÇO", "REF"]:
            return "REFORCO"
        
        # ARMADURA DE BORDO
        elif tipo_upper in ["BORDO", "BORDA"]:
            return "BORDO"
        
        # ARMADURA DE FURO/ABERTURA
        elif tipo_upper in ["FURO", "ABERTURA"]:
            return "ABERTURA"
        
        # ARRANQUE
        elif tipo_upper in ["ARRANQUE", "ARR", "ESPERA"]:
            return "ARRANQUE"
        
        # Padrão
        else:
            return "GENERICO"
    
    @staticmethod
    def obter_especificacoes_tela(tipo_tela):
        """
        Retorna as especificações técnicas da tela
        """
        especificacoes = {
            "Q92": {
                "peso_m2": 1.48,
                "bitola": 4.2,
                "espacamento": 15,
                "area_aco": 0.92
            },
            "Q138": {
                "peso_m2": 2.20,
                "bitola": 5.0,
                "espacamento": 15,
                "area_aco": 1.38
            },
            "Q196": {
                "peso_m2": 3.11,
                "bitola": 6.0,
                "espacamento": 15,
                "area_aco": 1.96
            },
            "Q283": {
                "peso_m2": 4.48,
                "bitola": 7.0,
                "espacamento": 15,
                "area_aco": 2.83
            },
            "Q396": {
                "peso_m2": 6.28,
                "bitola": 8.0,
                "espacamento": 15,
                "area_aco": 3.96
            }
        }
        
        for key in especificacoes:
            if key in tipo_tela:
                return especificacoes[key]
        
        return {
            "peso_m2": 3.0,
            "bitola": 5.0,
            "espacamento": 15,
            "area_aco": 1.5
        }
    
    @staticmethod
    def calcular_sobreposicao_tela(tipo_tela):
        """
        Calcula a sobreposição necessária para a tela
        """
        # Sobreposição padrão: 2 malhas ou 25cm (o que for maior)
        espec = AnalisadorGeometricoLajes.obter_especificacoes_tela(tipo_tela)
        espacamento = espec["espacamento"]
        
        sobreposicao_malhas = 2 * espacamento  # cm
        sobreposicao_minima = 25  # cm
        
        return max(sobreposicao_malhas, sobreposicao_minima)

class LajesApp(tk.Tk):
    def __init__(self):
        super().__init__()
        
        self.title("EngenhariaPlanPro - LAJES")
        self.geometry("1200x700")
        self.configure(bg="#0d2818")
        
        # Centralizar
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - 600
        y = (self.winfo_screenheight() // 2) - 350
        self.geometry(f"1200x700+{x}+{y}")
        
        self.arquivos_selecionados = []
        self.dados_processados = []
        self.total_kg = 0.0
        self.total_barras = 0
        self.tipos_personalizados = {}
        
        self._criar_interface()
    
    def _criar_interface(self):
        # Container principal
        main_container = tk.Frame(self, bg="#0d2818")
        main_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Título
        header = tk.Frame(main_container, bg="#0d2818")
        header.pack(fill="x", pady=(0, 10))
        
        tk.Label(
            header,
            text="🏢 PROCESSAMENTO DE LAJES",
            font=("Arial", 20, "bold"),
            bg="#0d2818",
            fg="#00bcd4"
        ).pack()
        
        # Frame de controles
        control_frame = tk.Frame(main_container, bg="#1a3d2e")
        control_frame.pack(fill="x", pady=5)
        
        # Linha 1 - Dados do projeto
        row1 = tk.Frame(control_frame, bg="#1a3d2e")
        row1.pack(fill="x", padx=10, pady=10)
        
        tk.Label(row1, text="Obra:", bg="#1a3d2e", fg="white", font=("Arial", 10)).pack(side="left", padx=5)
        self.var_obra = tk.StringVar(value="OBRA 001")
        tk.Entry(row1, textvariable=self.var_obra, width=25, font=("Arial", 10)).pack(side="left", padx=5)
        
        tk.Label(row1, text="Pavimento:", bg="#1a3d2e", fg="white", font=("Arial", 10)).pack(side="left", padx=20)
        self.var_pavimento = tk.StringVar(value="TIPO")
        tk.Entry(row1, textvariable=self.var_pavimento, width=20, font=("Arial", 10)).pack(side="left", padx=5)
        
        tk.Label(row1, text="Tipo Laje:", bg="#1a3d2e", fg="white", font=("Arial", 10)).pack(side="left", padx=20)
        self.var_tipo_laje = ttk.Combobox(row1, values=["Maciça", "Nervurada", "Protendida", "Steel Deck"], width=15)
        self.var_tipo_laje.set("Maciça")
        self.var_tipo_laje.pack(side="left", padx=5)
        
        # Linha 2 - Botões principais
        row2 = tk.Frame(control_frame, bg="#1a3d2e")
        row2.pack(fill="x", padx=10, pady=(0, 10))
        
        tk.Button(
            row2,
            text="📁 Selecionar Arquivos",
            command=self.selecionar_arquivos,
            bg="#00acc1",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        tk.Button(
            row2,
            text="⚙️ PROCESSAR",
            command=self.processar,
            bg="#00bcd4",
            fg="white",
            font=("Arial", 11, "bold"),
            padx=20,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        tk.Button(
            row2,
            text="📄 Gerar Romaneio",
            command=self.gerar_romaneio,
            bg="#1976d2",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        tk.Button(
            row2,
            text="🏷️ Etiquetas",
            command=self.gerar_etiquetas,
            bg="#9c27b0",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        tk.Button(
            row2,
            text="📐 Plano de Corte",
            command=self.gerar_plano_corte,
            bg="#e91e63",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        tk.Button(
            row2,
            text="📊 Exportar Excel",
            command=self.exportar_excel,
            bg="#4caf50",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        tk.Button(
            row2,
            text="🖨️ Imprimir",
            command=self.imprimir_direto,
            bg="#6a1b9a",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        tk.Button(
            row2,
            text="🔄 Limpar",
            command=self.limpar,
            bg="#757575",
            fg="white",
            font=("Arial", 10),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        # Frame da tabela
        table_frame = tk.Frame(main_container, bg="#0d2818")
        table_frame.pack(fill="both", expand=True, pady=10)
        
        # Criar Treeview com scrollbars
        tree_scroll_y = ttk.Scrollbar(table_frame)
        tree_scroll_y.pack(side="right", fill="y")
        
        tree_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal")
        tree_scroll_x.pack(side="bottom", fill="x")
        
        # Configurar estilo
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="#1e1e1e", foreground="white", fieldbackground="#1e1e1e")
        style.configure("Treeview.Heading", background="#00acc1", foreground="white", font=("Arial", 10, "bold"))
        
        # Criar Treeview
        self.tree = ttk.Treeview(
            table_frame,
            columns=("laje", "tipo", "especificacao", "qtde", "area_comp", "peso"),
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set
        )
        
        # Configurar colunas
        self.tree.heading("laje", text="LAJE")
        self.tree.heading("tipo", text="TIPO")
        self.tree.heading("especificacao", text="ESPECIFICAÇÃO")
        self.tree.heading("qtde", text="QUANTIDADE")
        self.tree.heading("area_comp", text="ÁREA/COMP")
        self.tree.heading("peso", text="PESO (kg)")
        
        self.tree.column("laje", width=100, anchor="center")
        self.tree.column("tipo", width=100, anchor="center")
        self.tree.column("especificacao", width=150, anchor="center")
        self.tree.column("qtde", width=100, anchor="center")
        self.tree.column("area_comp", width=100, anchor="center")
        self.tree.column("peso", width=100, anchor="center")
        
        self.tree.pack(fill="both", expand=True)
        
        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)
        
        # Status bar
        self.status_frame = tk.Frame(main_container, bg="#1a3d2e", relief="sunken", bd=1)
        self.status_frame.pack(fill="x", pady=(5, 0))
        
        self.status_label = tk.Label(
            self.status_frame,
            text="✅ Pronto para processar",
            bg="#1a3d2e",
            fg="white",
            font=("Arial", 10),
            anchor="w"
        )
        self.status_label.pack(side="left", padx=10, pady=5)
        
        self.info_label = tk.Label(
            self.status_frame,
            text="",
            bg="#1a3d2e",
            fg="#00bcd4",
            font=("Arial", 10, "bold"),
            anchor="e"
        )
        self.info_label.pack(side="right", padx=10, pady=5)
    
    def selecionar_arquivos(self):
        """Seleciona arquivos DXF/DWG"""
        arquivos = filedialog.askopenfilenames(
            title="Selecionar arquivos DXF/DWG de LAJES",
            filetypes=[
                ("Arquivos CAD", "*.dxf *.dwg"),
                ("DXF", "*.dxf"),
                ("DWG", "*.dwg"),
                ("Todos", "*.*")
            ]
        )
        
        if arquivos:
            self.arquivos_selecionados = list(arquivos)
            self.status_label.config(text=f"📁 {len(arquivos)} arquivo(s) selecionado(s)")
            
            nomes = [os.path.basename(f) for f in arquivos]
            if len(nomes) > 3:
                texto = f"{', '.join(nomes[:3])}, ..."
            else:
                texto = ', '.join(nomes)
            self.info_label.config(text=texto)
    
    def processar(self):
        """Processa os arquivos selecionados"""
        if not self.arquivos_selecionados:
            messagebox.showwarning("Atenção", "Por favor, selecione os arquivos DXF/DWG primeiro!")
            return
        
        try:
            self.status_label.config(text="⏳ Processando arquivos...")
            self.update()
            
            self.dados_processados, self.total_kg, self.total_barras = processar_lajes(self.arquivos_selecionados)
            
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            for dado in self.dados_processados:
                self.tree.insert("", "end", values=dado)
            
            self.status_label.config(text=f"✅ Processamento concluído")
            self.info_label.config(text=f"Total: {self.total_barras} itens | {self.total_kg:.2f} kg")
            
            if len(self.dados_processados) == 0:
                messagebox.showinfo("Informação", "Nenhuma laje foi encontrada nos arquivos.\n\nVerifique se o formato está correto:\n- Lajes: L1, L2, L3...\n- Telas: Q196, Q283...\n- Barras: N1, N2...")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar:\n{str(e)}")
            self.status_label.config(text="❌ Erro no processamento")
    
    def gerar_plano_corte(self):
        """Gera plano de corte otimizado para telas e barras"""
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return
        
        janela = tk.Toplevel(self)
        janela.title("📐 Plano de Corte - LAJES")
        janela.geometry("900x700")
        janela.configure(bg="#0d2818")
        
        janela.update_idletasks()
        x = (janela.winfo_screenwidth() // 2) - 450
        y = (janela.winfo_screenheight() // 2) - 350
        janela.geometry(f"900x700+{x}+{y}")
        
        # Frame superior com botões
        btn_frame = tk.Frame(janela, bg="#1a3d2e")
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(
            btn_frame,
            text="💾 Salvar",
            command=lambda: self.salvar_txt(texto.get("1.0", "end-1c")),
            bg="#2e7d32",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=10
        ).pack(side="left", padx=5)
        
        tk.Button(
            btn_frame,
            text="🖨️ Imprimir",
            command=lambda: self.imprimir_com_preview(texto.get("1.0", "end-1c")),
            bg="#6a1b9a",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=10
        ).pack(side="left", padx=5)
        
        tk.Button(
            btn_frame,
            text="❌ Fechar",
            command=janela.destroy,
            bg="#757575",
            fg="white",
            font=("Arial", 10),
            padx=10
        ).pack(side="right", padx=5)
        
        # ScrolledText para o plano
        texto = ScrolledText(
            janela,
            bg="#1e1e1e",
            fg="#00bcd4",
            font=("Courier New", 10),
            wrap="none"
        )
        texto.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Gerar conteúdo do plano de corte
        conteudo = []
        conteudo.append("=" * 80)
        conteudo.append("                    PLANO DE CORTE - LAJES")
        conteudo.append("=" * 80)
        conteudo.append(f"Obra:      {self.var_obra.get()}")
        conteudo.append(f"Pavimento: {self.var_pavimento.get()}")
        conteudo.append(f"Tipo:      {self.var_tipo_laje.get()}")
        conteudo.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        conteudo.append("=" * 80)
        conteudo.append("")
        
        # Separar telas e barras
        telas = []
        barras = []
        
        for dado in self.dados_processados:
            if dado[1] == "TELA":
                telas.append(dado)
            else:
                barras.append(dado)
        
        # SEÇÃO 1: TELAS SOLDADAS
        if telas:
            conteudo.append("╔" + "═" * 78 + "╗")
            conteudo.append("║" + " TELAS SOLDADAS".center(78) + "║")
            conteudo.append("╚" + "═" * 78 + "╝")
            conteudo.append("")
            
            # Agrupar telas por tipo
            telas_agrupadas = {}
            for tela in telas:
                tipo = tela[2]
                if tipo not in telas_agrupadas:
                    telas_agrupadas[tipo] = []
                telas_agrupadas[tipo].append(tela)
            
            for tipo_tela in sorted(telas_agrupadas.keys()):
                conteudo.append(f"TIPO: {tipo_tela}")
                conteudo.append("-" * 40)
                
                area_total = 0
                peso_total = 0
                
                for tela in telas_agrupadas[tipo_tela]:
                    laje = tela[0]
                    area = tela[4]
                    peso = tela[5]
                    
                    conteudo.append(f"  {laje}: Área = {area:.2f} m² | Peso = {peso:.2f} kg")
                    area_total += area
                    peso_total += peso
                
                # Calcular painéis necessários (padrão 2,45m x 6,00m = 14,7m²)
                area_painel = 14.7
                paineis_necessarios = math.ceil(area_total / area_painel)
                
                # Considerar sobreposição
                sobreposicao = AnalisadorGeometricoLajes.calcular_sobreposicao_tela(tipo_tela)
                area_com_sobreposicao = area_total * 1.15  # 15% de sobreposição
                paineis_com_sobreposicao = math.ceil(area_com_sobreposicao / area_painel)
                
                conteudo.append("")
                conteudo.append(f"  Área total: {area_total:.2f} m²")
                conteudo.append(f"  Área com sobreposição (15%): {area_com_sobreposicao:.2f} m²")
                conteudo.append(f"  Painéis necessários (2,45x6,00m): {paineis_com_sobreposicao}")
                conteudo.append(f"  Sobreposição recomendada: {sobreposicao} cm")
                conteudo.append(f"  Peso total: {peso_total:.2f} kg")
                conteudo.append("")
        
        # SEÇÃO 2: BARRAS DE AÇO
        if barras:
            conteudo.append("╔" + "═" * 78 + "╗")
            conteudo.append("║" + " BARRAS DE AÇO".center(78) + "║")
            conteudo.append("╚" + "═" * 78 + "╝")
            conteudo.append("")
            
            # Agrupar por tipo e bitola
            barras_agrupadas = {}
            for barra in barras:
                tipo = barra[1]
                espec = barra[2]
                
                # Extrair bitola da especificação
                bitola = 0
                if "ø" in espec:
                    try:
                        bitola = float(espec.split("ø")[1].split()[0])
                    except:
                        bitola = 0
                
                chave = f"{tipo}_{bitola}"
                if chave not in barras_agrupadas:
                    barras_agrupadas[chave] = []
                barras_agrupadas[chave].append(barra)
            
            # Ordenar por tipo e bitola
            for chave in sorted(barras_agrupadas.keys()):
                tipo, bitola = chave.split("_")
                bitola = float(bitola)
                
                conteudo.append(f"TIPO: {tipo} - BITOLA: ø{bitola:.1f}mm")
                conteudo.append("-" * 40)
                
                total_barras_tipo = 0
                total_metros = 0
                total_peso = 0
                
                for barra in barras_agrupadas[chave]:
                    laje = barra[0]
                    qtde = barra[3]
                    comp = barra[4]
                    peso = barra[5]
                    
                    conteudo.append(f"  {laje}: {qtde} barras x {comp:.2f}m = {qtde*comp:.2f}m | {peso:.2f}kg")
                    total_barras_tipo += qtde
                    total_metros += qtde * comp
                    total_peso += peso
                
                # Calcular barras de 12m necessárias
                barras_12m = math.ceil(total_metros / 12)
                
                conteudo.append("")
                conteudo.append(f"  Total: {total_barras_tipo} barras")
                conteudo.append(f"  Metragem total: {total_metros:.2f}m")
                conteudo.append(f"  Barras de 12m necessárias: {barras_12m}")
                conteudo.append(f"  Peso total: {total_peso:.2f}kg")
                conteudo.append("")
        
        conteudo.append("=" * 80)
        conteudo.append("RESUMO DO PLANO DE CORTE")
        conteudo.append("-" * 40)
        conteudo.append(f"Total geral: {self.total_kg:.2f} kg")
        conteudo.append("=" * 80)
        
        texto.insert("1.0", "\n".join(conteudo))
    
    def gerar_etiquetas(self):
        """Gera etiquetas de identificação para lajes"""
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return
        
        # Criar janela
        self.janela_etiq = tk.Toplevel(self)
        self.janela_etiq.title("Etiquetas de Identificação - LAJES")
        self.janela_etiq.geometry("900x750")
        self.janela_etiq.configure(bg="#f5f5f5")
        
        # Variáveis de controle de página
        self.pagina_atual = 0
        self.etiquetas_por_pagina = 4
        self.total_paginas = (len(self.dados_processados) + self.etiquetas_por_pagina - 1) // self.etiquetas_por_pagina
        
        # Frame superior com informações
        info_frame = tk.Frame(self.janela_etiq, bg="#00acc1")
        info_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Label(
            info_frame,
            text=f"Total de Etiquetas: {len(self.dados_processados)} | Total de Páginas: {self.total_paginas}",
            bg="#00acc1",
            fg="white",
            font=("Arial", 10)
        ).pack(pady=5)
        
        # Frame de navegação
        nav_frame = tk.Frame(self.janela_etiq, bg="#2c3e50")
        nav_frame.pack(fill="x", padx=5, pady=5)
        
        # Botões de navegação
        self.btn_primeira = tk.Button(
            nav_frame,
            text="⏮️ Primeira",
            command=self.primeira_pagina,
            bg="#34495e",
            fg="white",
            font=("Arial", 10, "bold"),
            width=10
        )
        self.btn_primeira.pack(side="left", padx=5, pady=5)
        
        self.btn_anterior = tk.Button(
            nav_frame,
            text="◀ Anterior",
            command=self.pagina_anterior,
            bg="#2196f3",
            fg="white",
            font=("Arial", 10, "bold"),
            width=10
        )
        self.btn_anterior.pack(side="left", padx=5, pady=5)
        
        self.label_pagina = tk.Label(
            nav_frame,
            text=f"Página {self.pagina_atual + 1} de {self.total_paginas}",
            bg="#2c3e50",
            fg="white",
            font=("Arial", 12, "bold"),
            width=20
        )
        self.label_pagina.pack(side="left", padx=20, pady=5)
        
        self.btn_proxima = tk.Button(
            nav_frame,
            text="Próxima ▶",
            command=self.proxima_pagina,
            bg="#2196f3",
            fg="white",
            font=("Arial", 10, "bold"),
            width=10
        )
        self.btn_proxima.pack(side="left", padx=5, pady=5)
        
        self.btn_ultima = tk.Button(
            nav_frame,
            text="Última ⏭️",
            command=self.ultima_pagina,
            bg="#34495e",
            fg="white",
            font=("Arial", 10, "bold"),
            width=10
        )
        self.btn_ultima.pack(side="left", padx=5, pady=5)
        
        # Canvas para etiquetas
        self.canvas_etiq = tk.Canvas(self.janela_etiq, bg="white", width=850, height=600)
        self.canvas_etiq.pack(pady=5)
        
        # Desenhar primeira página
        self.desenhar_pagina_etiquetas_lajes()
        self.atualizar_botoes_navegacao()
    
    def desenhar_pagina_etiquetas_lajes(self):
        """Desenha a página atual de etiquetas para lajes"""
        # Limpar canvas
        self.canvas_etiq.delete("all")
        
        # Configuração das etiquetas
        largura_etiq = 400
        altura_etiq = 280
        margem = 25
        
        # Calcular índices
        inicio = self.pagina_atual * self.etiquetas_por_pagina
        fim = min(inicio + self.etiquetas_por_pagina, len(self.dados_processados))
        
        # Título da página
        self.canvas_etiq.create_text(
            425, 10,
            text=f"PÁGINA {self.pagina_atual + 1} DE {self.total_paginas}",
            font=("Arial", 8),
            fill="gray"
        )
        
        # Desenhar etiquetas da página atual
        for i, idx in enumerate(range(inicio, fim)):
            dado = self.dados_processados[idx]
            
            # Posição na página (2x2)
            col = i % 2
            row = i // 2
            x = margem + col * (largura_etiq + margem)
            y = margem + row * (altura_etiq + margem) + 20
            
            # Extrair dados
            laje = dado[0]
            tipo = dado[1]
            especificacao = dado[2]
            qtde = dado[3]
            area_comp = dado[4]
            peso = dado[5]
            
            # Identificar tipo específico
            tipo_especifico = AnalisadorGeometricoLajes.identificar_tipo_laje(tipo, especificacao)
            
            # DESENHAR ETIQUETA
            # Borda com cor baseada no tipo
            cor_borda = "#00acc1"  # Padrão cyan
            if "TELA" in tipo_especifico:
                cor_borda = "#4caf50"  # Verde para telas
            elif "NEG" in tipo_especifico:
                cor_borda = "#f44336"  # Vermelho para negativos
            elif "POS" in tipo_especifico:
                cor_borda = "#2196f3"  # Azul para positivos
            
            self.canvas_etiq.create_rectangle(x, y, x+largura_etiq, y+altura_etiq, 
                                             outline=cor_borda, width=3, fill="white")
            
            # Número da etiqueta
            self.canvas_etiq.create_text(
                x + 10, y + 10,
                text=f"#{idx + 1}",
                font=("Arial", 8),
                fill="gray",
                anchor="w"
            )
            
            # Linha divisória horizontal
            self.canvas_etiq.create_line(x, y+45, x+largura_etiq, y+45, width=1, fill=cor_borda)
            
            # CABEÇALHO
            self.canvas_etiq.create_text(x+largura_etiq/2, y+22,
                                        text=f"OBRA: {self.var_obra.get()} - PAV: {self.var_pavimento.get()}",
                                        font=("Arial", 11, "bold"), fill="black")
            
            # LAJE E TIPO
            self.canvas_etiq.create_text(x+60, y+65,
                                        text=f"LAJE: {laje}",
                                        font=("Arial", 10, "bold"), fill="black", anchor="w")
            
            self.canvas_etiq.create_text(x+largura_etiq-60, y+65,
                                        text=f"TIPO: {tipo}",
                                        font=("Arial", 10, "bold"), fill="black", anchor="e")
            
            # ESPECIFICAÇÃO
            self.canvas_etiq.create_text(x+largura_etiq/2, y+90,
                                        text=f"{especificacao}",
                                        font=("Arial", 12, "bold"), fill=cor_borda)
            
            # DADOS (lado esquerdo)
            self.canvas_etiq.create_text(x+30, y+120,
                                        text=f"Quantidade: {qtde}",
                                        font=("Arial", 9), fill="black", anchor="w")
            
            if tipo == "TELA":
                self.canvas_etiq.create_text(x+30, y+140,
                                            text=f"Área: {area_comp:.2f} m²",
                                            font=("Arial", 9), fill="black", anchor="w")
            else:
                self.canvas_etiq.create_text(x+30, y+140,
                                            text=f"Comp: {area_comp:.2f} m",
                                            font=("Arial", 9), fill="black", anchor="w")
            
            self.canvas_etiq.create_text(x+30, y+160,
                                        text=f"Peso: {peso:.2f} kg",
                                        font=("Arial", 9, "bold"), fill="blue", anchor="w")
            
            # ÁREA DE DESENHO (centro-direita)
            desenho_x = x + 180
            desenho_y = y + 120
            
            # DESENHAR CONFORME O TIPO
            if "TELA" in tipo_especifico:
                # Desenhar grade representando tela
                for i in range(5):
                    # Linhas horizontais
                    self.canvas_etiq.create_line(desenho_x, desenho_y + i*15, 
                                                desenho_x + 120, desenho_y + i*15,
                                                width=1, fill="gray")
                    # Linhas verticais
                    self.canvas_etiq.create_line(desenho_x + i*30, desenho_y, 
                                                desenho_x + i*30, desenho_y + 60,
                                                width=1, fill="gray")
                
                # Especificações da tela
                if "Q" in especificacao:
                    espec = AnalisadorGeometricoLajes.obter_especificacoes_tela(especificacao)
                    self.canvas_etiq.create_text(desenho_x + 60, desenho_y + 80,
                                                text=f"ø{espec['bitola']}mm c/{espec['espacamento']}cm",
                                                font=("Arial", 8), fill="green")
            
            elif "NEG" in tipo_especifico:
                # Desenhar barra negativa (com dobras para cima)
                self.canvas_etiq.create_line(desenho_x, desenho_y+30, desenho_x, desenho_y,                                            width=4, fill="red")
                self.canvas_etiq.create_line(desenho_x, desenho_y, desenho_x+100, desenho_y,
                                            width=4, fill="red")
                self.canvas_etiq.create_line(desenho_x+100, desenho_y, desenho_x+100, desenho_y+30,
                                            width=4, fill="red")
                
                self.canvas_etiq.create_text(desenho_x + 50, desenho_y + 50,
                                            text="NEGATIVO",
                                            font=("Arial", 8), fill="red")
            
            elif "POS" in tipo_especifico:
                # Desenhar barra positiva (reta com ganchos)
                self.canvas_etiq.create_line(desenho_x, desenho_y+10, desenho_x, desenho_y+20,
                                            width=3, fill="blue")
                self.canvas_etiq.create_line(desenho_x, desenho_y+20, desenho_x+100, desenho_y+20,
                                            width=4, fill="blue")
                self.canvas_etiq.create_line(desenho_x+100, desenho_y+20, desenho_x+100, desenho_y+10,
                                            width=3, fill="blue")
                
                self.canvas_etiq.create_text(desenho_x + 50, desenho_y + 50,
                                            text="POSITIVO",
                                            font=("Arial", 8), fill="blue")
            
            elif "DIST" in tipo_especifico:
                # Desenhar barras de distribuição (paralelas)
                for i in range(3):
                    self.canvas_etiq.create_line(desenho_x, desenho_y + i*15,
                                                desenho_x+100, desenho_y + i*15,
                                                width=2, fill="purple")
                
                self.canvas_etiq.create_text(desenho_x + 50, desenho_y + 50,
                                            text="DISTRIBUIÇÃO",
                                            font=("Arial", 8), fill="purple")
            
            elif "REFORCO" in tipo_especifico or "REFORÇO" in tipo_especifico:
                # Desenhar reforço (X)
                self.canvas_etiq.create_line(desenho_x, desenho_y, desenho_x+80, desenho_y+40,
                                            width=3, fill="orange")
                self.canvas_etiq.create_line(desenho_x, desenho_y+40, desenho_x+80, desenho_y,
                                            width=3, fill="orange")
                
                self.canvas_etiq.create_text(desenho_x + 40, desenho_y + 55,
                                            text="REFORÇO",
                                            font=("Arial", 8), fill="orange")
            
            # CÓDIGO DE BARRAS (parte inferior)
            codigo_y = y + altura_etiq - 40
            codigo_x = x + (largura_etiq - 160)/2
            
            # Barras verticais
            for j in range(20):
                bar_w = 3 if j % 2 == 0 else 2
                bar_x = codigo_x + j * 8
                self.canvas_etiq.create_line(bar_x, codigo_y, bar_x, codigo_y+20,
                                            width=bar_w, fill="black")
            
            # Texto do código
            codigo = f"{laje}{tipo}{idx:03d}"
            self.canvas_etiq.create_text(x+largura_etiq/2, codigo_y+28,
                                        text=codigo, font=("Courier", 9), fill="black")
    
    def gerar_romaneio(self):
        """Gera romaneio detalhado com múltiplas visualizações"""
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return
        
        janela = tk.Toplevel(self)
        janela.title("Romaneio de LAJES - Múltiplas Visualizações")
        janela.geometry("1000x700")
        janela.configure(bg="#0d2818")
        
        janela.update_idletasks()
        x = (janela.winfo_screenwidth() // 2) - 500
        y = (janela.winfo_screenheight() // 2) - 350
        janela.geometry(f"1000x700+{x}+{y}")
        
        # Frame superior com botões
        btn_frame = tk.Frame(janela, bg="#1a3d2e")
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        # Variável para controlar aba ativa
        self.aba_ativa = None
        self.texto_ativo = None
        
        def salvar_aba_ativa():
            if self.texto_ativo:
                self.salvar_txt(self.texto_ativo.get("1.0", "end-1c"))
        
        def imprimir_aba_ativa():
            if self.texto_ativo:
                self.imprimir_com_preview(self.texto_ativo.get("1.0", "end-1c"))
        
        tk.Button(
            btn_frame,
            text="💾 Salvar TXT",
            command=salvar_aba_ativa,
            bg="#2e7d32",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=10
        ).pack(side="left", padx=5)
        
        tk.Button(
            btn_frame,
            text="🖨️ Imprimir",
            command=imprimir_aba_ativa,
            bg="#6a1b9a",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=10
        ).pack(side="left", padx=5)
        
        tk.Button(
            btn_frame,
            text="📊 Exportar Tudo Excel",
            command=self.exportar_excel,
            bg="#00acc1",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=10
        ).pack(side="left", padx=5)
        
        tk.Button(
            btn_frame,
            text="❌ Fechar",
            command=janela.destroy,
            bg="#757575",
            fg="white",
            font=("Arial", 10),
            padx=10
        ).pack(side="right", padx=5)
        
        # Notebook para abas
        notebook = ttk.Notebook(janela)
        notebook.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Estilo para o notebook
        style = ttk.Style()
        style.configure('TNotebook.Tab', padding=[20, 10])
        
        # ========== ABA 1: ROMANEIO GERAL ==========
        frame_geral = tk.Frame(notebook, bg="#1e1e1e")
        notebook.add(frame_geral, text="📋 Romaneio Geral")
        
        texto_geral = ScrolledText(
            frame_geral,
            bg="#1e1e1e",
            fg="#00bcd4",
            font=("Courier New", 10),
            wrap="none"
        )
        texto_geral.pack(fill="both", expand=True)
        
        # Gerar conteúdo do romaneio geral
        conteudo_geral = []
        conteudo_geral.append("=" * 80)
        conteudo_geral.append("                    ROMANEIO DE LAJES - GERAL")
        conteudo_geral.append("=" * 80)
        conteudo_geral.append(f"Obra:      {self.var_obra.get()}")
        conteudo_geral.append(f"Pavimento: {self.var_pavimento.get()}")
        conteudo_geral.append(f"Tipo Laje: {self.var_tipo_laje.get()}")
        conteudo_geral.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        conteudo_geral.append("=" * 80)
        conteudo_geral.append("")
        
        laje_atual = None
        for dado in self.dados_processados:
            if dado[0] != laje_atual:
                laje_atual = dado[0]
                conteudo_geral.append("")
                conteudo_geral.append(f">>> {laje_atual}")
                conteudo_geral.append("-" * 60)
            
            tipo = dado[1]
            espec = dado[2]
            qtde = dado[3]
            area_comp = dado[4]
            peso = dado[5]
            
            if tipo == "TELA":
                conteudo_geral.append(
                    f"    {tipo:<10} {espec:<15} Área: {area_comp:>6.2f}m²  Peso: {peso:>8.2f}kg"
                )
            else:
                conteudo_geral.append(
                    f"    {tipo:<10} {espec:<15} Qtd: {qtde:>3}  Comp: {area_comp:>5.2f}m  Peso: {peso:>8.2f}kg"
                )
        
        conteudo_geral.append("")
        conteudo_geral.append("=" * 80)
        conteudo_geral.append(f"TOTAL GERAL: {self.total_kg:.2f} kg")
        conteudo_geral.append("=" * 80)
        
        texto_geral.insert("1.0", "\n".join(conteudo_geral))
        
        # ========== ABA 2: TELAS SOLDADAS ==========
        frame_telas = tk.Frame(notebook, bg="#1e1e1e")
        notebook.add(frame_telas, text="🔲 Telas Soldadas")
        
        texto_telas = ScrolledText(
            frame_telas,
            bg="#1e1e1e",
            fg="#4caf50",
            font=("Courier New", 10),
            wrap="none"
        )
        texto_telas.pack(fill="both", expand=True)
        
        # Filtrar apenas telas
        telas = [d for d in self.dados_processados if d[1] == "TELA"]
        
        conteudo_telas = []
        conteudo_telas.append("=" * 80)
        conteudo_telas.append("                    ROMANEIO DE TELAS SOLDADAS")
        conteudo_telas.append("=" * 80)
        conteudo_telas.append(f"Obra:      {self.var_obra.get()}")
        conteudo_telas.append(f"Pavimento: {self.var_pavimento.get()}")
        conteudo_telas.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        conteudo_telas.append("=" * 80)
        conteudo_telas.append("")
        
        if telas:
            # Agrupar por tipo de tela
            telas_por_tipo = {}
            for tela in telas:
                tipo = tela[2]
                if tipo not in telas_por_tipo:
                    telas_por_tipo[tipo] = []
                telas_por_tipo[tipo].append(tela)
            
            for tipo_tela in sorted(telas_por_tipo.keys()):
                conteudo_telas.append(f"╔{'═' * 78}╗")
                conteudo_telas.append(f"║ TELA: {tipo_tela:<70} ║")
                conteudo_telas.append(f"╚{'═' * 78}╝")
                conteudo_telas.append("")
                
                # Especificações da tela
                espec = AnalisadorGeometricoLajes.obter_especificacoes_tela(tipo_tela)
                conteudo_telas.append(f"  Especificações:")
                conteudo_telas.append(f"    - Bitola: ø{espec['bitola']}mm")
                conteudo_telas.append(f"    - Espaçamento: {espec['espacamento']}cm")
                conteudo_telas.append(f"    - Peso/m²: {espec['peso_m2']}kg")
                conteudo_telas.append(f"    - Área de aço: {espec['area_aco']}cm²/m")
                conteudo_telas.append("")
                
                area_total = 0
                peso_total = 0
                
                conteudo_telas.append("  Lajes:")
                for tela in telas_por_tipo[tipo_tela]:
                    laje = tela[0]
                    area = tela[4]
                    peso = tela[5]
                    
                    conteudo_telas.append(f"    {laje}: {area:.2f}m² = {peso:.2f}kg")
                    area_total += area
                    peso_total += peso
                
                # Calcular painéis
                area_painel = 14.7  # 2,45m x 6,00m
                paineis = math.ceil(area_total * 1.15 / area_painel)  # 15% sobreposição
                
                conteudo_telas.append("")
                conteudo_telas.append(f"  Subtotal {tipo_tela}:")
                conteudo_telas.append(f"    - Área total: {area_total:.2f}m²")
                conteudo_telas.append(f"    - Área com sobreposição (15%): {area_total*1.15:.2f}m²")
                conteudo_telas.append(f"    - Painéis (2,45x6,00m): {paineis}")
                conteudo_telas.append(f"    - Peso total: {peso_total:.2f}kg")
                conteudo_telas.append("")
        else:
            conteudo_telas.append("Nenhuma tela soldada encontrada.")
        
        conteudo_telas.append("=" * 80)
        texto_telas.insert("1.0", "\n".join(conteudo_telas))
        
        # ========== ABA 3: BARRAS DE AÇO ==========
        frame_barras = tk.Frame(notebook, bg="#1e1e1e")
        notebook.add(frame_barras, text="📊 Barras de Aço")
        
        texto_barras = ScrolledText(
            frame_barras,
            bg="#1e1e1e",
            fg="#ff9800",
            font=("Courier New", 10),
            wrap="none"
        )
        texto_barras.pack(fill="both", expand=True)
        
        # Filtrar apenas barras
        barras = [d for d in self.dados_processados if d[1] != "TELA"]
        
        conteudo_barras = []
        conteudo_barras.append("=" * 80)
        conteudo_barras.append("                    ROMANEIO DE BARRAS DE AÇO")
        conteudo_barras.        conteudo_barras.append("=" * 80)
        conteudo_barras.append(f"Obra:      {self.var_obra.get()}")
        conteudo_barras.append(f"Pavimento: {self.var_pavimento.get()}")
        conteudo_barras.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        conteudo_barras.append("=" * 80)
        conteudo_barras.append("")
        
        if barras:
            # Agrupar por tipo de barra
            barras_por_tipo = {}
            for barra in barras:
                tipo = barra[1]
                if tipo not in barras_por_tipo:
                    barras_por_tipo[tipo] = []
                barras_por_tipo[tipo].append(barra)
            
            for tipo_barra in sorted(barras_por_tipo.keys()):
                conteudo_barras.append(f"╔{'═' * 78}╗")
                conteudo_barras.append(f"║ TIPO: {tipo_barra:<70} ║")
                conteudo_barras.append(f"╚{'═' * 78}╝")
                conteudo_barras.append("")
                
                total_barras_tipo = 0
                total_metros = 0
                total_peso = 0
                
                for barra in barras_por_tipo[tipo_barra]:
                    laje = barra[0]
                    espec = barra[2]
                    qtde = barra[3]
                    comp = barra[4]
                    peso = barra[5]
                    
                    conteudo_barras.append(
                        f"  {laje}: {espec:<15} Qtd: {qtde:>3} x {comp:.2f}m = {peso:.2f}kg"
                    )
                    total_barras_tipo += qtde
                    total_metros += qtde * comp
                    total_peso += peso
                
                conteudo_barras.append("")
                conteudo_barras.append(f"  Subtotal {tipo_barra}:")
                conteudo_barras.append(f"    - Total de barras: {total_barras_tipo}")
                conteudo_barras.append(f"    - Metragem total: {total_metros:.2f}m")
                conteudo_barras.append(f"    - Barras de 12m: {math.ceil(total_metros/12)}")
                conteudo_barras.append(f"    - Peso total: {total_peso:.2f}kg")
                conteudo_barras.append("")
        else:
            conteudo_barras.append("Nenhuma barra de aço encontrada.")
        
        conteudo_barras.append("=" * 80)
        texto_barras.insert("1.0", "\n".join(conteudo_barras))
        
        # ========== ABA 4: RESUMO EXECUTIVO ==========
        frame_resumo = tk.Frame(notebook, bg="#1e1e1e")
        notebook.add(frame_resumo, text="📊 Resumo Executivo")
        
        texto_resumo = ScrolledText(
            frame_resumo,
            bg="#1e1e1e",
            fg="#ffc107",
            font=("Courier New", 10),
            wrap="none"
        )
        texto_resumo.pack(fill="both", expand=True)
        
        conteudo_resumo = []
        conteudo_resumo.append("=" * 80)
        conteudo_resumo.append("                    RESUMO EXECUTIVO - LAJES")
        conteudo_resumo.append("=" * 80)
        conteudo_resumo.append(f"Obra:      {self.var_obra.get()}")
        conteudo_resumo.append(f"Pavimento: {self.var_pavimento.get()}")
        conteudo_resumo.append(f"Tipo Laje: {self.var_tipo_laje.get()}")
        conteudo_resumo.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        conteudo_resumo.append("=" * 80)
        conteudo_resumo.append("")
        
        # Estatísticas gerais
        conteudo_resumo.append("ESTATÍSTICAS GERAIS:")
        conteudo_resumo.append("-" * 40)
        
        # Contar lajes únicas
        lajes_unicas = set(dado[0] for dado in self.dados_processados)
        conteudo_resumo.append(f"  Total de Lajes:        {len(lajes_unicas)}")
        conteudo_resumo.append(f"  Total de Itens:        {len(self.dados_processados)}")
        conteudo_resumo.append(f"  Peso Total:            {self.total_kg:.2f} kg")
        conteudo_resumo.append("")
        
        # Resumo por tipo
        conteudo_resumo.append("RESUMO POR TIPO:")
        conteudo_resumo.append("-" * 40)
        
        resumo_tipos = {}
        for dado in self.dados_processados:
            tipo = dado[1]
            peso = dado[5]
            
            if tipo not in resumo_tipos:
                resumo_tipos[tipo] = {'peso': 0, 'qtde': 0}
            resumo_tipos[tipo]['peso'] += peso
            resumo_tipos[tipo]['qtde'] += 1
        
        for tipo in sorted(resumo_tipos.keys()):
            peso = resumo_tipos[tipo]['peso']
            qtde = resumo_tipos[tipo]['qtde']
            percentual = (peso / self.total_kg) * 100
            conteudo_resumo.append(
                f"  {tipo:<15}: {qtde:>3} itens | {peso:>8.2f} kg ({percentual:>5.1f}%)"
            )
        
        conteudo_resumo.append("")
        conteudo_resumo.append("=" * 80)
        conteudo_resumo.append("FIM DO RESUMO EXECUTIVO")
        conteudo_resumo.append("=" * 80)
        
        texto_resumo.insert("1.0", "\n".join(conteudo_resumo))
        
        # ========== ABA 5: LISTA DE COMPRAS ==========
        frame_compras = tk.Frame(notebook, bg="#1e1e1e")
        notebook.add(frame_compras, text="🛒 Lista de Compras")
        
        texto_compras = ScrolledText(
            frame_compras,
            bg="#1e1e1e",
            fg="#9c27b0",
            font=("Courier New", 10),
            wrap="none"
        )
        texto_compras.pack(fill="both", expand=True)
        
        conteudo_compras = []
        conteudo_compras.append("=" * 80)
        conteudo_compras.append("                    LISTA DE COMPRAS - LAJES")
        conteudo_compras.append("=" * 80)
        conteudo_compras.append(f"Obra:      {self.var_obra.get()}")
        conteudo_compras.append(f"Pavimento: {self.var_pavimento.get()}")
        conteudo_compras.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        conteudo_compras.append("=" * 80)
        conteudo_compras.append("")
        
        conteudo_compras.append("MATERIAL NECESSÁRIO (COM PERDAS):")
        conteudo_compras.append("-" * 40)
        conteudo_compras.append("")
        
        # Separar telas e barras
        telas_compra = {}
        barras_compra = {}
        
        for dado in self.dados_processados:
            if dado[1] == "TELA":
                tipo = dado[2]
                if tipo not in telas_compra:
                    telas_compra[tipo] = {'area': 0, 'peso': 0}
                telas_compra[tipo]['area'] += dado[4]
                telas_compra[tipo]['peso'] += dado[5]
            else:
                # Extrair bitola
                espec = dado[2]
                bitola = "0"
                if "ø" in espec:
                    try:
                        bitola = espec.split("ø")[1].split()[0]
                    except:
                        pass
                
                if bitola not in barras_compra:
                    barras_compra[bitola] = {'metros': 0, 'peso': 0, 'qtde': 0}
                barras_compra[bitola]['metros'] += dado[3] * dado[4]
                barras_compra[bitola]['peso'] += dado[5]
                barras_compra[bitola]['qtde'] += dado[3]
        
        # TELAS SOLDADAS
        if telas_compra:
            conteudo_compras.append("TELAS SOLDADAS:")
            conteudo_compras.append("")
            
            total_telas = 0
            for tipo in sorted(telas_compra.keys()):
                area = telas_compra[tipo]['area']
                area_com_perda = area * 1.15  # 15% de sobreposição
                paineis = math.ceil(area_com_perda / 14.7)  # Painéis de 2,45x6,00m
                peso_com_perda = telas_compra[tipo]['peso'] * 1.15
                
                conteudo_compras.append(f"  {tipo}:")
                conteudo_compras.append(f"    - Área sem perda:     {area:.2f} m²")
                conteudo_compras.append(f"    - Área com 15% perda: {area_com_perda:.2f} m²")
                conteudo_compras.append(f"    - Painéis (2,45x6m):  {paineis} unidades")
                conteudo_compras.append(f"    - Peso estimado:      {peso_com_perda:.2f} kg")
                conteudo_compras.append("")
                
                total_telas += peso_com_perda
        
        # BARRAS DE AÇO
        if barras_compra:
            conteudo_compras.append("BARRAS DE AÇO:")
            conteudo_compras.append("")
            
            total_barras = 0
            for bitola in sorted(barras_compra.keys(), key=lambda x: float(x) if x.replace('.','').isdigit() else 0):
                metros = barras_compra[bitola]['metros']
                metros_com_perda = metros * 1.10  # 10% de perda
                barras_12m = math.ceil(metros_com_perda / 12)
                peso_com_perda = barras_compra[bitola]['peso'] * 1.10
                
                conteudo_compras.append(f"  Ø {bitola} mm:")
                conteudo_compras.append(f"    - Metragem sem perda:  {metros:.2f} m")
                conteudo_compras.append(f"    - Metragem com 10%:    {metros_com_perda:.2f} m")
                conteudo_compras.append(f"    - Barras de 12m:       {barras_12m} barras")
                conteudo_compras.append(f"    - Peso estimado:       {peso_com_perda:.2f} kg")
                conteudo_compras.append("")
                
                total_barras += peso_com_perda
        
        # RESUMO FINAL
        conteudo_compras.append("-" * 40)
        conteudo_compras.append("RESUMO PARA PEDIDO:")
        conteudo_compras.append("")
        
        peso_total_compra = 0
        
        if telas_compra:
            conteudo_compras.append("TELAS:")
            for tipo in sorted(telas_compra.keys()):
                area_com_perda = telas_compra[tipo]['area'] * 1.15
                paineis = math.ceil(area_com_perda / 14.7)
                peso = telas_compra[tipo]['peso'] * 1.15
                conteudo_compras.append(f"  {tipo}: {paineis} painéis ({peso:.2f} kg)")
                peso_total_compra += peso
        
        if barras_compra:
            conteudo_compras.append("")
            conteudo_compras.append("BARRAS:")
            for bitola in sorted(barras_compra.keys(), key=lambda x: float(x) if x.replace('.','').isdigit() else 0):
                barras_12m = math.ceil(barras_compra[bitola]['metros'] * 1.10 / 12)
                peso = barras_compra[bitola]['peso'] * 1.10
                conteudo_compras.append(f"  Ø {bitola} mm: {barras_12m} barras de 12m ({peso:.2f} kg)")
                peso_total_compra += peso
        
        conteudo_compras.append("")
        conteudo_compras.append("-" * 40)
        conteudo_compras.append(f"PESO TOTAL PARA COMPRA: {peso_total_compra:.2f} kg")
        conteudo_compras.append(f"PESO TOTAL SEM PERDA:   {self.total_kg:.2f} kg")
        conteudo_compras.append(f"DIFERENÇA (PERDA):      {peso_total_compra - self.total_kg:.2f} kg")
        conteudo_compras.append("")
        conteudo_compras.append("=" * 80)
        conteudo_compras.append("FIM DA LISTA DE COMPRAS")
        conteudo_compras.append("=" * 80)
        
        texto_compras.insert("1.0", "\n".join(conteudo_compras))
        
        # Função para atualizar texto ativo quando mudar de aba
        def on_tab_changed(event):
            selected_tab = notebook.index(notebook.select())
            textos = [texto_geral, texto_telas, texto_barras, texto_resumo, texto_compras]
            self.texto_ativo = textos[selected_tab]
        
        notebook.bind("<<NotebookTabChanged>>", on_tab_changed)
        
        # Definir primeira aba como ativa
        self.texto_ativo = texto_geral
    
    def atualizar_botoes_navegacao(self):
        """Atualiza estado dos botões de navegação"""
        if self.pagina_atual == 0:
            self.btn_primeira.config(state="disabled")
            self.btn_anterior.config(state="disabled")
        else:
            self.btn_primeira.config(state="normal")
            self.btn_anterior.config(state="normal")
        
        if self.pagina_atual >= self.total_paginas - 1:
            self.btn_proxima.config(state="disabled")
            self.btn_ultima.config(state="disabled")
        else:
            self.btn_proxima.config(state="normal")
            self.btn_ultima.config(state="normal")
        
        self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {self.total_paginas}")
    
    def primeira_pagina(self):
        self.pagina_atual = 0
        self.desenhar_pagina_etiquetas_lajes()
        self.atualizar_botoes_navegacao()
    
    def pagina_anterior(self):
        if self.pagina_atual > 0:
            self.pagina_atual -= 1
            self.desenhar_pagina_etiquetas_lajes()
            self.atualizar_botoes_navegacao()
    
    def proxima_pagina(self):
        if self.pagina_atual < self.total_paginas - 1:
            self.pagina_atual += 1
            self.desenhar_pagina_etiquetas_lajes()
            self.atualizar_botoes_navegacao()
    
    def ultima_pagina(self):
        self.pagina_atual = self.total_paginas - 1
        self.desenhar_pagina_etiquetas_lajes()
        self.atualizar_botoes_navegacao()
    
    def exportar_excel(self):
        """Exporta dados para Excel com formatação profissional"""
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            
            # Aba 1: Dados Gerais
            ws1 = wb.active
            ws1.title = "Lajes - Geral"
            
            # Cabeçalho
            ws1.merge_cells('A1:F1')
            ws1['A1'] = 'RELATÓRIO DE LAJES - DETALHADO'
            ws1['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws1['A1'].fill = PatternFill(start_color="00ACC1", end_color="00ACC1", fill_type="solid")
            ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Informações do projeto
            ws1['A3'] = 'Obra:'
            ws1['B3'] = self.var_obra.get()
            ws1['A4'] = 'Pavimento:'
            ws1['B4'] = self.var_pavimento.get()
            ws1['A5'] = 'Tipo Laje:'
            ws1['B5'] = self.var_tipo_laje.get()
            ws1['A6'] = 'Data:'
            ws1['B6'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            ws1['D3'] = 'Total de Itens:'
            ws1['E3'] = len(self.dados_processados)
            ws1['D4'] = 'Peso Total (kg):'
            ws1['E4'] = f"{self.total_kg:.2f}"
            
            # Cabeçalhos da tabela
            headers = ['LAJE', 'TIPO', 'ESPECIFICAÇÃO', 'QTD', 'ÁREA/COMP', 'PESO (kg)']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=8, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="00ACC1", end_color="00ACC1", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Dados
            for row_idx, dado in enumerate(self.dados_processados, 9):
                for col_idx, valor in enumerate(dado, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx)
                    cell.value = valor
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Alternar cores das linhas
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
            
            # Total
            row_total = len(self.dados_processados) + 10
            ws1.merge_cells(f'A{row_total}:E{row_total}')
            ws1[f'A{row_total}'] = 'TOTAL GERAL'
            ws1[f'A{row_total}'].font = Font(bold=True, size=12)
            ws1[f'A{row_total}'].alignment = Alignment(horizontal='right')
            ws1[f'A{row_total}'].fill = PatternFill(start_color="B2EBF2", end_color="B2EBF2", fill_type="solid")
            
            ws1[f'F{row_total}'] = f"{self.total_kg:.2f}"
            ws1[f'F{row_total}'].font = Font(bold=True)
            ws1[f'F{row_total}'].fill = PatternFill(start_color="B2EBF2", end_color="B2EBF2", fill_type="solid")
            
            # Aba 2: Telas Soldadas
            ws2 = wb.create_sheet("Telas Soldadas")
            
            ws2['A1'] = 'RESUMO DE TELAS SOLDADAS'
            ws2['A1'].font = Font(size=14, bold=True)
            
            # Filtrar telas
            telas = [d for d in self.dados_processados if d[1] == "TELA"]
            
            if telas:
                # Cabeçalhos
                headers_tela = ['LAJE', 'TIPO TELA', 'ÁREA (m²)', 'PESO (kg)']
                for col, header in enumerate(headers_tela, 1):
                    cell = ws2.cell(row=3, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Dados das telas
                row = 4
                for tela in telas:
                    ws2[f'A{row}'] = tela[0]
                    ws2[f'B{row}'] = tela[2]
                    ws2[f'C{row}'] = tela[4]
                    ws2[f'D{row}'] = tela[5]
                    row += 1
                
                # Resumo por tipo
                row += 2
                ws2[f'A{row}'] = 'RESUMO POR TIPO:'
                ws2[f'A{row}'].font = Font(bold=True)
                row += 1
                
                telas_tipo = {}
                for tela in telas:
                    tipo = tela[2]
                    if tipo not in telas_tipo:
                        telas_tipo[tipo] = {'area': 0, 'peso': 0}
                    telas_tipo[tipo]['area'] += tela[4]
                    telas_tipo[tipo]['peso'] += tela[5]
                
                for tipo in sorted(telas_tipo.keys()):
                    ws2[f'A{row}'] = tipo
                    ws2[f'B{row}'] = f"{telas_tipo[tipo]['area']:.2f} m²"
                    ws2[f'C{row}'] = f"{math.ceil(telas_tipo[tipo]['area']*1.15/14.7)} painéis"
                    ws2[f'D{row}'] = f"{telas_tipo[tipo]['peso']:.2f} kg"
                    row += 1
            
            # Aba 3: Barras de Aço
            ws3 = wb.create_sheet("Barras de Aço")
            
            ws3['A1'] = 'RESUMO DE BARRAS DE AÇO'
            ws3['A1'].font = Font(size=14, bold=True)
            
            # Filtrar barras
            barras = [d for d in self.dados_processados if d[1] != "TELA"]
            
            if barras:
                # Cabeçalhos
                headers_barra = ['LAJE', 'TIPO', 'ESPECIFICAÇÃO', 'QTD', 'COMP (m)', 'PESO (kg)']
                for col, header in enumerate(headers_barra, 1):
                    cell = ws3.cell(row=3, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Dados das barras
                row = 4
                for barra in barras:
                    ws3[f'A{row}'] = barra[0]
                    ws3[f'B{row}'] = barra[1]
                    ws3[f'C{row}'] = barra[2]
                    ws3[f'D{row}'] = barra[3]
                    ws3[f'E{row}'] = barra[4]
                    ws3[f'F{row}'] = barra[5]
                    row += 1
            
            # Aba 4: Lista de Compras
            ws4 = wb.create_sheet("Lista de Compras")
            
            ws4['A1'] = 'LISTA DE COMPRAS - LAJES'
            ws4['A1'].font = Font(size=14, bold=True)
            
            row = 3
            ws4[f'A{row}'] = 'TELAS SOLDADAS (com 15% sobreposição):'
            ws4[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Calcular telas necessárias
            telas_compra = {}
            for dado in self.dados_processados:
                if dado[1] == "TELA":
                    tipo = dado[2]
                    if tipo not in telas_compra:
                        telas_compra[tipo] = 0
                    telas_compra[tipo] += dado[4]
            
            for tipo in sorted(telas_compra.keys()):
                area_com_perda = telas_compra[tipo] * 1.15
                paineis = math.ceil(area_com_perda / 14.7)
                ws4[f'A{row}'] = tipo
                ws4[f'B{row}'] = f"{paineis} painéis (2,45x6,00m)"
                ws4[f'C{row}'] = f"{area_com_perda:.2f} m²"
                row += 1
            
            row += 1
            ws4[f'A{row}'] = 'BARRAS DE AÇO (com 10% perda):'
            ws4[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Calcular barras necessárias
            barras_compra = {}
            for dado in self.dados_processados:
                if dado[1] != "TELA":
                    espec = dado[2]
                    bitola = "0"
                    if "ø" in espec:
                        try:
                            bitola = espec.split("ø")[1].split()[0]
                        except:
                            pass
                    
                    if bitola not in barras_compra:
                        barras_compra[bitola] = 0
                    barras_compra[bitola] += dado[3] * dado[4]
            
            for bitola in sorted(barras_compra.keys(), key=lambda x: float(x) if x.replace('.','').isdigit() else 0):
                metros_com_perda = barras_compra[bitola] * 1.10
                barras_12m = math.ceil(metros_com_perda / 12)
                ws4[f'A{row}'] = f"Ø {bitola} mm"
                ws4[f'B{row}'] = f"{barras_12m} barras de 12m"
                ws4[f'C{row}'] = f"{metros_com_perda:.2f} m"
                row += 1
            
            # Ajustar largura das colunas em todas as abas
            for ws in [ws1, ws2, ws3, ws4]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Salvar arquivo
            arquivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=f"lajes_{self.var_pavimento.get()}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )
            
            if arquivo:
                wb.save(arquivo)
                messagebox.showinfo("Sucesso", f"Arquivo Excel salvo com sucesso!\n{arquivo}")
                
                if messagebox.askyesno("Abrir arquivo", "Deseja abrir o arquivo Excel agora?"):
                    os.startfile(arquivo)
        
        except ImportError:
            messagebox.showerror("Erro", "Biblioteca openpyxl não instalada!\n\nExecute: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar Excel:\n{str(e)}")
    
    def imprimir_direto(self):
        """Imprime relatório diretamente"""
        if not self.dados_processados:
            messagebox.showwarning("Atenção", "Processe os arquivos primeiro!")
            return
        
        try:
            import tempfile
            
            conteudo = []
            conteudo.append("=" * 70)
            conteudo.append("                    RELATÓRIO DE LAJES")
            conteudo.append("=" * 70)
            conteudo.append(f"Obra:      {self.var_obra.get()}")
            conteudo.append(f"Pavimento: {self.var_pavimento.get()}")
            conteudo.append(f"Tipo:      {self.var_tipo_laje.get()}")
            conteudo.append(f"Data:      {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            conteudo.append("=" * 70)
            conteudo.append("")
            
            laje_atual = None
            for dado in self.dados_processados:
                if dado[0] != laje_atual:
                    laje_atual = dado[0]
                    conteudo.append("")
                    conteudo.append(f">>> {laje_atual}")
                    conteudo.append("-" * 50)
                
                tipo = dado[1]
                espec = dado[2]
                qtde = dado[3]
                area_comp = dado[4]
                peso = dado[5]
                
                if tipo == "TELA":
                    conteudo.append(
                        f"    {tipo:<10} {espec:<15} Área: {area_comp:>6.2f}m²  Peso: {peso:>8.2f}kg"
                    )
                else:
                    conteudo.append(
                        f"    {tipo:<10} {espec:<15} Qtd: {qtde:>3}  Comp: {area_comp:>5.2f}m  Peso: {peso:>8.2f}kg"
                    )
            
            conteudo.append("")
            conteudo.append("=" * 70)
            conteudo.append(f"TOTAL GERAL:")
            conteudo.append(f"  Total de itens: {len(self.dados_processados)}")
            conteudo.append(f"  Peso total: {self.total_kg:.2f} kg")
            conteudo.append("=" * 70)
            
            texto_final = "\n".join(conteudo)
            
            # Criar arquivo temporário
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.txt', encoding='utf-8') as f:
                f.write(texto_final)
                temp_file = f.name
            
            # Abrir para impressão
            if os.name == 'nt':  # Windows
                os.startfile(temp_file, "print")
                messagebox.showinfo("Impressão", "Documento enviado para impressora!")
            else:
                messagebox.showinfo("Impressão", f"Arquivo criado em:\n{temp_file}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao imprimir:\n{str(e)}")
    
    def limpar(self):
        """Limpa todos os dados"""
        self.arquivos_selecionados = []
        self.dados_processados = []
        self.total_kg = 0.0
        self.total_barras = 0
        self.tipos_personalizados = {}
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.status_label.config(text="✅ Pronto para processar")
        self.info_label.config(text="")
    
    def imprimir_com_preview(self, conteudo):
        """Imprime com preview"""
        janela_preview = tk.Toplevel(self)
        janela_preview.title("Preview de Impressão")
        janela_preview.geometry("800x600")
        
        frame_botoes = tk.Frame(janela_preview, bg="#2c3e50")
        frame_botoes.pack(fill="x", padx=5, pady=5)
        
        tk.Label(
            frame_botoes,
            text="📄 PREVIEW DO DOCUMENTO",
            bg="#2c3e50",
            fg="white",
            font=("Arial", 12, "bold")
        ).pack(side="left", padx=10, pady=5)
        
        tk.Button(
            frame_botoes,
            text="🖨️ Confirmar Impressão",
            command=lambda: self.confirmar_impressao(conteudo, janela_preview),
            bg="#27ae60",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side="right", padx=5, pady=5)
        
        tk.Button(
            frame_botoes,
            text="❌ Cancelar",
            command=janela_preview.destroy,
            bg="#e74c3c",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side="right", padx=5, pady=5)
        
        texto_preview = ScrolledText(
            janela_preview,
            font=("Courier New", 10),
            bg="white",
            fg="black",
            wrap="word"
        )
        texto_preview.pack(fill="both", expand=True, padx=5, pady=5)
        texto_preview.insert("1.0", conteudo)
        texto_preview.config(state="disabled")
    
    def confirmar_impressao(self, conteudo, janela_preview):
        """Confirma e envia para impressão"""
        try:
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.txt', encoding='utf-8') as f:
                f.write(conteudo)
                temp_file = f.name
            
            if os.name == 'nt':
                os.startfile(temp_file, "print")
                messagebox.showinfo("Impressão", "Documento enviado para impressora!")
            else:
                messagebox.showinfo("Impressão", f"Arquivo criado em:\n{temp_file}")
            
            janela_preview.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao imprimir:\n{str(e)}")
    
    def salvar_txt(self, conteudo):
        """Salva conteúdo em arquivo texto"""
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Texto", "*.txt")],
            initialfile=f"romaneio_lajes_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
        )
        if arquivo:
            with open(arquivo, "w", encoding="utf-8") as f:
                f.write(conteudo)
            messagebox.showinfo("Sucesso", "Arquivo salvo com sucesso!")

# Executar aplicação
if __name__ == "__main__":
    app = LajesApp()
    app.mainloop()